#!/usr/bin/env python3
"""
Auto-Sync Pipeline - COMPLETE ZERO-TOUCH VERSION

Runs the entire pipeline automatically with no user input required.
"""

import json
import logging
import os
import subprocess
import sys
import time
from pathlib import Path
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AutoSyncPipeline:
    """Completely automated pipeline."""
    
    def __init__(self):
        self.script_dir = Path(__file__).parent.absolute()
        self.config = self._load_config()
        
    def _load_config(self) -> dict:
        """Load configuration."""
        config_path = self.script_dir / "company_config.json"
        if config_path.exists():
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {"company_name": "Company", "location": "Location"}
    
    def consolidate_all_data(self) -> list:
        """Automatically consolidate all employee data."""
        print("🔄 Auto-consolidating employee data...")
        
        all_employees = []
        processed_files = []
        
        potential_files = [
            "linkedin_candidates.json",
            "merged_employees.json",
            "verified_employee_data.json",
            "website_employees.json",
            "temp_employee_data.json"
        ]
        
        existing_ids = set()
        
        for filename in potential_files:
            file_path = self.script_dir / filename
            if file_path.exists():
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    if isinstance(data, list):
                        for item in data:
                            first_name = item.get('first_name', '').lower().strip()
                            last_name = item.get('last_name', '').lower().strip()
                            emp_id = f"{first_name}_{last_name}"
                            
                            if emp_id and emp_id not in existing_ids and first_name and last_name:
                                standardized = {
                                    'first_name': item.get('first_name', '').strip().title(),
                                    'last_name': item.get('last_name', '').strip().title(),
                                    'title': item.get('title', ''),
                                    'company_name': item.get('company_name', self.config.get('company_name', '')),
                                    'location': item.get('location', self.config.get('location', '')),
                                    'source': item.get('source', filename.replace('.json', '')),
                                    'confidence': item.get('confidence', 'medium'),
                                    'link': (item.get('link') or item.get('linkedin_url') or item.get('source_link', '')),
                                    'last_updated': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                }
                                
                                all_employees.append(standardized)
                                existing_ids.add(emp_id)
                        
                        processed_files.append(filename)
                        
                except:
                    continue
        
        if all_employees:
            # Auto-save consolidated data
            consolidated_path = self.script_dir / "auto_consolidated_employees.json"
            with open(consolidated_path, 'w', encoding='utf-8') as f:
                json.dump(all_employees, f, indent=4)
            
            print(f"✅ Auto-consolidated {len(all_employees)} employees from {len(processed_files)} files")
        
        return all_employees
    
    def auto_generate_excel(self, employees: list) -> bool:
        """Automatically generate Excel report."""
        try:
            # Try to run existing Excel script first
            excel_script = self.script_dir / "script4_excel_output.py"
            if excel_script.exists():
                result = subprocess.run([sys.executable, str(excel_script)], 
                                      capture_output=True, timeout=120)
                if result.returncode == 0:
                    print("📊 Excel report generated automatically")
                    return True
            
            # Fallback: create simple Excel inline
            try:
                subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Auto-Generated Employees"
                
                # Headers
                headers = ["First Name", "Last Name", "Title", "Company", "Location", "Source", "Link"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Data
                for row, emp in enumerate(employees, 2):
                    ws.cell(row=row, column=1, value=emp.get('first_name', ''))
                    ws.cell(row=row, column=2, value=emp.get('last_name', ''))
                    ws.cell(row=row, column=3, value=emp.get('title', ''))
                    ws.cell(row=row, column=4, value=emp.get('company_name', ''))
                    ws.cell(row=row, column=5, value=emp.get('location', ''))
                    ws.cell(row=row, column=6, value=emp.get('source', ''))
                    ws.cell(row=row, column=7, value=emp.get('link', ''))
                
                # Save
                company_name = self.config.get('company_name', 'Company').replace(' ', '_')
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{company_name}_Auto_Report_{timestamp}.xlsx"
                filepath = self.script_dir / filename
                
                wb.save(filepath)
                print(f"📊 Auto-generated Excel report: {filename}")
                
                # Try to open it
                try:
                    if sys.platform == 'win32':
                        os.startfile(str(filepath))
                    elif sys.platform == 'darwin':
                        subprocess.call(['open', str(filepath)])
                    else:
                        subprocess.call(['xdg-open', str(filepath)])
                except:
                    pass
                
                return True
                
            except Exception as e:
                print(f"⚠️ Excel generation failed: {e}")
                return False
                
        except Exception as e:
            print(f"⚠️ Auto-Excel error: {e}")
            return False
    
    def run_complete_pipeline(self):
        """Run the complete pipeline automatically."""
        start_time = time.time()
        
        print("🚀 AUTO-SYNC PIPELINE - ZERO-TOUCH MODE")
        print("=" * 50)
        
        # Step 1: Consolidate data
        employees = self.consolidate_all_data()
        
        if not employees:
            print("⚠️ No employee data found")
            return
        
        # Step 2: Auto-verify LinkedIn (if script5 exists)
        script5_path = self.script_dir / "script5_linkedin_verification.py"
        if script5_path.exists():
            try:
                subprocess.run([sys.executable, str(script5_path)], 
                             capture_output=True, timeout=180)
                print("🔗 LinkedIn auto-verification attempted")
            except:
                print("⚠️ LinkedIn verification skipped")
        
        # Step 3: Generate Excel
        self.auto_generate_excel(employees)
        
        # Step 4: Auto-save everything
        final_data_path = self.script_dir / "final_employee_data.json"
        with open(final_data_path, 'w', encoding='utf-8') as f:
            json.dump(employees, f, indent=4)
        
        elapsed = time.time() - start_time
        print(f"✅ Complete pipeline finished in {elapsed:.1f} seconds")
        print(f"📊 {len(employees)} employees processed")

def main():
    """Main function - completely automated."""
    pipeline = AutoSyncPipeline()
    pipeline.run_complete_pipeline()

if __name__ == "__main__":
    main()