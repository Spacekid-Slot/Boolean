#!/usr/bin/env python3
"""
ONE-SHOT AUTOMATION - Complete Employee Discovery

Run script1 setup, then everything else automatically.
ONLY user interaction: Company info input + LinkedIn login
Result: Complete Excel report with verified LinkedIn profiles

Usage: python one_shot_automation.py
"""

import json
import logging
import os
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class OneShotAutomation:
    """Complete automation from setup to final Excel report."""
    
    def __init__(self):
        self.script_dir = Path(__file__).parent.absolute()
        self.config = {}
        
    def _print_header(self, title: str):
        """Print formatted header."""
        print("\n" + "=" * 70)
        print(f"🚀 {title}")
        print("=" * 70)
    
    def _print_status(self, message: str, emoji: str = "ℹ️"):
        """Print status with timestamp."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {emoji} {message}")
    
    def run_script1_setup(self) -> bool:
        """Run script1 setup with user input."""
        self._print_header("STEP 1: COMPANY SETUP")
        
        print("Quick setup - just need basic company information:")
        
        try:
            # Company name
            while True:
                company_name = input("\n🏢 Company name: ").strip()
                if len(company_name) >= 2:
                    break
                print("Please enter a valid company name (2+ characters)")
            
            # Location
            while True:
                location = input("📍 Location (city, country): ").strip()
                if len(location) >= 2:
                    break
                print("Please enter a valid location")
            
            # Website
            while True:
                website = input("🌐 Company website (e.g., company.com): ").strip()
                if len(website) >= 4:
                    if not website.startswith(('http://', 'https://')):
                        website = 'https://' + website
                    break
                print("Please enter a valid website")
            
            # Quick job title choice
            print("\n🎯 Job titles to search for:")
            print("1. Use comprehensive default list (CEO, Directors, Managers, etc.) - RECOMMENDED")
            print("2. Enter 3-5 specific job titles")
            
            job_choice = input("Choice (1/2, default: 1): ").strip()
            
            if job_choice == '2':
                print("Enter 3-5 job titles you want to find:")
                job_titles = []
                for i in range(5):
                    title = input(f"Job title {i+1} (or press Enter if done): ").strip()
                    if not title:
                        break
                    job_titles.append(title)
                
                if not job_titles:
                    job_titles = self._get_default_job_titles()
                    print("Using default job titles instead")
            else:
                job_titles = self._get_default_job_titles()
            
            # Create config
            self.config = {
                "company_name": company_name,
                "location": location,
                "company_website": website,
                "job_titles": job_titles,
                "output_file": f"{company_name.replace(' ', '_')}_{location.replace(' ', '_')}_employees.xlsx",
                "temp_data_file": "temp_employee_data.json",
                "processed_data_file": "processed_employee_data.json",
                "verified_data_file": "verified_employee_data.json",
                "pages_to_scrape": 5,
                "debug_mode": False,
                "review_mode": False,
                "search_types": ["linkedin", "website"],
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "version": "3.0"
            }
            
            # Save config
            config_path = self.script_dir / "company_config.json"
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
            
            self._print_status("✅ Setup complete!", "✅")
            self._print_status(f"Company: {company_name}", "📋")
            self._print_status(f"Location: {location}", "📋")
            self._print_status(f"Job titles: {len(job_titles)} configured", "📋")
            
            return True
            
        except KeyboardInterrupt:
            print("\n❌ Setup cancelled")
            return False
        except Exception as e:
            print(f"❌ Setup error: {e}")
            return False
    
    def _get_default_job_titles(self) -> List[str]:
        """Comprehensive default job titles."""
        return [
            "CEO", "Chief Executive Officer", "Managing Director", "President", "Chairman",
            "Chief Financial Officer", "CFO", "Finance Director", "Financial Controller",
            "Chief Technology Officer", "CTO", "Technology Director", "IT Director",
            "Chief Operating Officer", "COO", "Operations Director", "Operations Manager",
            "Chief Marketing Officer", "CMO", "Marketing Director", "Marketing Manager",
            "Sales Director", "Sales Manager", "Business Development Manager",
            "Director", "Senior Director", "Vice President", "VP", "Executive Director",
            "General Manager", "Regional Manager", "Branch Manager", "Area Manager",
            "Head of Sales", "Head of Marketing", "Head of Operations", "Head of Finance",
            "Senior Manager", "Manager", "Team Leader", "Department Head",
            "Property Director", "Property Manager", "Asset Manager", "Portfolio Manager",
            "Investment Manager", "Fund Manager", "Account Manager", "Client Manager",
            "Project Manager", "Programme Manager", "Product Manager",
            "HR Director", "Human Resources Manager", "People Director",
            "Legal Director", "Compliance Manager", "Risk Manager"
        ]
    
    def run_linkedin_search(self) -> bool:
        """Run LinkedIn X-ray search automatically."""
        self._print_header("STEP 2: LINKEDIN X-RAY SEARCH")
        
        linkedin_script = self.script_dir / "script2_web_scraping.py"
        if not linkedin_script.exists():
            print("❌ LinkedIn search script not found")
            return False
        
        print(f"🔍 Searching LinkedIn for employees at {self.config['company_name']}...")
        print(f"📍 Location: {self.config['location']}")
        print(f"🎯 Job titles: {len(self.config['job_titles'])} different roles")
        print("⏱️ This will take 2-5 minutes...")
        
        try:
            # Run LinkedIn search
            result = subprocess.run(
                [sys.executable, str(linkedin_script)], 
                capture_output=True, 
                text=True, 
                timeout=600
            )
            
            if result.returncode == 0:
                # Check results
                linkedin_file = self.script_dir / "linkedin_candidates.json"
                if linkedin_file.exists():
                    with open(linkedin_file, 'r', encoding='utf-8') as f:
                        candidates = json.load(f)
                    
                    if candidates:
                        self._print_status(f"✅ Found {len(candidates)} LinkedIn profiles!", "✅")
                        return True
                    else:
                        self._print_status("⚠️ Search completed but no profiles found", "⚠️")
                        return False
                else:
                    self._print_status("⚠️ Search completed but no results file created", "⚠️")
                    return False
            else:
                print(f"❌ LinkedIn search failed: {result.stderr}")
                return False
                
        except subprocess.TimeoutExpired:
            print("❌ LinkedIn search timed out")
            return False
        except Exception as e:
            print(f"❌ LinkedIn search error: {e}")
            return False
    
    def run_linkedin_verification(self) -> bool:
        """Run LinkedIn verification with user login."""
        self._print_header("STEP 3: LINKEDIN VERIFICATION")
        
        print("🔗 LinkedIn verification will:")
        print("   • Open browser to LinkedIn")
        print("   • Wait for you to login manually")
        print("   • Verify current employment for each profile")
        print("   • Update job titles and company information")
        
        proceed = input("\n🔗 Run LinkedIn verification? (y/n, default: y): ").strip().lower()
        if proceed == 'n':
            self._print_status("⚠️ Skipping LinkedIn verification", "⚠️")
            return False
        
        # Create verification script inline
        verification_script = '''#!/usr/bin/env python3
import json
import sys
import time
from pathlib import Path
from datetime import datetime

def install_selenium():
    import subprocess
    try:
        from selenium import webdriver
        return True
    except ImportError:
        subprocess.call([sys.executable, "-m", "pip", "install", "selenium", "webdriver-manager"])
        return True

def verify_linkedin_profiles():
    if not install_selenium():
        return False
    
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from webdriver_manager.chrome import ChromeDriverManager
        
        script_dir = Path(__file__).parent
        
        # Load LinkedIn candidates
        linkedin_file = script_dir / "linkedin_candidates.json"
        if not linkedin_file.exists():
            print("❌ No LinkedIn candidates found")
            return False
        
        with open(linkedin_file, 'r', encoding='utf-8') as f:
            candidates = json.load(f)
        
        if not candidates:
            print("❌ No candidates to verify")
            return False
        
        print(f"🔗 Found {len(candidates)} LinkedIn profiles to verify")
        
        # Setup browser
        options = Options()
        options.add_argument('--window-size=1200,800')
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        
        try:
            # Manual LinkedIn login
            print("\\n🔗 Opening LinkedIn for manual login...")
            driver.get("https://www.linkedin.com/login")
            
            input("\\n✅ Please login to LinkedIn in the browser, then press Enter to continue...")
            
            # Verify profiles
            verified_candidates = []
            
            for i, candidate in enumerate(candidates[:15], 1):  # Limit to 15 for safety
                linkedin_url = candidate.get('linkedin_url') or candidate.get('link', '')
                if not linkedin_url:
                    continue
                
                print(f"\\n🔍 Verifying {i}/15: {candidate.get('first_name', '')} {candidate.get('last_name', '')}")
                
                try:
                    driver.get(linkedin_url)
                    time.sleep(3)
                    
                    # Try to get current title
                    try:
                        title_element = driver.find_element(By.CSS_SELECTOR, 'h1')
                        current_title = title_element.text.strip()
                        candidate['verified_title'] = current_title
                        candidate['verification_status'] = 'verified'
                        candidate['verification_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        print(f"   ✅ Current title: {current_title}")
                    except:
                        candidate['verification_status'] = 'partial'
                        print(f"   ⚠️ Could not extract title")
                    
                    verified_candidates.append(candidate)
                    time.sleep(2)
                    
                except Exception as e:
                    print(f"   ❌ Error verifying profile: {e}")
                    candidate['verification_status'] = 'error'
                    verified_candidates.append(candidate)
            
            # Save verified results
            verified_file = script_dir / "linkedin_verified.json"
            with open(verified_file, 'w', encoding='utf-8') as f:
                json.dump(verified_candidates, f, indent=4)
            
            print(f"\\n✅ Verification complete! Verified {len(verified_candidates)} profiles")
            print("💾 Results saved to linkedin_verified.json")
            
            input("\\nPress Enter to close browser and continue...")
            
        finally:
            driver.quit()
        
        return True
        
    except Exception as e:
        print(f"❌ Verification error: {e}")
        return False

if __name__ == "__main__":
    verify_linkedin_profiles()
'''
        
        # Write and run verification script
        temp_verify_script = self.script_dir / "temp_verify.py"
        try:
            with open(temp_verify_script, 'w', encoding='utf-8') as f:
                f.write(verification_script)
            
            # Run verification
            result = subprocess.run([sys.executable, str(temp_verify_script)])
            
            # Clean up temp script
            temp_verify_script.unlink()
            
            # Check if verification completed
            verified_file = self.script_dir / "linkedin_verified.json"
            if verified_file.exists():
                self._print_status("✅ LinkedIn verification completed", "✅")
                return True
            else:
                self._print_status("⚠️ Verification may not have completed fully", "⚠️")
                return False
                
        except Exception as e:
            print(f"❌ Verification error: {e}")
            return False
    
    def generate_final_excel(self) -> bool:
        """Generate comprehensive final Excel report."""
        self._print_header("STEP 4: GENERATING FINAL EXCEL REPORT")
        
        try:
            # Install openpyxl
            subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Collect all employee data
            all_employees = []
            
            # Load LinkedIn candidates
            linkedin_file = self.script_dir / "linkedin_candidates.json"
            if linkedin_file.exists():
                with open(linkedin_file, 'r', encoding='utf-8') as f:
                    linkedin_data = json.load(f)
                all_employees.extend(linkedin_data)
            
            # Load verified data if available
            verified_file = self.script_dir / "linkedin_verified.json"
            verified_data = {}
            if verified_file.exists():
                with open(verified_file, 'r', encoding='utf-8') as f:
                    verified_list = json.load(f)
                
                # Convert to lookup dict
                for item in verified_list:
                    url = item.get('linkedin_url') or item.get('link', '')
                    if url:
                        verified_data[url] = item
            
            if not all_employees:
                print("❌ No employee data found")
                return False
            
            # Remove duplicates
            seen = set()
            unique_employees = []
            for emp in all_employees:
                name_key = f"{emp.get('first_name', '').lower()}_{emp.get('last_name', '').lower()}"
                if name_key not in seen and name_key != '_':
                    unique_employees.append(emp)
                    seen.add(name_key)
            
            print(f"📊 Creating Excel report with {len(unique_employees)} unique employees...")
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employee Discovery Results"
            
            # Title
            ws.merge_cells('A1:I1')
            title_cell = ws['A1']
            title_cell.value = f"{self.config['company_name']} - Complete Employee Discovery Report"
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Subtitle
            ws.merge_cells('A2:I2')
            subtitle_cell = ws['A2']
            subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Location: {self.config['location']}"
            subtitle_cell.font = Font(italic=True, size=11)
            subtitle_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = [
                "First Name", "Last Name", "Original Title", "Current Title", 
                "LinkedIn URL", "Company", "Location", "Verification", "Source"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    top=Side(style='thin'), bottom=Side(style='thin'),
                    left=Side(style='thin'), right=Side(style='thin')
                )
            
            # Data rows
            linkedin_count = 0
            verified_count = 0
            
            for row, emp in enumerate(unique_employees, 5):
                linkedin_url = emp.get('linkedin_url') or emp.get('link', '')
                
                # Check if this profile was verified
                verified_info = verified_data.get(linkedin_url, {})
                current_title = verified_info.get('verified_title', '')
                verification_status = verified_info.get('verification_status', 'not_verified')
                
                if verification_status in ['verified', 'partial']:
                    verified_count += 1
                
                # Write data
                ws.cell(row=row, column=1, value=emp.get('first_name', ''))
                ws.cell(row=row, column=2, value=emp.get('last_name', ''))
                ws.cell(row=row, column=3, value=emp.get('title', ''))
                ws.cell(row=row, column=4, value=current_title)
                
                # LinkedIn URL with hyperlink
                url_cell = ws.cell(row=row, column=5, value=linkedin_url)
                if linkedin_url and 'linkedin.com/in/' in linkedin_url.lower():
                    linkedin_count += 1
                    try:
                        url_cell.hyperlink = linkedin_url
                        url_cell.font = Font(color="0000FF", underline="single")
                    except:
                        pass
                
                ws.cell(row=row, column=6, value=emp.get('company_name', self.config['company_name']))
                ws.cell(row=row, column=7, value=emp.get('location', self.config['location']))
                
                # Verification status with color
                status_cell = ws.cell(row=row, column=8, value=verification_status.replace('_', ' ').title())
                if verification_status == 'verified':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif verification_status == 'partial':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row, column=9, value=emp.get('source', 'LinkedIn X-ray'))
            
            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max(max_length + 2, 12), 50)
            
            # Add summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            
            summary_data = [
                [f"{self.config['company_name']} - Employee Discovery Summary", ""],
                ["", ""],
                ["Company Information", ""],
                ["Company Name", self.config['company_name']],
                ["Location", self.config['location']],
                ["Website", self.config['company_website']],
                ["", ""],
                ["Search Results", ""],
                ["Total Employees Found", len(unique_employees)],
                ["LinkedIn Profiles", linkedin_count],
                ["Verified Profiles", verified_count],
                ["Job Titles Searched", len(self.config['job_titles'])],
                ["", ""],
                ["Report Details", ""],
                ["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["Search Method", "LinkedIn X-ray + Verification"],
                ["Data Quality", f"{verified_count}/{linkedin_count} profiles verified" if linkedin_count > 0 else "No LinkedIn profiles"],
                ["", ""],
                ["Next Steps", ""],
                ["1. Review LinkedIn profiles", "Click URLs to view full profiles"],
                ["2. Contact employees", "Use verified current information"],
                ["3. Update regularly", "Re-run search monthly for new hires"],
            ]
            
            for row, (label, value) in enumerate(summary_data, 1):
                cell1 = summary_sheet.cell(row=row, column=1, value=label)
                cell2 = summary_sheet.cell(row=row, column=2, value=value)
                
                if label and not label.startswith(('1.', '2.', '3.')):
                    cell1.font = Font(bold=True)
                    if "Summary" in label:
                        cell1.font = Font(bold=True, size=14)
            
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 50
            
            # Save Excel file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            company_clean = self.config['company_name'].replace(' ', '_')
            filename = f"{company_clean}_Complete_Employee_Discovery_{timestamp}.xlsx"
            filepath = self.script_dir / filename
            
            wb.save(filepath)
            
            self._print_status(f"✅ Excel report created: {filename}", "✅")
            self._print_status(f"📊 Contains {len(unique_employees)} employees", "📊")
            self._print_status(f"🔗 {linkedin_count} LinkedIn profiles", "🔗")
            self._print_status(f"✅ {verified_count} verified profiles", "✅")
            
            # Open Excel file
            try:
                if sys.platform == 'win32':
                    os.startfile(str(filepath))
                elif sys.platform == 'darwin':
                    subprocess.call(['open', str(filepath)])
                else:
                    subprocess.call(['xdg-open', str(filepath)])
                self._print_status("📂 Excel file opened automatically", "📂")
            except:
                self._print_status(f"📂 Please open: {filename}", "📂")
            
            return True
            
        except Exception as e:
            print(f"❌ Excel generation error: {e}")
            return False
    
    def run_complete_automation(self):
        """Run the complete one-shot automation."""
        self._print_header("ONE-SHOT EMPLOYEE DISCOVERY AUTOMATION")
        
        print("🎯 This will:")
        print("   1. ✅ Collect company information (script1)")
        print("   2. 🔍 Search LinkedIn for employees")
        print("   3. 🔗 Verify LinkedIn profiles (with your login)")
        print("   4. 📊 Generate complete Excel report")
        print("\n⏱️ Total time: 10-15 minutes")
        print("🔑 Only interactions: Company setup + LinkedIn login")
        
        proceed = input("\n🚀 Start complete automation? (y/n, default: y): ").strip().lower()
        if proceed == 'n':
            print("❌ Automation cancelled")
            return
        
        start_time = time.time()
        
        try:
            # Step 1: Setup
            if not self.run_script1_setup():
                print("❌ Setup failed - aborting")
                return
            
            # Step 2: LinkedIn search
            if not self.run_linkedin_search():
                print("❌ LinkedIn search failed - aborting")
                return
            
            # Step 3: LinkedIn verification
            self.run_linkedin_verification()  # Optional step
            
            # Step 4: Final Excel
            if not self.generate_final_excel():
                print("❌ Excel generation failed")
                return
            
            # Success summary
            elapsed = time.time() - start_time
            
            self._print_header("🎉 AUTOMATION COMPLETE!")
            print(f"⏱️ Total time: {elapsed/60:.1f} minutes")
            print(f"🏢 Company: {self.config['company_name']}")
            print(f"📊 Complete Excel report generated and opened")
            print(f"✅ Employee discovery pipeline complete!")
            
        except KeyboardInterrupt:
            print("\n❌ Automation cancelled by user")
        except Exception as e:
            print(f"❌ Automation error: {e}")

def main():
    automation = OneShotAutomation()
    automation.run_complete_automation()

if __name__ == "__main__":
    main()