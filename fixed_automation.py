#!/usr/bin/env python3
"""
FIXED AUTOMATION - Windows Compatible

Fixed version that:
1. Handles Windows console encoding issues
2. Fixes Excel merge cell errors
3. Shows search method menu (LinkedIn, Website, etc.)
4. Only requires company setup + LinkedIn login
"""

import json
import logging
import os
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict
from datetime import datetime

# Set UTF-8 encoding for Windows console
if sys.platform == 'win32':
    try:
        # Try to set console to UTF-8
        os.system('chcp 65001 >nul 2>&1')
    except:
        pass

# Configure logging without unicode
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FixedAutomation:
    """Fixed automation that works on Windows and shows search menu."""
    
    def __init__(self):
        self.script_dir = Path(__file__).parent.absolute()
        self.config = {}
        
    def _print_header(self, title: str):
        """Print header without emojis."""
        print("\n" + "=" * 70)
        print(f">>> {title}")
        print("=" * 70)
    
    def _print_status(self, message: str, prefix: str = "INFO"):
        """Print status without emojis."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {prefix}: {message}")
    
    def run_script1_setup(self) -> bool:
        """Run company setup."""
        self._print_header("STEP 1: COMPANY SETUP")
        
        print("Quick setup - enter basic company information:")
        
        try:
            # Company name
            while True:
                company_name = input("\nCompany name: ").strip()
                if len(company_name) >= 2:
                    break
                print("Please enter a valid company name")
            
            # Location  
            while True:
                location = input("Location (city, country): ").strip()
                if len(location) >= 2:
                    break
                print("Please enter a valid location")
            
            # Website
            while True:
                website = input("Company website (e.g., company.com): ").strip()
                if len(website) >= 4:
                    if not website.startswith(('http://', 'https://')):
                        website = 'https://' + website
                    break
                print("Please enter a valid website")
            
            # Job titles
            print("\nJob titles to search for:")
            print("1. Use comprehensive default list (CEO, Directors, Managers, etc.)")
            print("2. Enter specific job titles")
            
            job_choice = input("Choice (1/2, default: 1): ").strip()
            
            if job_choice == '2':
                print("Enter job titles (one per line, empty line to finish):")
                job_titles = []
                while True:
                    title = input("Job title: ").strip()
                    if not title:
                        break
                    job_titles.append(title)
                    print(f"Added: {title}")
                
                if not job_titles:
                    job_titles = self._get_default_job_titles()
                    print("No titles entered, using defaults")
            else:
                job_titles = self._get_default_job_titles()
            
            # Create config
            self.config = {
                "company_name": company_name,
                "location": location, 
                "company_website": website,
                "job_titles": job_titles,
                "output_file": f"{company_name.replace(' ', '_')}_{location.replace(' ', '_')}_employees.xlsx",
                "temp_data_file": "temp_employee_data.json",
                "processed_data_file": "processed_employee_data.json", 
                "verified_data_file": "verified_employee_data.json",
                "pages_to_scrape": 3,
                "debug_mode": False,
                "review_mode": False,
                "search_types": ["linkedin"],
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "version": "3.0"
            }
            
            # Save config
            config_path = self.script_dir / "company_config.json"
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
            
            self._print_status("Setup complete!")
            self._print_status(f"Company: {company_name}")
            self._print_status(f"Location: {location}")
            self._print_status(f"Job titles: {len(job_titles)} configured")
            
            return True
            
        except KeyboardInterrupt:
            print("\nSetup cancelled")
            return False
        except Exception as e:
            print(f"Setup error: {e}")
            return False
    
    def _get_default_job_titles(self) -> List[str]:
        """Default job titles."""
        return [
            "CEO", "Chief Executive Officer", "Managing Director", "President",
            "Chief Financial Officer", "CFO", "Finance Director", 
            "Chief Technology Officer", "CTO", "Technology Director",
            "Chief Operating Officer", "COO", "Operations Director",
            "Sales Director", "Sales Manager", "Business Development Manager",
            "Marketing Director", "Marketing Manager", "Head of Marketing",
            "Director", "Senior Director", "Vice President", "VP",
            "General Manager", "Regional Manager", "Branch Manager",
            "Head of Sales", "Head of Operations", "Head of Finance",
            "Senior Manager", "Manager", "Team Leader", "Department Head"
        ]
    
    def show_search_menu(self) -> str:
        """Show search method selection menu."""
        self._print_header("STEP 2: SELECT SEARCH METHOD")
        
        print("Available search methods:")
        print("1. LinkedIn Search (via Bing X-ray) - RECOMMENDED")
        print("2. Company Website Search")
        print("3. Both LinkedIn and Website")
        print("4. Exit")
        
        while True:
            choice = input("\nSelect search method (1-4, default: 1): ").strip()
            
            if choice == '' or choice == '1':
                return 'linkedin'
            elif choice == '2':
                return 'website'
            elif choice == '3':
                return 'both'
            elif choice == '4':
                return 'exit'
            else:
                print("Please enter 1, 2, 3, or 4")
    
    def run_linkedin_search(self) -> bool:
        """Run LinkedIn search with fixed encoding."""
        self._print_header("LINKEDIN X-RAY SEARCH")
        
        print(f"Searching LinkedIn for employees at {self.config['company_name']}")
        print(f"Location: {self.config['location']}")
        print(f"Job titles: {len(self.config['job_titles'])} different roles")
        print("This will take 3-7 minutes...")
        
        # Fix the LinkedIn script encoding first
        linkedin_script = self.script_dir / "script2_web_scraping.py"
        if not linkedin_script.exists():
            print("LinkedIn search script not found")
            return False
        
        # Create a fixed version of the LinkedIn script
        self._create_fixed_linkedin_script()
        
        try:
            # Run the fixed LinkedIn search
            fixed_script = self.script_dir / "script2_fixed.py"
            
            # Set environment to handle encoding
            env = os.environ.copy()
            env['PYTHONIOENCODING'] = 'utf-8'
            
            result = subprocess.run(
                [sys.executable, str(fixed_script)],
                capture_output=True,
                text=True,
                timeout=600,
                env=env,
                encoding='utf-8',
                errors='replace'
            )
            
            if result.returncode == 0:
                # Check for results
                linkedin_file = self.script_dir / "linkedin_candidates.json"
                if linkedin_file.exists():
                    with open(linkedin_file, 'r', encoding='utf-8') as f:
                        candidates = json.load(f)
                    
                    if candidates:
                        self._print_status(f"SUCCESS: Found {len(candidates)} LinkedIn profiles!")
                        return True
                    else:
                        self._print_status("Search completed but no profiles found")
                        return False
                else:
                    self._print_status("Search completed but no results file created")
                    return False
            else:
                print(f"LinkedIn search failed. Error: {result.stderr}")
                return False
                
        except subprocess.TimeoutExpired:
            print("LinkedIn search timed out")
            return False
        except Exception as e:
            print(f"LinkedIn search error: {e}")
            return False
    
    def _create_fixed_linkedin_script(self):
        """Create a fixed version of the LinkedIn script without emoji issues."""
        fixed_content = '''#!/usr/bin/env python3
"""
Fixed LinkedIn Search - Windows Compatible
"""
import json
import os
import sys
from pathlib import Path

# Add the original script directory to path
script_dir = Path(__file__).parent.absolute()
sys.path.insert(0, str(script_dir))

def main():
    """Run LinkedIn search with fixed encoding."""
    try:
        # Import the original script module
        import script2_web_scraping
        
        # Replace print statements that cause encoding issues
        original_print = print
        
        def safe_print(*args, **kwargs):
            """Safe print function that handles encoding."""
            try:
                original_print(*args, **kwargs)
            except UnicodeEncodeError:
                # Replace problematic characters
                safe_args = []
                for arg in args:
                    if isinstance(arg, str):
                        # Replace common emoji with text
                        safe_arg = arg.replace('🔍', '[SEARCH]')
                        safe_arg = safe_arg.replace('✅', '[OK]')
                        safe_arg = safe_arg.replace('❌', '[ERROR]')
                        safe_arg = safe_arg.replace('⚠️', '[WARNING]')
                        safe_arg = safe_arg.replace('🔗', '[LINK]')
                        safe_arg = safe_arg.replace('📊', '[DATA]')
                        safe_arg = safe_arg.replace('🚀', '[LAUNCH]')
                        safe_arg = safe_arg.replace('💡', '[INFO]')
                        safe_args.append(safe_arg)
                    else:
                        safe_args.append(arg)
                original_print(*safe_args, **kwargs)
        
        # Monkey patch print
        import builtins
        builtins.print = safe_print
        
        # Run the original main function
        script2_web_scraping.main()
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
'''
        
        fixed_script_path = self.script_dir / "script2_fixed.py"
        with open(fixed_script_path, 'w', encoding='utf-8') as f:
            f.write(fixed_content)
    
    def create_fixed_excel_report(self) -> bool:
        """Create Excel report with fixed merge cell handling."""
        self._print_header("GENERATING EXCEL REPORT")
        
        try:
            # Install openpyxl
            subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Load employee data
            all_employees = []
            
            # Check for LinkedIn results
            linkedin_file = self.script_dir / "linkedin_candidates.json"
            if linkedin_file.exists():
                with open(linkedin_file, 'r', encoding='utf-8') as f:
                    linkedin_data = json.load(f)
                all_employees.extend(linkedin_data)
            
            # Check for merged results
            merged_file = self.script_dir / "merged_employees.json"
            if merged_file.exists():
                with open(merged_file, 'r', encoding='utf-8') as f:
                    merged_data = json.load(f)
                all_employees.extend(merged_data)
            
            if not all_employees:
                print("No employee data found")
                return False
            
            # Remove duplicates
            seen = set()
            unique_employees = []
            for emp in all_employees:
                name_key = f"{emp.get('first_name', '').lower()}_{emp.get('last_name', '').lower()}"
                if name_key not in seen and name_key != '_':
                    unique_employees.append(emp)
                    seen.add(name_key)
            
            print(f"Creating Excel report with {len(unique_employees)} unique employees...")
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LinkedIn Search Results"
            
            # Title (avoid merge cells that cause issues)
            title_cell = ws['A1']
            title_cell.value = f"{self.config['company_name']} - LinkedIn Search Results"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Subtitle
            subtitle_cell = ws['A2']
            subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            subtitle_cell.font = Font(italic=True, size=10)
            
            # Headers
            headers = [
                "First Name", "Last Name", "Job Title", "LinkedIn URL",
                "Company", "Location", "Source", "Confidence"
            ]
            
            header_row = 4
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            linkedin_count = 0
            for row_num, emp in enumerate(unique_employees, header_row + 1):
                ws.cell(row=row_num, column=1, value=emp.get('first_name', ''))
                ws.cell(row=row_num, column=2, value=emp.get('last_name', ''))
                ws.cell(row=row_num, column=3, value=emp.get('title', ''))
                
                # LinkedIn URL
                linkedin_url = emp.get('linkedin_url') or emp.get('link', '')
                url_cell = ws.cell(row=row_num, column=4, value=linkedin_url)
                
                if linkedin_url and 'linkedin.com/in/' in linkedin_url.lower():
                    linkedin_count += 1
                    try:
                        url_cell.hyperlink = linkedin_url
                        url_cell.font = Font(color="0000FF", underline="single")
                    except:
                        pass
                
                ws.cell(row=row_num, column=5, value=emp.get('company_name', self.config['company_name']))
                ws.cell(row=row_num, column=6, value=emp.get('location', self.config['location']))
                ws.cell(row=row_num, column=7, value=emp.get('source', 'LinkedIn Search'))
                
                # Confidence with colors
                conf_cell = ws.cell(row=row_num, column=8, value=emp.get('confidence', 'medium').title())
                confidence = emp.get('confidence', 'medium').lower()
                if confidence == 'high':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif confidence == 'medium':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Auto-adjust columns (fixed method)
            for col_num in range(1, len(headers) + 1):
                column_letter = ws.cell(row=1, column=col_num).column_letter
                max_length = 0
                
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            company_clean = self.config['company_name'].replace(' ', '_')
            filename = f"{company_clean}_LinkedIn_Results_{timestamp}.xlsx"
            filepath = self.script_dir / filename
            
            wb.save(filepath)
            
            self._print_status(f"SUCCESS: Excel report created - {filename}")
            self._print_status(f"Contains {len(unique_employees)} employees")
            self._print_status(f"{linkedin_count} LinkedIn profiles found")
            
            # Open Excel file
            try:
                if sys.platform == 'win32':
                    os.startfile(str(filepath))
                    self._print_status("Excel file opened automatically")
                else:
                    subprocess.call(['open', str(filepath)])
            except:
                self._print_status(f"Please open manually: {filename}")
            
            return True
            
        except Exception as e:
            print(f"Excel creation error: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def run_linkedin_verification(self) -> bool:
        """Run LinkedIn verification."""
        self._print_header("LINKEDIN VERIFICATION")
        
        print("LinkedIn verification will:")
        print("  - Open browser to LinkedIn login")
        print("  - Wait for you to login manually")
        print("  - Verify current employment automatically")
        
        proceed = input("\nRun LinkedIn verification? (y/n, default: y): ").strip().lower()
        if proceed == 'n':
            self._print_status("Skipping LinkedIn verification")
            return False
        
        # Check if script5 exists
        script5_path = self.script_dir / "script5_linkedin_verification.py"
        if script5_path.exists():
            try:
                self._print_status("Launching LinkedIn verification...")
                result = subprocess.run([sys.executable, str(script5_path)])
                
                # Check if verification completed
                verified_file = self.script_dir / "linkedin_verified.json"
                if verified_file.exists():
                    self._print_status("LinkedIn verification completed")
                    return True
                else:
                    self._print_status("Verification may not have completed")
                    return False
            except Exception as e:
                print(f"Verification error: {e}")
                return False
        else:
            self._print_status("LinkedIn verification script not found")
            return False
    
    def run_complete_automation(self):
        """Run complete automation with menu."""
        self._print_header("FIXED EMPLOYEE DISCOVERY AUTOMATION")
        
        print("This automation will:")
        print("  1. Collect company information")
        print("  2. Show search method menu")
        print("  3. Run selected search method")
        print("  4. Generate Excel report")
        print("  5. Optional LinkedIn verification")
        print("\nTotal time: 10-15 minutes")
        print("Interactions: Company setup + search selection + LinkedIn login")
        
        proceed = input("\nStart automation? (y/n, default: y): ").strip().lower()
        if proceed == 'n':
            print("Automation cancelled")
            return
        
        start_time = time.time()
        
        try:
            # Step 1: Setup
            if not self.run_script1_setup():
                print("Setup failed - stopping")
                return
            
            # Step 2: Search method selection
            search_method = self.show_search_menu()
            
            if search_method == 'exit':
                print("Automation cancelled")
                return
            
            # Step 3: Run search
            search_success = False
            
            if search_method in ['linkedin', 'both']:
                search_success = self.run_linkedin_search()
            
            if search_method in ['website', 'both']:
                # Could add website search here
                print("Website search not implemented yet")
            
            if not search_success:
                print("Search failed - stopping")
                return
            
            # Step 4: Create Excel report
            if not self.create_fixed_excel_report():
                print("Excel generation failed")
                return
            
            # Step 5: Optional verification
            self.run_linkedin_verification()
            
            # Success
            elapsed = time.time() - start_time
            
            self._print_header("AUTOMATION COMPLETE!")
            print(f"Total time: {elapsed/60:.1f} minutes")
            print(f"Company: {self.config['company_name']}")
            print("Excel report generated and opened")
            print("Employee discovery complete!")
            
        except KeyboardInterrupt:
            print("\nAutomation cancelled by user")
        except Exception as e:
            print(f"Automation error: {e}")

def main():
    automation = FixedAutomation()
    automation.run_complete_automation()

if __name__ == "__main__":
    main()