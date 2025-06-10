#!/usr/bin/env python3
"""
Fixed Excel Output Generator

This script creates Excel reports from any employee data JSON files.
Completely rewritten to be simple, reliable, and actually work.
"""

import json
import os
import subprocess
import sys
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

def install_openpyxl():
    """Install openpyxl if not available."""
    try:
        import openpyxl
        return True
    except ImportError:
        print("Installing openpyxl...")
        try:
            subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
            return True
        except:
            print("Failed to install openpyxl")
            return False

def load_config() -> Dict:
    """Load configuration file."""
    script_dir = Path(__file__).parent.absolute()
    config_path = script_dir / "company_config.json"
    
    config = {}
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
        except:
            pass
    
    return config

def find_employee_data() -> tuple:
    """Find employee data from available JSON files."""
    script_dir = Path(__file__).parent.absolute()
    
    # Priority order for data files
    potential_files = [
        "verified_employee_data.json",
        "linkedin_candidates.json",
        "merged_employees.json",
        "script3a_verified_employees.json",
        "validated_employees.json",
        "processed_employee_data.json",
        "website_employees.json",
        "temp_employee_data.json"
    ]
    
    for filename in potential_files:
        file_path = script_dir / filename
        if file_path.exists():
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if isinstance(data, list) and data:
                    print(f"Found employee data in: {filename}")
                    print(f"Number of employees: {len(data)}")
                    return data, filename
            except Exception as e:
                print(f"Error reading {filename}: {e}")
                continue
    
    return [], None

def create_excel_report(employees: List[Dict], config: Dict, source_file: str) -> bool:
    """Create comprehensive Excel report from employee data."""
    if not install_openpyxl():
        return False
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    try:
        print(f"Creating Excel report with {len(employees)} employees...")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Data"
        
        # Title row
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"{config.get('company_name', 'Company')} - Employee Directory"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtitle row
        ws.merge_cells('A2:I2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} from {source_file}"
        subtitle_cell.font = Font(italic=True, size=10)
        subtitle_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            "First Name", "Last Name", "Job Title", "Company", 
            "Location", "Source", "Confidence", "Profile Link", "Notes"
        ]
        
        header_row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
        
        # Data rows
        current_row = header_row + 1
        linkedin_count = 0
        
        for employee in employees:
            # Basic data
            ws.cell(row=current_row, column=1, value=employee.get('first_name', ''))
            ws.cell(row=current_row, column=2, value=employee.get('last_name', ''))
            ws.cell(row=current_row, column=3, value=employee.get('title', ''))
            ws.cell(row=current_row, column=4, value=employee.get('company_name', config.get('company_name', '')))
            ws.cell(row=current_row, column=5, value=employee.get('location', config.get('location', '')))
            ws.cell(row=current_row, column=6, value=employee.get('source', ''))
            
            # Confidence with color coding
            conf_cell = ws.cell(row=current_row, column=7)
            confidence = employee.get('confidence', 'low').lower()
            conf_cell.value = confidence.title()
            
            if confidence == 'high':
                conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confidence == 'medium':
                conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Profile link with hyperlink
            link_cell = ws.cell(row=current_row, column=8)
            link = (employee.get('link') or 
                   employee.get('linkedin_url') or 
                   employee.get('source_link', ''))
            
            if link:
                link_cell.value = link
                try:
                    link_cell.hyperlink = link
                    link_cell.font = Font(color="0000FF", underline="single")
                    if 'linkedin.com/in/' in link.lower():
                        linkedin_count += 1
                except:
                    pass
            else:
                link_cell.value = "N/A"
            
            # Notes
            notes = []
            if employee.get('needs_verification'):
                notes.append("Needs verification")
            if 'linkedin' in employee.get('source', '').lower():
                notes.append("LinkedIn profile")
            
            ws.cell(row=current_row, column=9, value="; ".join(notes))
            
            current_row += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_sheet = wb.create_sheet(title="Summary")
        
        # Calculate statistics
        sources = {}
        confidence_levels = {}
        
        for emp in employees:
            source = emp.get('source', 'Unknown')
            sources[source] = sources.get(source, 0) + 1
            
            conf = emp.get('confidence', 'low')
            confidence_levels[conf] = confidence_levels.get(conf, 0) + 1
        
        summary_data = [
            ["Employee Data Summary Report", ""],
            ["", ""],
            ["Company", config.get('company_name', '')],
            ["Location", config.get('location', '')],
            ["Total Employees", len(employees)],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Data Source File", source_file],
            ["", ""],
            ["Data Sources", ""],
        ]
        
        # Add source breakdown
        for source, count in sorted(sources.items(), key=lambda x: x[1], reverse=True):
            summary_data.append([f"  {source}", count])
        
        summary_data.extend([
            ["", ""],
            ["Confidence Levels", ""],
        ])
        
        for conf, count in sorted(confidence_levels.items()):
            summary_data.append([f"  {conf.title()}", count])
        
        if linkedin_count > 0:
            summary_data.extend([
                ["", ""],
                ["LinkedIn Profiles", linkedin_count],
                ["Verification Available", "Yes" if linkedin_count > 0 else "No"],
                ["", ""],
                ["Next Steps", ""],
                ["1. Review LinkedIn profiles manually", ""],
                ["2. Run LinkedIn verification script", ""],
                ["3. Export data to CRM/database", ""],
            ])
        
        # Write summary data
        for row, (label, value) in enumerate(summary_data, 1):
            cell1 = summary_sheet.cell(row=row, column=1, value=label)
            cell2 = summary_sheet.cell(row=row, column=2, value=value)
            
            if label and not label.startswith(('  ', '1.', '2.', '3.')):
                cell1.font = Font(bold=True)
                if label == "Employee Data Summary Report":
                    cell1.font = Font(bold=True, size=14)
        
        # Auto-adjust summary columns
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 25
        
        # Add instructions sheet
        instructions_sheet = wb.create_sheet(title="Instructions")
        
        instructions_data = [
            ["How to Use This Employee Data", ""],
            ["", ""],
            ["Data Quality", ""],
            ["• High Confidence", "Most reliable data - use for important outreach"],
            ["• Medium Confidence", "Good data - verify before using"],
            ["• Low Confidence", "Needs manual verification"],
            ["", ""],
            ["LinkedIn Profiles", ""],
            ["• Click profile links", "Opens LinkedIn profiles in browser"],
            ["• Verify current employment", "Check if they still work at target company"],
            ["• Run verification script", "Use script5_linkedin_verification.py for bulk verification"],
            ["", ""],
            ["Export Options", ""],
            ["• Copy to CRM", "Select data and copy to customer management system"],
            ["• Save as CSV", "File > Save As > CSV for database import"],
            ["• Email lists", "Use first/last names for email address generation"],
            ["", ""],
            ["Data Privacy", ""],
            ["• Use responsibly", "Respect privacy laws and LinkedIn terms"],
            ["• Business use only", "Don't use for spam or unsolicited marketing"],
            ["• Keep updated", "Re-run scripts periodically for fresh data"],
        ]
        
        for row, (instruction, detail) in enumerate(instructions_data, 1):
            cell1 = instructions_sheet.cell(row=row, column=1, value=instruction)
            cell2 = instructions_sheet.cell(row=row, column=2, value=detail)
            
            if instruction and not instruction.startswith('•'):
                cell1.font = Font(bold=True)
                if instruction == "How to Use This Employee Data":
                    cell1.font = Font(bold=True, size=14)
        
        instructions_sheet.column_dimensions['A'].width = 25
        instructions_sheet.column_dimensions['B'].width = 60
        
        # Generate filename
        company_clean = config.get('company_name', 'Company').replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{company_clean}_Employee_Report_{timestamp}.xlsx"
        
        # Save workbook
        script_dir = Path(__file__).parent.absolute()
        filepath = script_dir / filename
        
        wb.save(filepath)
        
        print(f"✅ Excel report created: {filename}")
        print(f"📊 Contains {len(employees)} employees across {len(wb.sheetnames)} sheets")
        print(f"📁 Location: {filepath}")
        
        if linkedin_count > 0:
            print(f"🔗 {linkedin_count} LinkedIn profiles available for verification")
        
        # Try to open the file
        try:
            if sys.platform == 'win32':
                os.startfile(str(filepath))
            elif sys.platform == 'darwin':
                subprocess.call(['open', str(filepath)])
            else:
                subprocess.call(['xdg-open', str(filepath)])
            print("📂 Excel file opened automatically")
        except:
            print("📂 Please open the Excel file manually")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating Excel report: {e}")
        import traceback
        traceback.print_exc()
        return False

def show_data_preview(employees: List[Dict], max_show: int = 5):
    """Show preview of employee data."""
    print(f"\n👥 Sample employee data:")
    for i, emp in enumerate(employees[:max_show], 1):
        name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}"
        title = emp.get('title', 'Unknown')
        source = emp.get('source', 'Unknown')
        confidence = emp.get('confidence', 'unknown')
        
        # Check for LinkedIn
        link = (emp.get('link') or 
               emp.get('linkedin_url') or 
               emp.get('source_link', ''))
        linkedin_indicator = " 🔗" if link and 'linkedin.com/in/' in link.lower() else ""
        
        print(f"  {i}. {name} - {title} ({confidence} confidence){linkedin_indicator}")
        print(f"     Source: {source}")
    
    if len(employees) > max_show:
        print(f"  ... and {len(employees) - max_show} more employees")

def main():
    """Main function to create Excel from employee data."""
    print("=" * 60)
    print("EXCEL REPORT GENERATOR")
    print("=" * 60)
    print("Creates comprehensive Excel reports from employee data")
    
    # Load configuration
    config = load_config()
    print(f"\nCompany: {config.get('company_name', 'Not specified')}")
    print(f"Location: {config.get('location', 'Not specified')}")
    
    # Find employee data
    employees, source_file = find_employee_data()
    
    if not employees:
        print("\n❌ No employee data found!")
        print("\nAvailable files in directory:")
        script_dir = Path(__file__).parent.absolute()
        json_files = list(script_dir.glob("*.json"))
        if json_files:
            for file in sorted(json_files):
                print(f"  - {file.name}")
        else:
            print("  - No JSON files found")
        
        print("\n💡 To generate employee data, run:")
        print("   1. script1_input_collection.py (setup)")
        print("   2. script2_web_scraping.py (LinkedIn search)")
        return
    
    # Show data preview
    show_data_preview(employees)
    
    # Show statistics
    sources = {}
    confidence_counts = {}
    linkedin_count = 0
    
    for emp in employees:
        source = emp.get('source', 'Unknown')
        sources[source] = sources.get(source, 0) + 1
        
        conf = emp.get('confidence', 'low')
        confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
        
        link = (emp.get('link') or 
               emp.get('linkedin_url') or 
               emp.get('source_link', ''))
        if link and 'linkedin.com/in/' in link.lower():
            linkedin_count += 1
    
    print(f"\n📈 Data Statistics:")
    print(f"   Total employees: {len(employees)}")
    print(f"   Data sources: {len(sources)}")
    if linkedin_count > 0:
        print(f"   LinkedIn profiles: {linkedin_count}")
    
    print(f"\n📊 By source:")
    for source, count in sorted(sources.items(), key=lambda x: x[1], reverse=True):
        print(f"   - {source}: {count}")
    
    print(f"\n🎯 By confidence:")
    for conf, count in sorted(confidence_counts.items()):
        print(f"   - {conf.title()}: {count}")
    
    # Create Excel report
    print(f"\n📋 Creating Excel report...")
    
    if create_excel_report(employees, config, source_file):
        print(f"\n🎉 Success! Excel report created successfully!")
        
        if linkedin_count > 0:
            print(f"\n💡 Next Steps:")
            print(f"   • Review LinkedIn profiles manually by clicking links")
            print(f"   • Run script5_linkedin_verification.py for bulk verification")
            print(f"   • Export verified data to your CRM or database")
        else:
            print(f"\n💡 Next Steps:")
            print(f"   • Review employee data for accuracy")
            print(f"   • Use contact information for outreach")
            print(f"   • Consider running LinkedIn search to find more profiles")
    else:
        print("\n❌ Failed to create Excel report")

if __name__ == "__main__":
    main()