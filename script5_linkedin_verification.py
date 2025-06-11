#!/usr/bin/env python3
"""
LinkedIn Profile Verification Script - PDF EXTRACTION VERSION

Verifies LinkedIn profiles by saving them as PDFs and extracting structured data.
PDFs are automatically cleaned up after processing to save storage space.
"""

import json
import logging
import os
import random
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import tempfile

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def install_dependencies():
    """Install required packages automatically."""
    required_packages = [
        'selenium', 'webdriver-manager', 'openpyxl', 
        'PyPDF2', 'pdfplumber', 'requests'
    ]
    
    for package in required_packages:
        try:
            if package == 'webdriver-manager':
                import webdriver_manager
            elif package == 'openpyxl':
                import openpyxl
            elif package == 'PyPDF2':
                import PyPDF2
            elif package == 'pdfplumber':
                import pdfplumber
            else:
                __import__(package)
        except ImportError:
            logger.info(f"Installing {package}...")
            # Handle special cases for package installation
            if package == 'PyPDF2':
                subprocess.call([sys.executable, "-m", "pip", "install", "PyPDF2"])
            elif package == 'pdfplumber':
                # pdfplumber sometimes has dependency issues, install with all dependencies
                subprocess.call([sys.executable, "-m", "pip", "install", "pdfplumber[all]"])
            else:
                subprocess.call([sys.executable, "-m", "pip", "install", package])

def install_dependencies():
    """Install required packages automatically."""
    required_packages = [
        'selenium', 'webdriver-manager', 'openpyxl', 
        'PyPDF2', 'pdfplumber', 'requests'
    ]
    
    print("🔧 Checking and installing required dependencies...")
    
    for package in required_packages:
        try:
            if package == 'webdriver-manager':
                import webdriver_manager  # type: ignore
            elif package == 'openpyxl':
                import openpyxl  # type: ignore
            elif package == 'PyPDF2':
                import PyPDF2  # type: ignore
            elif package == 'pdfplumber':
                import pdfplumber  # type: ignore
            elif package == 'selenium':
                import selenium  # type: ignore
            elif package == 'requests':
                import requests  # type: ignore
            else:
                __import__(package)  # type: ignore
            print(f"✅ {package} - already installed")
        except ImportError:
            print(f"📦 Installing {package}...")
            try:
                # Handle special cases for package installation
                if package == 'PyPDF2':
                    subprocess.call([sys.executable, "-m", "pip", "install", "PyPDF2"])
                elif package == 'pdfplumber':
                    # Try installing pdfplumber with all dependencies
                    result = subprocess.call([sys.executable, "-m", "pip", "install", "pdfplumber"])
                    if result != 0:
                        print(f"⚠️ pdfplumber installation failed, will use PyPDF2 only")
                        continue
                else:
                    subprocess.call([sys.executable, "-m", "pip", "install", package])
                print(f"✅ {package} - installed successfully")
            except Exception as e:
                print(f"❌ Failed to install {package}: {e}")
                if package in ['PyPDF2', 'pdfplumber']:
                    print(f"   PDF processing may be limited without {package}")
                else:
                    print(f"   This may cause issues with the script")

# Call installation before any imports
install_dependencies()

print("🔄 Loading libraries...")

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    from webdriver_manager.chrome import ChromeDriverManager
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import requests
    print("✅ Core libraries loaded successfully")
    
    # Import PDF libraries with fallback handling
    PDF_LIBRARIES_AVAILABLE = []
    
    print("🔍 Checking PDF processing libraries...")
    try:
        import pdfplumber  # type: ignore
        PDF_LIBRARIES_AVAILABLE.append('pdfplumber')
        print("✅ pdfplumber - available")
    except ImportError as e:
        print(f"⚠️ pdfplumber not available: {e}")
        print("   Will attempt to install...")
        try:
            subprocess.call([sys.executable, "-m", "pip", "install", "pdfplumber"])
            import pdfplumber  # type: ignore
            PDF_LIBRARIES_AVAILABLE.append('pdfplumber')
            print("✅ pdfplumber - installed and loaded")
        except Exception as install_error:
            print(f"❌ pdfplumber installation failed: {install_error}")
    
    try:
        import PyPDF2  # type: ignore
        PDF_LIBRARIES_AVAILABLE.append('PyPDF2')
        print("✅ PyPDF2 - available")
    except ImportError as e:
        print(f"⚠️ PyPDF2 not available: {e}")
        print("   Will attempt to install...")
        try:
            subprocess.call([sys.executable, "-m", "pip", "install", "PyPDF2"])
            import PyPDF2  # type: ignore
            PDF_LIBRARIES_AVAILABLE.append('PyPDF2')
            print("✅ PyPDF2 - installed and loaded")
        except Exception as install_error:
            print(f"❌ PyPDF2 installation failed: {install_error}")
    
    if not PDF_LIBRARIES_AVAILABLE:
        print("❌ ERROR: No PDF processing libraries available!")
        print("   Please manually install one of these:")
        print("   pip install pdfplumber")
        print("   pip install PyPDF2")
        sys.exit(1)
    else:
        print(f"✅ PDF processing available using: {', '.join(PDF_LIBRARIES_AVAILABLE)}")
        
except ImportError as e:
    print(f"❌ Failed to import required packages: {e}")
    print("   Please install missing packages manually:")
    print("   pip install selenium webdriver-manager openpyxl PyPDF2 pdfplumber requests")
    sys.exit(1)

class LinkedInPDFVerifier:
    """LinkedIn profile verifier using PDF extraction method."""
    
    def __init__(self, config_path: str):
        self.config = self._load_config(config_path)
        self.script_dir = Path(__file__).parent.absolute()
        self.temp_dir = Path(tempfile.gettempdir()) / "linkedin_pdfs"
        self.temp_dir.mkdir(exist_ok=True)
        self.driver = None
        
        # Company information for verification
        self.target_company = self.config.get('company_name', '').lower()
        self.target_location = self.config.get('location', '').lower()
        
        logger.info(f"LinkedIn PDF Verifier initialized for {self.target_company}")
        
    def _load_config(self, config_path: str) -> dict:
        """Load configuration."""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Configuration error: {e}")
            return {}
    
    def _setup_browser_for_pdf(self) -> Optional[webdriver.Chrome]:
        """Setup Chrome browser optimized for PDF generation."""
        try:
            logger.info("Setting up Chrome browser for PDF extraction...")
            
            options = Options()
            
            # PDF generation settings
            pdf_settings = {
                "printing.print_preview_sticky_settings.appState": json.dumps({
                    "recentDestinations": [{
                        "id": "Save as PDF",
                        "origin": "local",
                        "account": ""
                    }],
                    "selectedDestinationId": "Save as PDF",
                    "version": 2,
                    "marginsType": 0,
                    "isLandscapeEnabled": False,
                    "isCssBackgroundEnabled": True,
                    "mediaSize": {
                        "height_microns": 297000,
                        "name": "ISO_A4",
                        "width_microns": 210000
                    }
                })
            }
            
            options.add_experimental_option('prefs', {
                'plugins.always_open_pdf_externally': True,
                'download.default_directory': str(self.temp_dir),
                'download.prompt_for_download': False,
                'download.directory_upgrade': True,
                'safebrowsing.enabled': True,
                **pdf_settings
            })
            
            # Standard options for LinkedIn access
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-extensions')
            options.add_argument('--disable-gpu')
            options.add_argument('--window-size=1200,800')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option('excludeSwitches', ['enable-automation'])
            options.add_experimental_option('useAutomationExtension', False)
            
            # Random user agent
            user_agents = [
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15'
            ]
            options.add_argument(f'user-agent={random.choice(user_agents)}')
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            driver.set_page_load_timeout(30)
            driver.implicitly_wait(10)
            
            # Anti-detection
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            logger.info("✅ Browser setup successful for PDF extraction")
            return driver
            
        except Exception as e:
            logger.error(f"❌ Browser setup error: {e}")
            return None
    
    def _save_profile_as_pdf(self, linkedin_url: str, profile_name: str) -> Optional[Path]:
        """Save LinkedIn profile as PDF and return the file path."""
        try:
            logger.info(f"Navigating to LinkedIn profile: {profile_name}")
            
            # Navigate to profile
            self.driver.get(linkedin_url)
            time.sleep(random.uniform(3, 5))
            
            # Wait for profile content to load
            try:
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '[data-generated-suggestion-target]'))
                )
            except TimeoutException:
                # Fallback - wait for any main content
                try:
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'main'))
                    )
                except TimeoutException:
                    logger.warning(f"Profile content may not have loaded fully for {profile_name}")
            
            # Scroll to load more content
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight/2);")
            time.sleep(2)
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            
            # Generate PDF filename
            safe_name = re.sub(r'[^\w\-_]', '_', profile_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename = f"linkedin_{safe_name}_{timestamp}.pdf"
            pdf_path = self.temp_dir / pdf_filename
            
            # Use Chrome's print to PDF functionality
            print_options = {
                'landscape': False,
                'displayHeaderFooter': False,
                'printBackground': True,
                'preferCSSPageSize': True,
                'paperWidth': 8.27,  # A4 width in inches
                'paperHeight': 11.7,  # A4 height in inches
                'marginTop': 0.4,
                'marginBottom': 0.4,
                'marginLeft': 0.4,
                'marginRight': 0.4
            }
            
            # Execute Chrome DevTools print command
            result = self.driver.execute_cdp_cmd('Page.printToPDF', print_options)
            
            # Save PDF data to file
            with open(pdf_path, 'wb') as f:
                f.write(base64.b64decode(result['data']))
            
            if pdf_path.exists():
                logger.info(f"✅ PDF saved: {pdf_filename} ({pdf_path.stat().st_size} bytes)")
                return pdf_path
            else:
                logger.error(f"❌ PDF file not created: {pdf_filename}")
                return None
                
        except Exception as e:
            logger.error(f"❌ Error saving PDF for {profile_name}: {e}")
            return None
    
    def _extract_data_from_pdf(self, pdf_path: Path, profile_name: str) -> Dict:
        """Extract structured data from LinkedIn PDF."""
        try:
            logger.info(f"Extracting data from PDF: {pdf_path.name}")
            
            extracted_data = {
                'name': profile_name,
                'current_company': None,
                'current_title': None,
                'location': None,
                'employment_verified': False,
                'still_employed': False,
                'extraction_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'raw_text_sample': None
            }
            
            # Try available PDF libraries in order of preference
            text_content = ""
            extraction_method = "none"
            
            # Try pdfplumber first (better for structured extraction)
            if 'pdfplumber' in PDF_LIBRARIES_AVAILABLE:
                try:
                    import pdfplumber  # type: ignore # Import only when needed
                    with pdfplumber.open(pdf_path) as pdf:
                        for page in pdf.pages:
                            page_text = page.extract_text()
                            if page_text:
                                text_content += page_text + "\n"
                    extraction_method = "pdfplumber"
                    logger.info(f"✅ Used pdfplumber for {pdf_path.name}")
                except Exception as e:
                    logger.warning(f"pdfplumber failed for {pdf_path.name}: {e}")
                    text_content = ""
            
            # Fallback to PyPDF2 if pdfplumber failed or not available
            if not text_content and 'PyPDF2' in PDF_LIBRARIES_AVAILABLE:
                try:
                    import PyPDF2  # type: ignore # Import only when needed
                    with open(pdf_path, 'rb') as file:
                        pdf_reader = PyPDF2.PdfReader(file)
                        for page in pdf_reader.pages:
                            text_content += page.extract_text() + "\n"
                    extraction_method = "PyPDF2"
                    logger.info(f"✅ Used PyPDF2 for {pdf_path.name}")
                except Exception as e:
                    logger.error(f"PyPDF2 also failed for {pdf_path.name}: {e}")
                    text_content = ""
            
            if not text_content.strip():
                logger.error(f"No text extracted from PDF: {pdf_path.name}")
                extracted_data['extraction_method'] = "failed"
                return extracted_data
            
            extracted_data['extraction_method'] = extraction_method
            
            # Store sample of raw text for debugging
            extracted_data['raw_text_sample'] = text_content[:500]
            
            # Extract current employment information
            current_employment = self._parse_current_employment(text_content)
            extracted_data.update(current_employment)
            
            # Extract location
            location = self._parse_location(text_content)
            if location:
                extracted_data['location'] = location
            
            # Verify employment with target company
            if extracted_data['current_company']:
                extracted_data['employment_verified'] = True
                extracted_data['still_employed'] = self._verify_company_match(
                    extracted_data['current_company']
                )
            
            logger.info(f"✅ Data extracted for {profile_name} using {extraction_method}")
            if extracted_data['current_company']:
                logger.info(f"   Current: {extracted_data['current_title']} at {extracted_data['current_company']}")
                logger.info(f"   Still employed at target: {extracted_data['still_employed']}")
            
            return extracted_data
            
        except Exception as e:
            logger.error(f"❌ Error extracting data from PDF {pdf_path.name}: {e}")
            extracted_data['extraction_method'] = "error"
            return extracted_data
    
    def _parse_current_employment(self, text: str) -> Dict:
        """Parse current employment information from PDF text."""
        employment_data = {
            'current_company': None,
            'current_title': None
        }
        
        try:
            # Split into lines for analysis
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # Look for employment patterns
            employment_patterns = [
                # "Title at Company" patterns
                r'([^·\n]+?)\s+at\s+([^·\n]+?)(?:\s*·|\s*$|\s*\n)',
                
                # "Company" followed by "Title" patterns
                r'(?:^|\n)([A-Z][^·\n]+?)\s*\n([^·\n]+?)(?:\s*·|\s*$)',
                
                # Experience section patterns
                r'Experience\s*\n([^·\n]+?)\s*\n([^·\n]+?)(?:\s*·|\s*$)',
                
                # Direct title/company patterns
                r'([A-Z][a-zA-Z\s]+?(?:Manager|Director|Engineer|Analyst|Specialist|Officer|Executive|President|VP|Lead))\s*\n([A-Z][^·\n]+?)(?:\s*·|\s*$)'
            ]
            
            for pattern in employment_patterns:
                matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
                
                for match in matches:
                    if len(match.groups()) >= 2:
                        part1, part2 = match.groups()[:2]
                        
                        # Determine which is title and which is company
                        if self._looks_like_job_title(part1) and self._looks_like_company(part2):
                            employment_data['current_title'] = part1.strip()
                            employment_data['current_company'] = part2.strip()
                            break
                        elif self._looks_like_company(part1) and self._looks_like_job_title(part2):
                            employment_data['current_company'] = part1.strip()
                            employment_data['current_title'] = part2.strip()
                            break
                
                if employment_data['current_company']:
                    break
            
            # If still no company found, look for target company specifically
            if not employment_data['current_company'] and self.target_company:
                company_pattern = rf'\b{re.escape(self.target_company)}\b'
                if re.search(company_pattern, text, re.IGNORECASE):
                    employment_data['current_company'] = self.target_company.title()
                    
                    # Try to find title near the company mention
                    company_context = self._extract_context_around_company(text, self.target_company)
                    if company_context:
                        title = self._extract_title_from_context(company_context)
                        if title:
                            employment_data['current_title'] = title
            
        except Exception as e:
            logger.error(f"Error parsing employment: {e}")
        
        return employment_data
    
    def _looks_like_job_title(self, text: str) -> bool:
        """Check if text looks like a job title."""
        if not text or len(text) < 3 or len(text) > 100:
            return False
        
        title_indicators = [
            'manager', 'director', 'engineer', 'analyst', 'specialist',
            'officer', 'executive', 'president', 'vice president', 'vp',
            'lead', 'senior', 'junior', 'head', 'chief', 'coordinator',
            'consultant', 'developer', 'designer', 'architect'
        ]
        
        text_lower = text.lower()
        return any(indicator in text_lower for indicator in title_indicators)
    
    def _looks_like_company(self, text: str) -> bool:
        """Check if text looks like a company name."""
        if not text or len(text) < 2 or len(text) > 80:
            return False
        
        # Companies often have these patterns
        company_indicators = [
            'ltd', 'limited', 'inc', 'corp', 'corporation', 'llc',
            'company', 'group', 'holdings', 'solutions', 'services',
            'technologies', 'systems', 'associates', 'partners'
        ]
        
        text_lower = text.lower()
        
        # Check for company suffixes
        if any(text_lower.endswith(indicator) for indicator in company_indicators):
            return True
        
        # Check for company words
        if any(indicator in text_lower for indicator in company_indicators):
            return True
        
        # Check if it starts with capital letter and doesn't look like a title
        if text[0].isupper() and not self._looks_like_job_title(text):
            return True
        
        return False
    
    def _extract_context_around_company(self, text: str, company: str) -> str:
        """Extract text context around company mention."""
        try:
            company_pattern = rf'.{{0,100}}\b{re.escape(company)}\b.{{0,100}}'
            match = re.search(company_pattern, text, re.IGNORECASE | re.DOTALL)
            if match:
                return match.group(0)
        except Exception:
            pass
        return ""
    
    def _extract_title_from_context(self, context: str) -> Optional[str]:
        """Extract job title from context around company."""
        # Look for title patterns in the context
        title_patterns = [
            r'([A-Z][a-zA-Z\s]*(?:Manager|Director|Engineer|Analyst|Specialist|Officer|Executive|President|VP|Lead))',
            r'((?:Senior|Junior|Lead|Principal|Chief)\s+[A-Z][a-zA-Z\s]*)',
        ]
        
        for pattern in title_patterns:
            match = re.search(pattern, context, re.IGNORECASE)
            if match:
                title = match.group(1).strip()
                if self._looks_like_job_title(title):
                    return title
        
        return None
    
    def _parse_location(self, text: str) -> Optional[str]:
        """Parse location information from PDF text."""
        try:
            # Look for location patterns
            location_patterns = [
                r'(?:Location|Based in|Located in)\s*[:\-]?\s*([^·\n]+)',
                r'([A-Z][a-z]+,\s*[A-Z][a-z\s]+?)(?:\s*·|\s*$|\s*\n)',  # City, Country/State
                r'(?:^|\n)([A-Z][a-z]+,\s*[A-Z]{2,3})(?:\s*·|\s*$)',  # City, State/Country code
            ]
            
            for pattern in location_patterns:
                match = re.search(pattern, text, re.MULTILINE)
                if match:
                    location = match.group(1).strip()
                    if len(location) > 3 and len(location) < 50:
                        return location
            
        except Exception as e:
            logger.error(f"Error parsing location: {e}")
        
        return None
    
    def _verify_company_match(self, current_company: str) -> bool:
        """Verify if current company matches target company."""
        if not current_company or not self.target_company:
            return False
        
        current_lower = current_company.lower()
        target_lower = self.target_company.lower()
        
        # Exact match
        if current_lower == target_lower:
            return True
        
        # Partial match (target company name in current company)
        if target_lower in current_lower or current_lower in target_lower:
            return True
        
        # Handle common variations
        # Remove common suffixes for comparison
        common_suffixes = ['ltd', 'limited', 'inc', 'corp', 'corporation', 'llc', 'plc']
        
        def clean_company_name(name):
            name = name.lower().strip()
            for suffix in common_suffixes:
                if name.endswith(' ' + suffix):
                    name = name[:-len(suffix)-1].strip()
            return name
        
        clean_current = clean_company_name(current_company)
        clean_target = clean_company_name(self.target_company)
        
        return clean_current == clean_target
    
    def _cleanup_pdf(self, pdf_path: Path) -> bool:
        """Clean up PDF file to save storage space."""
        try:
            if pdf_path.exists():
                pdf_path.unlink()
                logger.info(f"🗑️ Cleaned up PDF: {pdf_path.name}")
                return True
        except Exception as e:
            logger.warning(f"Failed to cleanup PDF {pdf_path.name}: {e}")
        return False
    
    def _find_linkedin_profiles(self) -> List[Dict]:
        """Find LinkedIn profiles from data files."""
        potential_files = [
            "linkedin_candidates.json",
            "merged_employees.json", 
            "verified_employee_data.json",
            "script3a_verified_employees.json"
        ]
        
        linkedin_profiles = []
        
        for filename in potential_files:
            file_path = self.script_dir / filename
            if file_path.exists():
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    for item in data:
                        linkedin_url = (item.get('link') or 
                                      item.get('linkedin_url') or 
                                      item.get('source_link', ''))
                        
                        if linkedin_url and 'linkedin.com/in/' in linkedin_url.lower():
                            # Create profile record
                            profile = {
                                'first_name': item.get('first_name', ''),
                                'last_name': item.get('last_name', ''),
                                'title': item.get('title', ''),
                                'linkedin_url': linkedin_url,
                                'confidence': item.get('confidence', 'low'),
                                'source': item.get('source', 'Unknown'),
                                'original_data': item
                            }
                            linkedin_profiles.append(profile)
                except Exception as e:
                    logger.warning(f"Error reading {filename}: {e}")
                    continue
        
        # Remove duplicates based on LinkedIn URL
        unique_profiles = {}
        for profile in linkedin_profiles:
            url = profile['linkedin_url']
            if url not in unique_profiles:
                unique_profiles[url] = profile
        
        return list(unique_profiles.values())
    
    def verify_linkedin_profiles(self) -> bool:
        """Main method to verify LinkedIn profiles using PDF extraction."""
        try:
            print("🔗 Starting LinkedIn PDF Verification...")
            print(f"Target Company: {self.config.get('company_name', 'Not specified')}")
            print(f"Target Location: {self.config.get('location', 'Not specified')}")
            
            # Find LinkedIn profiles
            linkedin_profiles = self._find_linkedin_profiles()
            
            if not linkedin_profiles:
                print("⚠️ No LinkedIn profiles found for verification")
                return False
            
            print(f"✅ Found {len(linkedin_profiles)} LinkedIn profiles to verify")
            
            # Setup browser
            self.driver = self._setup_browser_for_pdf()
            if not self.driver:
                print("❌ Failed to setup browser")
                return False
            
            verified_profiles = []
            
            try:
                for i, profile in enumerate(linkedin_profiles, 1):
                    profile_name = f"{profile['first_name']} {profile['last_name']}"
                    print(f"\n🔍 Verifying {i}/{len(linkedin_profiles)}: {profile_name}")
                    
                    # Save profile as PDF
                    pdf_path = self._save_profile_as_pdf(profile['linkedin_url'], profile_name)
                    
                    if pdf_path:
                        # Extract data from PDF
                        extracted_data = self._extract_data_from_pdf(pdf_path, profile_name)
                        
                        # Merge with original profile data
                        verified_profile = profile['original_data'].copy()
                        verified_profile.update({
                            'verification_method': 'pdf_extraction',
                            'verification_date': extracted_data['extraction_date'],
                            'current_company_verified': extracted_data['current_company'],
                            'current_title_verified': extracted_data['current_title'],
                            'location_verified': extracted_data['location'],
                            'still_employed_at_target': extracted_data['still_employed'],
                            'employment_verification_status': 'verified' if extracted_data['employment_verified'] else 'unverified',
                            'pdf_extraction_successful': True
                        })
                        
                        verified_profiles.append(verified_profile)
                        
                        # Cleanup PDF
                        self._cleanup_pdf(pdf_path)
                        
                        # Add delay between profiles to avoid rate limiting
                        time.sleep(random.uniform(2, 4))
                    else:
                        # Handle failed PDF generation
                        failed_profile = profile['original_data'].copy()
                        failed_profile.update({
                            'verification_method': 'pdf_extraction',
                            'verification_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            'pdf_extraction_successful': False,
                            'employment_verification_status': 'failed'
                        })
                        verified_profiles.append(failed_profile)
                        
                        logger.warning(f"PDF extraction failed for {profile_name}")
            
            finally:
                # Close browser
                if self.driver:
                    try:
                        self.driver.quit()
                    except:
                        pass
            
            # Save verification results
            if verified_profiles:
                self._save_verification_results(verified_profiles)
                self._create_verification_excel(verified_profiles)
                
                # Print summary
                successful_verifications = sum(1 for p in verified_profiles if p.get('pdf_extraction_successful', False))
                still_employed = sum(1 for p in verified_profiles if p.get('still_employed_at_target', False))
                
                print(f"\n📊 Verification Summary:")
                print(f"   Total profiles: {len(verified_profiles)}")
                print(f"   Successful extractions: {successful_verifications}")
                print(f"   Still employed at {self.target_company}: {still_employed}")
                
                return True
            else:
                print("❌ No profiles were successfully verified")
                return False
                
        except Exception as e:
            logger.error(f"Verification process failed: {e}")
            return False
        finally:
            # Cleanup temp directory
            self._cleanup_temp_directory()
    
    def _save_verification_results(self, verified_profiles: List[Dict]) -> bool:
        """Save verification results to JSON file."""
        try:
            output_file = self.script_dir / "linkedin_verification_results.json"
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(verified_profiles, f, indent=4, ensure_ascii=False)
            
            logger.info(f"✅ Verification results saved to {output_file.name}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to save verification results: {e}")
            return False
    
    def _create_verification_excel(self, verified_profiles: List[Dict]) -> bool:
        """Create Excel report with verification results."""
        try:
            logger.info("Creating LinkedIn verification Excel report...")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LinkedIn Verification"
            
            # Headers
            headers = [
                "First Name", "Last Name", "Original Title", "LinkedIn URL",
                "Verification Status", "Current Company (Verified)", 
                "Current Title (Verified)", "Location (Verified)",
                "Still Employed at Target", "Verification Date", "Confidence"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            
            # Write data
            for row, profile in enumerate(verified_profiles, 2):
                ws.cell(row=row, column=1, value=profile.get('first_name', ''))
                ws.cell(row=row, column=2, value=profile.get('last_name', ''))
                ws.cell(row=row, column=3, value=profile.get('title', ''))
                
                # LinkedIn URL with hyperlink
                linkedin_url = (profile.get('link') or 
                              profile.get('linkedin_url') or 
                              profile.get('source_link', ''))
                url_cell = ws.cell(row=row, column=4, value=linkedin_url)
                if linkedin_url:
                    url_cell.hyperlink = linkedin_url
                    url_cell.font = Font(color="0000FF", underline="single")
                
                # Verification status with color coding
                status = profile.get('employment_verification_status', 'unknown')
                status_cell = ws.cell(row=row, column=5, value=status.title())
                if status == 'verified':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif status == 'failed':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                ws.cell(row=row, column=6, value=profile.get('current_company_verified', ''))
                ws.cell(row=row, column=7, value=profile.get('current_title_verified', ''))
                ws.cell(row=row, column=8, value=profile.get('location_verified', ''))
                
                # Still employed status with color coding
                still_employed = profile.get('still_employed_at_target', False)
                employed_cell = ws.cell(row=row, column=9, value="Yes" if still_employed else "No")
                if still_employed:
                    employed_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    employed_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row, column=10, value=profile.get('verification_date', ''))
                ws.cell(row=row, column=11, value=profile.get('confidence', ''))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            
            # Calculate statistics
            total_profiles = len(verified_profiles)
            successful_verifications = sum(1 for p in verified_profiles if p.get('pdf_extraction_successful', False))
            still_employed = sum(1 for p in verified_profiles if p.get('still_employed_at_target', False))
            verification_rate = (successful_verifications / total_profiles * 100) if total_profiles > 0 else 0
            
            summary_data = [
                ["LinkedIn Verification Summary", ""],
                ["", ""],
                ["Company", self.config.get('company_name', '')],
                ["Location", self.config.get('location', '')],
                ["Verification Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
                ["Total Profiles Processed", total_profiles],
                ["Successful Extractions", successful_verifications],
                ["Verification Success Rate", f"{verification_rate:.1f}%"],
                ["Still Employed at Target", still_employed],
                ["Employment Retention Rate", f"{(still_employed / total_profiles * 100):.1f}%" if total_profiles > 0 else "0%"],
                ["", ""],
                ["Verification Method", "PDF Extraction"],
                ["PDFs Auto-Cleaned", "Yes"],
                ["", ""],
                ["Key Insights", ""],
                ["• High retention rate indicates stable workforce", ""],
                ["• Low retention may suggest talent movement", ""],
                ["• Failed extractions may need manual review", ""],
            ]
            
            for row, (label, value) in enumerate(summary_data, 1):
                cell1 = summary_sheet.cell(row=row, column=1, value=label)
                cell2 = summary_sheet.cell(row=row, column=2, value=value)
                
                if label and not label.startswith(('•', '  ')):
                    cell1.font = Font(bold=True)
                    if label == "LinkedIn Verification Summary":
                        cell1.font = Font(bold=True, size=14)
            
            summary_sheet.column_dimensions['A'].width = 35
            summary_sheet.column_dimensions['B'].width = 25
            
            # Save Excel file
            company_clean = self.config.get('company_name', 'Company').replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"LinkedIn_Verification_{company_clean}_{timestamp}.xlsx"
            filepath = self.script_dir / filename
            
            wb.save(filepath)
            logger.info(f"✅ LinkedIn verification Excel saved: {filename}")
            
            # Try to open the file
            try:
                if sys.platform == 'win32':
                    os.startfile(str(filepath))
                elif sys.platform == 'darwin':
                    subprocess.call(['open', str(filepath)])
                else:
                    subprocess.call(['xdg-open', str(filepath)])
                logger.info("📂 Excel file opened automatically")
            except:
                logger.info("📂 Please open the Excel file manually")
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to create verification Excel: {e}")
            return False
    
    def _cleanup_temp_directory(self) -> bool:
        """Clean up temporary PDF directory."""
        try:
            if self.temp_dir.exists():
                # Remove any remaining PDF files
                for pdf_file in self.temp_dir.glob("*.pdf"):
                    try:
                        pdf_file.unlink()
                    except:
                        pass
                
                # Try to remove the directory if empty
                try:
                    self.temp_dir.rmdir()
                    logger.info("🗑️ Temporary PDF directory cleaned up")
                except OSError:
                    # Directory not empty, that's okay
                    pass
            
            return True
            
        except Exception as e:
            logger.warning(f"Failed to cleanup temp directory: {e}")
            return False

# Add missing import for base64
import base64

def main():
    """Main execution function."""
    try:
        script_dir = Path(__file__).parent.absolute()
        config_path = script_dir / "company_config.json"
        
        print("=" * 70)
        print("LINKEDIN PDF VERIFICATION - ENHANCED VERSION")
        print("=" * 70)
        print("This script verifies LinkedIn profiles using PDF extraction")
        print("PDFs are automatically cleaned up after processing")
        
        if not config_path.exists():
            print("⚠️ Configuration file not found")
            print("Please run script1_input_collection.py first")
            return False
        
        # Initialize verifier
        verifier = LinkedInPDFVerifier(str(config_path))
        
        # Start verification process
        success = verifier.verify_linkedin_profiles()
        
        if success:
            print("\n🎉 LinkedIn verification completed successfully!")
            print("📊 Check the Excel report for detailed results")
            print("💾 Verification data saved to JSON file")
            
            # Optionally launch next script
            launch_next = input("\n🚀 Launch Excel output script? (y/n, default: y): ").strip().lower()
            if launch_next != 'n':
                excel_script = script_dir / "script4_excel_output.py"
                if excel_script.exists():
                    try:
                        if sys.platform == 'win32':
                            subprocess.Popen([sys.executable, str(excel_script)], 
                                           creationflags=subprocess.CREATE_NEW_CONSOLE)
                        else:
                            subprocess.Popen([sys.executable, str(excel_script)])
                        print("📊 Excel generation launched!")
                    except Exception as e:
                        print(f"⚠️ Could not launch Excel script: {e}")
        else:
            print("\n❌ LinkedIn verification failed")
            print("Check the logs for details")
        
        return success
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        return False
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(f"\n❌ An error occurred: {e}")
        return False

if __name__ == "__main__":
    main()