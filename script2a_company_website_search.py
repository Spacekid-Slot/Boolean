#!/usr/bin/env python3
"""
Enhanced Company Website Employee Search - Clean Version with Job Title Search

This script searches company websites for employee information
with comprehensive filtering, job title targeting, and Excel generation.
"""

import json
import logging
import os
import random
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict, Optional, Set
from urllib.parse import quote_plus, urlparse
from dataclasses import dataclass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Install dependencies if needed
def install_dependencies():
    """Install required packages if not available."""
    required_packages = ['selenium', 'webdriver-manager']
    
    for package in required_packages:
        try:
            if package == 'webdriver-manager':
                import webdriver_manager
            else:
                __import__(package)
        except ImportError:
            logger.info(f"Installing {package}...")
            subprocess.call([sys.executable, "-m", "pip", "install", package])

install_dependencies()

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, 
        WebDriverException, SessionNotCreatedException
    )
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError as e:
    logger.error(f"Failed to import Selenium: {e}")
    sys.exit(1)

@dataclass
class Employee:
    """Employee data structure for website search results."""
    first_name: str
    last_name: str
    title: str
    source_link: str
    company_name: str
    location: str
    confidence: str
    source: str = 'Website Search Results'

class WebsiteEmployeeScraper:
    """Enhanced website-based employee scraper with job title targeting."""
    
    def __init__(self, config_path: str):
        """Initialize the website scraper."""
        self.config = self._load_config(config_path)
        self.script_dir = Path(__file__).parent.absolute()
        self.driver = None
        self.max_retries = 3
        self.processed_names: Set[str] = set()
        
        # Job titles to search for
        self.job_titles = self._get_job_titles_to_search()
        
        # User agents for rotation
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0'
        ]
    
    def _load_config(self, config_path: str) -> dict:
        """Load configuration from JSON file."""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.error(f"Configuration error: {e}")
            sys.exit(1)
    
    def _get_job_titles_to_search(self) -> List[str]:
        """Get list of job titles to search for based on user input or defaults."""
        # Check if job titles are provided in config
        if 'job_titles' in self.config and self.config['job_titles']:
            return self.config['job_titles']
        
        # Default comprehensive list of job titles
        default_titles = [
            # Executive/Leadership
            "CEO", "Chief Executive Officer", "Managing Director", "President",
            "Chief Financial Officer", "CFO", "Chief Technology Officer", "CTO",
            "Chief Operating Officer", "COO", "Chief Marketing Officer", "CMO",
            
            # Management
            "Director", "Manager", "Head of", "Vice President", "VP",
            "Senior Manager", "Regional Manager", "General Manager",
            
            # Property/Real Estate (common in Edinburgh)
            "Property Manager", "Estate Agent", "Letting Agent", "Property Factor",
            "Property Director", "Asset Manager", "Development Manager",
            "Property Investment Manager", "Facilities Manager",
            
            # Professional/Technical
            "Engineer", "Developer", "Analyst", "Consultant", "Specialist",
            "Senior Engineer", "Lead Developer", "Principal Consultant",
            "Technical Lead", "Project Manager", "Account Manager",
            
            # Finance/Accounting
            "Accountant", "Financial Analyst", "Investment Manager", "Fund Manager",
            "Financial Controller", "Finance Manager", "Treasury Manager",
            
            # Sales/Marketing
            "Sales Manager", "Marketing Manager", "Business Development",
            "Account Executive", "Sales Director", "Marketing Director",
            
            # Operations
            "Operations Manager", "Operations Director", "Process Manager",
            "Quality Manager", "Compliance Manager", "Risk Manager"
        ]
        
        return default_titles
    
    def _setup_browser(self) -> webdriver.Chrome:
        """Configure and return a Chrome WebDriver."""
        logger.info("Setting up browser...")
        
        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument(f'user-agent={random.choice(self.user_agents)}')
        
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(30)
            driver.implicitly_wait(10)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            return driver
        except Exception as e:
            logger.error(f"Browser setup error: {e}")
            raise
    
    def _is_valid_employee_name(self, first_name: str, last_name: str) -> bool:
        """Check if extracted name parts are likely to be valid employee names."""
        # Basic validation
        if not first_name or not last_name or len(first_name) < 2 or len(last_name) < 2:
            return False
        
        if len(first_name) > 25 or len(last_name) > 30:
            return False
        
        if not re.match(r"^[a-zA-Z\-']+$", first_name) or not re.match(r"^[a-zA-Z\-']+$", last_name):
            return False
        
        # Enhanced business term detection
        business_terms = {
            'appoints', 'announces', 'welcomes', 'promotes', 'elevates', 'names',
            'chief', 'officer', 'executive', 'financial', 'revenue', 'operating',
            'strategy', 'technology', 'interim', 'board', 'chair', 'director',
            'positions', 'hedge', 'fund', 'managers', 'portfolio', 'operations',
            'leadership', 'general', 'terms', 'acquisition', 'clearwater', 
            'analytics', 'finalizes', 'investor', 'quarter', 'press', 'release',
            'client', 'success', 'traders', 'demo', 'back', 'build', 'nothing',
            'move', 'learning', 'read', 'more', 'open', 'second', 'first',
            'company', 'corporation', 'enterprise', 'group', 'holdings',
            'management', 'development', 'consulting', 'services', 'solutions',
            'keeps', 'your', 'office', 'enfusion', 'sustains', 'global'
        }
        
        if (first_name.lower() in business_terms or 
            last_name.lower() in business_terms):
            return False
        
        # Check for title combinations
        full_name_lower = f"{first_name.lower()} {last_name.lower()}"
        title_combinations = [
            'chief executive', 'chief financial', 'chief technology', 'chief operating',
            'executive officer', 'financial officer', 'technology officer', 'operating officer',
            'vice president', 'board chair', 'board member', 'managing director',
            'keeps your', 'back office', 'enfusion names', 'enfusion announces',
            'enfusion sustains', 'global growth'
        ]
        
        for combo in title_combinations:
            if combo in full_name_lower:
                return False
        
        return True
    
    def _extract_employees_from_text(self, text: str, company_name: str, source_link: str) -> List[Employee]:
        """Extract employee information from text content with improved accuracy and job title matching."""
        employees = []
        
        try:
            # Enhanced employee patterns including job title-specific searches
            employee_patterns = [
                # "appoints/names/promotes FirstName LastName as/to Title"
                r'(?:appoints?|names?|promotes?|welcomes?|announces?)\s+([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?)\s+(?:as|to|as\s+the|to\s+the\s+position\s+of)\s+([^.]{5,50})',
                
                # "FirstName LastName, Title" or "FirstName LastName as Title"
                r'([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?),?\s+(?:as\s+)?([^.,]{10,60}(?:Chief|Director|Manager|Officer|President|VP|Vice\s+President)[^.,]{0,20})',
                
                # "Title FirstName LastName" (when title comes first)
                r'(?:Chief|Director|Manager|Officer|President|VP|Vice\s+President)\s+([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?)',
                
                # Look for patterns like "Brad Herring has been appointed"
                r'([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?)\s+(?:has\s+been|was|is)\s+(?:appointed|named|promoted|hired)',
                
                # Job title specific patterns
                r'([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?)\s+(?:joins|joined)\s+.*?(?:as|as\s+a|as\s+the)\s+([^.,]{5,50})',
                
                # Pattern for "John Smith is our new Property Manager"
                r'([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?)\s+is\s+(?:our\s+)?(?:new\s+)?([^.,]{5,50}(?:Manager|Director|Officer|Executive|Analyst|Specialist)[^.,]{0,20})'
            ]
            
            # Add patterns for specific job titles we're searching for
            for job_title in self.job_titles[:10]:  # Use first 10 to avoid too many patterns
                # Pattern: "John Smith, Property Manager"
                title_pattern = rf'([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?),?\s+{re.escape(job_title)}'
                employee_patterns.append(title_pattern)
                
                # Pattern: "Property Manager John Smith"
                reverse_pattern = rf'{re.escape(job_title)}\s+([A-Z][a-z]+)\s+([A-Z][a-z]+(?:-[A-Z][a-z]+)?)'
                employee_patterns.append(reverse_pattern)
            
            for pattern in employee_patterns:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                
                for match in matches:
                    try:
                        groups = match.groups()
                        
                        if len(groups) >= 2:
                            first_name = groups[0].strip()
                            last_name = groups[1].strip()
                            
                            # Extract title if available
                            title = "Unknown"
                            if len(groups) >= 3 and groups[2]:
                                title = groups[2].strip()
                                # Clean up title
                                title = re.sub(r'^(as\s+|to\s+|the\s+)', '', title, flags=re.IGNORECASE)
                                title = title.strip('.,')
                            else:
                                # Try to infer title from the pattern match
                                title = self._infer_title_from_context(match.group(0), first_name, last_name)
                            
                            # Validate this is a real person name
                            if not self._is_valid_employee_name(first_name, last_name):
                                continue
                            
                            # Check if already processed
                            name_key = f"{first_name.lower()}_{last_name.lower()}"
                            if name_key in self.processed_names:
                                continue
                            
                            # Determine confidence based on job title match and context
                            confidence = self._determine_confidence_level(title, match.group(0))
                            
                            # Create employee record
                            employee = Employee(
                                first_name=first_name,
                                last_name=last_name,
                                title=title,
                                source_link=source_link,
                                company_name=company_name,
                                location=self.config.get('location', ''),
                                confidence=confidence,
                                source='Website Search Results'
                            )
                            
                            employees.append(employee)
                            self.processed_names.add(name_key)
                            logger.info(f"Found: {first_name} {last_name} - {title} (confidence: {confidence})")
                            
                    except Exception as e:
                        logger.debug(f"Error processing employee pattern match: {e}")
                        continue
            
        except Exception as e:
            logger.error(f"Error extracting employees from text: {e}")
        
        return employees
    
    def _infer_title_from_context(self, match_text: str, first_name: str, last_name: str) -> str:
        """Infer job title from the context of the match."""
        # Look for title keywords in the surrounding text
        title_keywords = {
            'manager': 'Manager',
            'director': 'Director', 
            'executive': 'Executive',
            'officer': 'Officer',
            'president': 'President',
            'head': 'Head',
            'lead': 'Lead',
            'chief': 'Chief',
            'senior': 'Senior',
            'principal': 'Principal'
        }
        
        match_lower = match_text.lower()
        for keyword, title in title_keywords.items():
            if keyword in match_lower:
                return title
        
        # Check against our job title list
        for job_title in self.job_titles:
            if job_title.lower() in match_lower:
                return job_title
        
        return "Unknown"
    
    def _determine_confidence_level(self, title: str, context: str) -> str:
        """Determine confidence level based on title match and context."""
        score = 0
        
        # Base score for having a title
        if title and title != "Unknown":
            score += 2
        
        # Boost score if title matches our search list
        if title != "Unknown":
            for search_title in self.job_titles:
                if search_title.lower() in title.lower() or title.lower() in search_title.lower():
                    score += 3
                    break
        
        # Context indicators
        context_lower = context.lower()
        positive_indicators = ['appointed', 'promoted', 'joins', 'joined', 'announces', 'welcomes']
        if any(indicator in context_lower for indicator in positive_indicators):
            score += 2
        
        # Job title quality indicators
        quality_titles = ['director', 'manager', 'executive', 'officer', 'president', 'chief', 'head']
        if any(quality in title.lower() for quality in quality_titles):
            score += 1
        
        # Determine confidence level
        if score >= 5:
            return 'high'
        elif score >= 3:
            return 'medium'
        else:
            return 'low'
    
    def _process_search_results_page(self, company_name: str) -> List[Employee]:
        """Process a page of search results for employee information."""
        employees = []
        
        try:
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'li.b_algo'))
            )
            
            result_elements = self.driver.find_elements(By.CSS_SELECTOR, 'li.b_algo')
            logger.info(f"Found {len(result_elements)} search results")
            
            for result in result_elements:
                try:
                    title_element = result.find_element(By.CSS_SELECTOR, 'h2 a')
                    title = title_element.text
                    link = title_element.get_attribute('href')
                    
                    if not link:
                        continue
                    
                    # Extract description
                    description = ""
                    try:
                        desc_element = result.find_element(By.CSS_SELECTOR, '.b_caption p')
                        description = desc_element.text
                    except NoSuchElementException:
                        pass
                    
                    # Combine title and description for analysis
                    content = f"{title}\n{description}"
                    
                    # Extract employees from the content
                    page_employees = self._extract_employees_from_text(content, company_name, link)
                    
                    for emp in page_employees:
                        employees.append(emp)
                        logger.info(f"Found employee: {emp.first_name} {emp.last_name}")
                    
                except Exception as e:
                    logger.debug(f"Error processing search result: {e}")
                    continue
            
        except TimeoutException:
            logger.warning("Timeout waiting for search results")
        except Exception as e:
            logger.error(f"Error processing search results page: {e}")
        
        return employees
    
    def search_employees(self) -> List[Dict]:
        """Main employee search method with job title integration."""
        company_name = self.config['company_name']
        website = self.config['company_website']
        location = self.config['location']
        max_pages = self.config.get('pages_to_scrape', 5)
        
        all_employees = []
        
        logger.info(f"Starting website search for {company_name}")
        logger.info(f"Will search for {len(self.job_titles)} different job titles")
        
        try:
            self.driver = self._setup_browser()
            
            # Create search queries with job title integration
            parsed_url = urlparse(website if website.startswith('http') else f'https://{website}')
            domain = parsed_url.netloc or parsed_url.path
            if domain.startswith('www.'):
                domain = domain[4:]
            
            site_query = f"site:{domain}"
            
            # Base queries
            queries = [
                f'{site_query} (team OR staff OR people OR "our team")',
                f'{site_query} ("about us" OR "leadership" OR "management team")',
                f'{site_query} (joined OR appointed OR promoted)',
                f'{site_query} (director OR manager OR executive)',
            ]
            
            # Add job title specific queries
            for job_title in self.job_titles[:10]:  # Limit to first 10 titles
                title_queries = [
                    f'{site_query} "{job_title}"',
                    f'{site_query} "{job_title}" (appointed OR joined OR promoted)',
                    f'{site_query} "new {job_title}" OR "{job_title} joins"',
                ]
                queries.extend(title_queries)
            
            # Alternative patterns with job titles
            alt_queries = [
                f'{site_query} "Property Manager" OR "Asset Manager" OR "Development Manager"',
                f'{site_query} "Sales Director" OR "Marketing Director" OR "Operations Director"',
                f'{site_query} "Chief Executive" OR "Managing Director" OR "President"',
                f'{site_query} "announces" (Manager OR Director OR Executive)',
                f'{site_query} "welcomes" (Manager OR Director OR Executive)',
            ]
            queries.extend(alt_queries)
            
            logger.info(f"Created {len(queries)} search queries including job title searches")
            
            for query_idx, query in enumerate(queries, 1):
                if len(all_employees) >= 30:  # Reasonable limit for website search
                    break
                
                logger.info(f"Processing query {query_idx}/{len(queries)}: {query}")
                
                try:
                    search_url = f"https://www.bing.com/search?q={quote_plus(query)}"
                    self.driver.get(search_url)
                    time.sleep(random.uniform(2, 4))
                    
                    # Process pages
                    for page_num in range(min(max_pages, 3)):  # Limit to 3 pages for website search
                        logger.info(f"Processing page {page_num + 1}")
                        
                        page_employees = self._process_search_results_page(company_name)
                        all_employees.extend(page_employees)
                        
                        if len(all_employees) >= 30:
                            break
                    
                    time.sleep(random.uniform(3, 6))
                    
                except Exception as e:
                    logger.error(f"Error processing query: {e}")
                    continue
            
        except Exception as e:
            logger.error(f"Search error: {e}")
        
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except Exception as e:
                    logger.warning(f"Error closing browser: {e}")
        
        logger.info(f"Website search completed. Found {len(all_employees)} employees")
        return [emp.__dict__ for emp in all_employees]
    
    def create_excel_report(self, employees: List[Dict]) -> bool:
        """Create Excel report from website search results."""
        if not employees:
            logger.warning("No employees to include in Excel report")
            print("No employees found for Excel report")
            return False
        
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                logger.info("Installing openpyxl...")
                subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            
            logger.info("Creating Excel report...")
            print(f"Creating Excel report with {len(employees)} employees...")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Website Search Results"
            
            headers = ["First Name", "Last Name", "Job Title", "Source", "Confidence", "Source Link", "Location"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, emp in enumerate(employees, 2):
                ws.cell(row=row, column=1, value=emp.get('first_name', ''))
                ws.cell(row=row, column=2, value=emp.get('last_name', ''))
                ws.cell(row=row, column=3, value=emp.get('title', ''))
                ws.cell(row=row, column=4, value=emp.get('source', ''))
                
                conf_cell = ws.cell(row=row, column=5, value=emp.get('confidence', '').upper())
                if conf_cell.value == 'HIGH':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif conf_cell.value == 'MEDIUM':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                link = emp.get('source_link', '') or emp.get('link', '')
                link_cell = ws.cell(row=row, column=6, value=link)
                if link:
                    link_cell.hyperlink = link
                    link_cell.font = Font(color="0000FF", underline="single")
                
                ws.cell(row=row, column=7, value=emp.get('location', ''))
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            company_name = self.config.get('company_name', 'Company').replace(' ', '_')
            location = self.config.get('location', 'Location').replace(' ', '_')
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_file = f"{company_name}_{location}_Website_Search_{timestamp}.xlsx"
            output_path = self.script_dir / output_file
            
            wb.save(output_path)
            
            # Verify file creation
            if output_path.exists():
                file_size = output_path.stat().st_size
                logger.info(f"Excel report saved to {output_file}")
                print(f"Excel report saved: {output_file} ({file_size} bytes)")
                
                try:
                    if sys.platform == 'win32':
                        os.startfile(str(output_path))
                    elif sys.platform == 'darwin':
                        subprocess.call(['open', str(output_path)])
                    else:
                        subprocess.call(['xdg-open', str(output_path)])
                except Exception:
                    logger.info("Please open the Excel file manually")
                
                return True
            else:
                logger.error("Excel file was not created")
                print("ERROR: Excel file was not created")
                return False
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            print(f"ERROR creating Excel report: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def save_results(self, employees: List[Dict]) -> bool:
        """Save search results to files."""
        try:
            website_file = self.script_dir / "website_employees.json"
            with open(website_file, 'w', encoding='utf-8') as f:
                json.dump(employees, f, indent=4, ensure_ascii=False)
            
            logger.info(f"Results saved: {len(employees)} employees")
            return True
            
        except Exception as e:
            logger.error(f"Error saving results: {e}")
            return False

def main():
    """Main execution function."""
    try:
        script_dir = Path(__file__).parent.absolute()
        config_path = script_dir / "company_config.json"
        
        if not config_path.exists():
            logger.error("Configuration file not found. Please run script1_input_collection.py first.")
            sys.exit(1)
        
        print("=" * 60)
        print("WEBSITE EMPLOYEE SCRAPER (ENHANCED WITH JOB TITLES)")
        print("=" * 60)
        
        scraper = WebsiteEmployeeScraper(str(config_path))
        
        print(f"Company: {scraper.config['company_name']}")
        print(f"Website: {scraper.config['company_website']}")
        print(f"Location: {scraper.config['location']}")
        print(f"Job titles to search: {len(scraper.job_titles)}")
        
        if not scraper.config.get('company_website'):
            logger.error("No company website found in configuration.")
            sys.exit(1)
        
        # Display some job titles being searched
        print(f"\nSample job titles: {', '.join(scraper.job_titles[:5])}...")
        
        employees = scraper.search_employees()
        
        if employees:
            print(f"\nWebsite search completed successfully!")
            print(f"Found {len(employees)} employees")
            
            if scraper.save_results(employees):
                # Show summary by confidence and job titles
                confidence_counts = {}
                job_title_counts = {}
                for emp in employees:
                    conf = emp.get('confidence', 'unknown')
                    confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
                    
                    title = emp.get('title', 'Unknown')
                    if title != 'Unknown':
                        job_title_counts[title] = job_title_counts.get(title, 0) + 1
                
                print("\nConfidence breakdown:")
                for conf, count in sorted(confidence_counts.items()):
                    print(f"  - {conf.title()}: {count}")
                
                print(f"\nFound {len(job_title_counts)} unique job titles")
                if job_title_counts:
                    print("Top job titles found:")
                    for title, count in sorted(job_title_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                        print(f"  - {title}: {count}")
                
                print("\nGenerating Excel report...")
                if scraper.create_excel_report(employees):
                    print("Excel report generated successfully!")
                else:
                    print("Warning: Excel report generation failed")
            else:
                print("Failed to save results")
                sys.exit(1)
        else:
            print("No employees found.")
    
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()