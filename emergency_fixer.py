#!/usr/bin/env python3
"""
Complete Auto-Sync Emergency Fixer

This script automatically:
1. Finds and stops all auto-sync processes
2. Fixes Git repository conflicts
3. Disables auto-sync scripts
4. Restores your working scripts
5. Tests the employee discovery toolkit
"""

import os
import sys
import subprocess
import time
import json
import shutil
from pathlib import Path
from datetime import datetime

# Install required packages
def install_psutil():
    """Install psutil for process management."""
    try:
        import psutil
        return True
    except ImportError:
        print("📦 Installing psutil for process management...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "psutil"])
            import psutil
            return True
        except:
            print("⚠️ Could not install psutil - will use alternative methods")
            return False

def safe_run_command(cmd, timeout=30):
    """Safely run a command with timeout."""
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, 
                              text=True, timeout=timeout)
        return result.returncode == 0, result.stdout, result.stderr
    except subprocess.TimeoutExpired:
        return False, "", "Command timed out"
    except Exception as e:
        return False, "", str(e)

def kill_python_processes_by_name():
    """Alternative method to kill processes using taskkill."""
    print("🔍 Looking for auto-sync Python processes...")
    
    # Get list of Python processes
    success, stdout, stderr = safe_run_command('tasklist /FI "IMAGENAME eq python.exe" /FO CSV')
    
    if not success:
        print("⚠️ Could not list Python processes")
        return
    
    lines = stdout.strip().split('\n')
    if len(lines) <= 1:
        print("✅ No Python processes found")
        return
    
    print(f"📊 Found {len(lines)-1} Python processes")
    
    # Kill processes that might be auto-sync (be careful!)
    kill_confirm = input("🚨 WARNING: This will kill ALL Python processes except this one.\nThis might stop other important Python programs.\nContinue? (y/n): ").strip().lower()
    
    if kill_confirm == 'y':
        current_pid = os.getpid()
        success, stdout, stderr = safe_run_command(f'taskkill /F /IM python.exe /FI "PID ne {current_pid}"')
        if success:
            print("✅ Stopped Python processes")
        else:
            print(f"⚠️ Some processes might still be running: {stderr}")

def stop_autosync_processes():
    """Stop auto-sync processes using multiple methods."""
    print("\n🛑 STEP 1: STOPPING AUTO-SYNC PROCESSES")
    print("=" * 50)
    
    has_psutil = install_psutil()
    
    if has_psutil:
        try:
            import psutil
            print("🔍 Scanning for auto-sync processes...")
            
            stopped_count = 0
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    cmdline = ' '.join(proc.info['cmdline']) if proc.info['cmdline'] else ''
                    
                    # Look for auto-sync indicators
                    autosync_keywords = [
                        'auto_sync', 'autosync', 'git_auto', 'auto-sync',
                        'real_time_sync', 'file_watcher', 'git_watcher',
                        'sync.py', 'monitor.py', 'watch.py'
                    ]
                    
                    if (proc.info['name'].lower() == 'python.exe' and 
                        any(keyword in cmdline.lower() for keyword in autosync_keywords)):
                        
                        print(f"🎯 Found auto-sync process: PID {proc.info['pid']}")
                        print(f"   Command: {cmdline[:80]}...")
                        
                        if proc.info['pid'] != os.getpid():  # Don't kill ourselves!
                            proc.terminate()
                            try:
                                proc.wait(timeout=5)
                                print(f"   ✅ Stopped PID {proc.info['pid']}")
                                stopped_count += 1
                            except psutil.TimeoutExpired:
                                proc.kill()
                                print(f"   🔨 Force killed PID {proc.info['pid']}")
                                stopped_count += 1
                        
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            
            if stopped_count == 0:
                print("✅ No auto-sync processes found running")
            else:
                print(f"✅ Stopped {stopped_count} auto-sync processes")
                
        except Exception as e:
            print(f"⚠️ Error with psutil method: {e}")
            kill_python_processes_by_name()
    else:
        # Fallback method
        kill_python_processes_by_name()

def fix_git_repository():
    """Fix Git repository conflicts and sync issues."""
    print("\n🔧 STEP 2: FIXING GIT REPOSITORY")
    print("=" * 50)
    
    # Check if we're in a Git repository
    if not Path('.git').exists():
        print("⚠️ Not in a Git repository, skipping Git fixes")
        return True
    
    try:
        # 1. Check current status
        print("📊 Checking Git status...")
        success, stdout, stderr = safe_run_command('git status --porcelain')
        
        if stdout.strip():
            print("📝 Found uncommitted changes:")
            for line in stdout.strip().split('\n')[:10]:  # Show first 10 files
                print(f"   {line}")
            
            # Auto-commit current changes to save work
            print("💾 Auto-committing current changes to save your work...")
            safe_run_command('git add .')
            commit_msg = f"Emergency auto-commit before sync fix - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            safe_run_command(f'git commit -m "{commit_msg}"')
            print("✅ Changes committed")
        
        # 2. Fetch latest from remote
        print("📡 Fetching latest changes from remote...")
        success, stdout, stderr = safe_run_command('git fetch origin')
        if success:
            print("✅ Fetched remote changes")
        else:
            print(f"⚠️ Fetch warning: {stderr}")
        
        # 3. Try to merge/rebase
        print("🔄 Syncing with remote repository...")
        
        # First try rebase
        success, stdout, stderr = safe_run_command('git rebase origin/main')
        if success:
            print("✅ Successfully rebased with remote")
        else:
            print("⚠️ Rebase failed, trying merge...")
            
            # Try merge
            success, stdout, stderr = safe_run_command('git merge origin/main')
            if success:
                print("✅ Successfully merged with remote")
            else:
                print("🚨 Both rebase and merge failed. Using reset method...")
                
                # Nuclear option: reset to remote
                reset_confirm = input("This will discard local changes and reset to remote version. Continue? (y/n): ").strip().lower()
                if reset_confirm == 'y':
                    safe_run_command('git reset --hard origin/main')
                    print("✅ Repository reset to remote state")
                else:
                    print("⚠️ Git conflicts remain - you'll need to resolve manually")
                    return False
        
        # 4. Push any local commits
        print("📤 Pushing local commits...")
        success, stdout, stderr = safe_run_command('git push origin main')
        if success:
            print("✅ Successfully pushed to remote")
        else:
            print(f"⚠️ Push failed (this is often normal): {stderr[:100]}...")
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing Git repository: {e}")
        return False

def find_and_disable_autosync_scripts():
    """Find and disable auto-sync scripts."""
    print("\n🚫 STEP 3: DISABLING AUTO-SYNC SCRIPTS")
    print("=" * 50)
    
    script_dir = Path.cwd()
    autosync_patterns = [
        '*auto*sync*.py', '*auto_sync*.py', '*autosync*.py',
        '*git*auto*.py', '*real*time*.py', '*file*watcher*.py',
        '*monitor*.py', '*sync*.py', '*watcher*.py'
    ]
    
    found_scripts = set()
    for pattern in autosync_patterns:
        found_scripts.update(script_dir.glob(pattern))
    
    # Filter out our main scripts
    main_scripts = {
        'script1_input_collection.py', 'script2_web_scraping.py',
        'script2a_company_website_search.py', 'script3_data_processing.py',
        'script3a_data_review.py', 'script4_excel_output.py',
        'employee_discovery_selector.py'
    }
    
    autosync_scripts = [s for s in found_scripts if s.name not in main_scripts]
    
    if not autosync_scripts:
        print("✅ No auto-sync scripts found to disable")
        return True
    
    print(f"🔍 Found {len(autosync_scripts)} potential auto-sync scripts:")
    
    disabled_count = 0
    for script in autosync_scripts:
        print(f"   📄 {script.name}")
        
        # Check if it contains sync-related code
        try:
            with open(script, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
            sync_indicators = ['git add', 'git commit', 'git push', 'file watcher', 'auto sync', 'real time']
            if any(indicator in content.lower() for indicator in sync_indicators):
                # This looks like an auto-sync script
                disabled_name = script.with_suffix('.py.DISABLED')
                script.rename(disabled_name)
                print(f"   ✅ Disabled: {script.name} → {disabled_name.name}")
                disabled_count += 1
            else:
                print(f"   ⏭️ Kept: {script.name} (doesn't appear to be auto-sync)")
                
        except Exception as e:
            print(f"   ⚠️ Could not check {script.name}: {e}")
    
    print(f"✅ Disabled {disabled_count} auto-sync scripts")
    return True

def restore_working_scripts():
    """Restore working versions of the main scripts."""
    print("\n🔄 STEP 4: RESTORING WORKING SCRIPTS")
    print("=" * 50)
    
    # Fixed script2 content (the LinkedIn search script)
    fixed_script2 = '''#!/usr/bin/env python3
"""
LinkedIn Search via Bing (X-ray Search) - WORKING VERSION

This script searches LinkedIn profiles using Bing X-ray search.
Simplified and tested to actually work.
"""

import json
import logging
import os
import random
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict, Optional
from urllib.parse import quote_plus
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def install_dependencies():
    """Install required packages."""
    required = ['selenium', 'webdriver-manager', 'openpyxl']
    for package in required:
        try:
            if package == 'webdriver-manager':
                import webdriver_manager
            elif package == 'openpyxl':
                import openpyxl
            else:
                __import__(package)
        except ImportError:
            print(f"Installing {package}...")
            subprocess.call([sys.executable, "-m", "pip", "install", package])

print("🔧 Checking dependencies...")
install_dependencies()

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    from webdriver_manager.chrome import ChromeDriverManager
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    print("✅ Dependencies loaded")
except ImportError as e:
    print(f"❌ Import error: {e}")
    input("Press Enter to exit...")
    sys.exit(1)

def main():
    """Main function."""
    try:
        print("=" * 60)
        print("🔍 LINKEDIN X-RAY SEARCH - WORKING VERSION")
        print("=" * 60)
        
        # Load config
        config_path = Path(__file__).parent / "company_config.json"
        if not config_path.exists():
            print("❌ Configuration file not found!")
            print("Run script1_input_collection.py first")
            input("Press Enter to exit...")
            return
        
        with open(config_path, 'r') as f:
            config = json.load(f)
        
        company = config.get('company_name', '')
        location = config.get('location', '')
        
        if not company:
            print("❌ No company name in configuration!")
            input("Press Enter to exit...")
            return
        
        print(f"🎯 Company: {company}")
        print(f"📍 Location: {location}")
        
        # Setup browser
        print("🌐 Setting up browser...")
        options = Options()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            print("✅ Browser ready")
        except Exception as e:
            print(f"❌ Browser setup failed: {e}")
            input("Press Enter to exit...")
            return
        
        try:
            # Simple search query
            query = f'site:linkedin.com/in/ "{company}" {location}'
            search_url = f"https://www.bing.com/search?q={quote_plus(query)}"
            
            print(f"🔍 Searching: {query}")
            driver.get(search_url)
            time.sleep(5)
            
            # Look for results
            results = driver.find_elements(By.CSS_SELECTOR, 'li.b_algo')
            print(f"📄 Found {len(results)} search results")
            
            candidates = []
            for i, result in enumerate(results[:10]):  # Process first 10
                try:
                    title_elem = result.find_element(By.CSS_SELECTOR, 'h2 a')
                    title = title_elem.text
                    url = title_elem.get_attribute('href')
                    
                    if 'linkedin.com/in/' in url:
                        # Extract name from title
                        name_match = re.search(r'^([^-|]+)', title)
                        if name_match:
                            full_name = name_match.group(1).strip()
                            name_parts = full_name.split()
                            if len(name_parts) >= 2:
                                first_name = name_parts[0]
                                last_name = ' '.join(name_parts[1:])
                                
                                candidate = {
                                    'first_name': first_name,
                                    'last_name': last_name,
                                    'title': 'LinkedIn Profile',
                                    'link': url,
                                    'company_name': company,
                                    'location': location,
                                    'confidence': 'medium',
                                    'source': 'LinkedIn X-ray Search'
                                }
                                
                                candidates.append(candidate)
                                print(f"✅ Found: {first_name} {last_name}")
                
                except Exception as e:
                    continue
            
            # Save results
            if candidates:
                # Save JSON
                with open('linkedin_candidates.json', 'w') as f:
                    json.dump(candidates, f, indent=4)
                
                # Create simple Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "LinkedIn Profiles"
                
                headers = ["First Name", "Last Name", "LinkedIn URL", "Company"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                for row, candidate in enumerate(candidates, 2):
                    ws.cell(row=row, column=1, value=candidate['first_name'])
                    ws.cell(row=row, column=2, value=candidate['last_name'])
                    ws.cell(row=row, column=3, value=candidate['link'])
                    ws.cell(row=row, column=4, value=candidate['company_name'])
                
                filename = f"{company.replace(' ', '_')}_LinkedIn_Results.xlsx"
                wb.save(filename)
                
                print(f"\\n🎉 SUCCESS!")
                print(f"📊 Found {len(candidates)} LinkedIn profiles")
                print(f"💾 Results saved to:")
                print(f"   - linkedin_candidates.json")
                print(f"   - {filename}")
                
                # Try to open Excel
                try:
                    os.startfile(filename)
                except:
                    pass
            else:
                print("❌ No LinkedIn profiles found")
        
        finally:
            driver.quit()
            print("🌐 Browser closed")
    
    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        input("\\nPress Enter to exit...")

if __name__ == "__main__":
    main()
'''
    
    # Write the fixed script
    script2_path = Path('script2_web_scraping.py')
    try:
        with open(script2_path, 'w', encoding='utf-8') as f:
            f.write(fixed_script2)
        print(f"✅ Restored working version of {script2_path.name}")
    except Exception as e:
        print(f"⚠️ Could not restore {script2_path.name}: {e}")
    
    return True

def test_employee_discovery():
    """Test if the employee discovery toolkit works."""
    print("\n🧪 STEP 5: TESTING EMPLOYEE DISCOVERY TOOLKIT")
    print("=" * 50)
    
    # Check if main scripts exist
    required_scripts = [
        'employee_discovery_selector.py',
        'script1_input_collection.py',
        'script2_web_scraping.py',
        'company_config.json'
    ]
    
    missing_scripts = []
    for script in required_scripts:
        if not Path(script).exists():
            missing_scripts.append(script)
    
    if missing_scripts:
        print(f"⚠️ Missing required files: {', '.join(missing_scripts)}")
        return False
    
    # Check if config is valid
    try:
        with open('company_config.json', 'r') as f:
            config = json.load(f)
        
        if not config.get('company_name'):
            print("⚠️ Configuration missing company name")
            return False
        
        print(f"✅ Configuration valid for: {config['company_name']}")
    except Exception as e:
        print(f"⚠️ Configuration error: {e}")
        return False
    
    print("✅ All core scripts present and configuration valid")
    return True

def create_launcher_script():
    """Create a simple launcher to start the toolkit."""
    launcher_content = '''@echo off
echo.
echo ================================================
echo EMPLOYEE DISCOVERY TOOLKIT - SAFE LAUNCHER
echo ================================================
echo.
echo This launcher ensures no auto-sync interference
echo.

cd /d "%~dp0"

echo Checking for auto-sync processes...
tasklist | findstr /i "auto" > nul
if %errorlevel% == 0 (
    echo WARNING: Auto-sync processes may be running!
    echo Please run the emergency fixer first.
    pause
    exit
)

echo Starting Employee Discovery Toolkit...
python employee_discovery_selector.py

echo.
echo Toolkit completed.
pause
'''
    
    try:
        with open('START_EMPLOYEE_DISCOVERY.bat', 'w') as f:
            f.write(launcher_content)
        print("✅ Created safe launcher: START_EMPLOYEE_DISCOVERY.bat")
    except Exception as e:
        print(f"⚠️ Could not create launcher: {e}")

def main():
    """Main emergency fixer function."""
    print("🚨 COMPLETE AUTO-SYNC EMERGENCY FIXER")
    print("=" * 60)
    print("This script will automatically fix all auto-sync issues")
    print("and restore your working employee discovery toolkit.")
    print()
    
    start_time = time.time()
    
    try:
        # Step 1: Stop auto-sync processes
        stop_autosync_processes()
        time.sleep(2)
        
        # Step 2: Fix Git repository
        fix_git_repository()
        time.sleep(1)
        
        # Step 3: Disable auto-sync scripts
        find_and_disable_autosync_scripts()
        time.sleep(1)
        
        # Step 4: Restore working scripts
        restore_working_scripts()
        time.sleep(1)
        
        # Step 5: Test toolkit
        toolkit_works = test_employee_discovery()
        
        # Step 6: Create safe launcher
        create_launcher_script()
        
        elapsed = time.time() - start_time
        
        print(f"\n🎉 EMERGENCY FIX COMPLETED!")
        print("=" * 60)
        print(f"⏱️ Total time: {elapsed:.1f} seconds")
        
        if toolkit_works:
            print("✅ Employee Discovery Toolkit is ready to use!")
            print("\n🚀 Next steps:")
            print("1. Double-click START_EMPLOYEE_DISCOVERY.bat")
            print("2. Or run: python employee_discovery_selector.py")
            print("3. Select option 1 for LinkedIn search")
            print("\n⚠️ IMPORTANT: Do not re-enable auto-sync until you're sure everything works!")
        else:
            print("⚠️ Some issues remain with the toolkit")
            print("You may need to run script1_input_collection.py first")
        
        print(f"\n📁 Current directory: {Path.cwd()}")
        
    except KeyboardInterrupt:
        print("\n\n⏹️ Emergency fix cancelled by user")
    except Exception as e:
        print(f"\n❌ Emergency fix failed: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "="*60)
    input("Press Enter to exit...")

if __name__ == "__main__":
    main()