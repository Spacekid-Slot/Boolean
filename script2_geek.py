#!/usr/bin/env python3
"""
Enhanced Recruitment Geek Tool - WITH SITE-SPECIFIC DOMAINS + LOCATION

This script uses the actual Recruitment Geek LinkedIn search tool at:
https://recruitmentgeek.com/tools/linkedin

ENHANCED IMPROVEMENTS:
- Uses site:uk.linkedin.com, site:fr.linkedin.com etc. in Boolean queries
- Combines domain targeting with location terms for better filtering
- Works around Recruitment Geek's poor location filter
- Enhanced data capture with better validation
- Country-specific LinkedIn domain integration
"""

import json
import logging
import os
import random
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict, Optional, Set
from urllib.parse import quote_plus, urljoin
from dataclasses import dataclass
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def install_dependencies():
    """Install required packages if not available."""
    required_packages = ['selenium', 'webdriver-manager', 'openpyxl']
    
    for package in required_packages:
        try:
            if package == 'webdriver-manager':
                import webdriver_manager
            elif package == 'openpyxl':
                import openpyxl
            else:
                __import__(package)
        except ImportError:
            print(f"Installing {package}...")
            subprocess.call([sys.executable, "-m", "pip", "install", package])

install_dependencies()

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import Select
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, 
        WebDriverException, SessionNotCreatedException,
        ElementNotInteractableException, StaleElementReferenceException
    )
    from webdriver_manager.chrome import ChromeDriverManager
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"Failed to import required packages: {e}")
    sys.exit(1)

@dataclass
class LinkedInCandidate:
    """LinkedIn profile candidate from Enhanced Recruitment Geek tool."""
    first_name: str
    last_name: str
    linkedin_url: str
    company_name: str
    location: str
    confidence: str
    source: str
    extraction_date: str
    linkedin_domain: str
    search_query: str

class EnhancedRecruitmentGeekScraper:
    """Enhanced Recruitment Geek scraper with site-specific domains + location targeting."""
    
    def __init__(self, config_path: str):
        """Initialize the enhanced scraper."""
        self.config = self._load_config(config_path)
        self.script_dir = Path(__file__).parent.absolute()
        self.driver = None
        self.processed_urls: Set[str] = set()
        self.job_titles = self._get_job_titles_to_search()
        
        # Enhanced LinkedIn domain mapping
        self.linkedin_domains = self._get_linkedin_domains_mapping()
        
        # Determine the target LinkedIn domain and location terms
        self.target_domain, self.location_terms = self._determine_target_domain_and_location()
        
        # Recruitment Geek tool URL
        self.recruitment_geek_url = "https://recruitmentgeek.com/tools/linkedin#gsc.tab=0"
        
        # User agents
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0'
        ]
        
        logger.info(f"Enhanced Recruitment Geek initialized")
        logger.info(f"Target domain: {self.target_domain}")
        logger.info(f"Location terms: {self.location_terms}")
        
    def _load_config(self, config_path: str) -> dict:
        """Load configuration from JSON file."""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Configuration error: {e}")
            sys.exit(1)
    
    def _get_linkedin_domains_mapping(self) -> Dict[str, str]:
        """Get mapping of countries/regions to LinkedIn domains."""
        return {
            # Major English-speaking markets
            'uk': 'uk.linkedin.com',
            'united kingdom': 'uk.linkedin.com',
            'england': 'uk.linkedin.com',
            'scotland': 'uk.linkedin.com',
            'wales': 'uk.linkedin.com',
            'northern ireland': 'uk.linkedin.com',
            'britain': 'uk.linkedin.com',
            'great britain': 'uk.linkedin.com',
            
            # UK Cities
            'london': 'uk.linkedin.com',
            'edinburgh': 'uk.linkedin.com',
            'glasgow': 'uk.linkedin.com',
            'manchester': 'uk.linkedin.com',
            'birmingham': 'uk.linkedin.com',
            'bristol': 'uk.linkedin.com',
            'leeds': 'uk.linkedin.com',
            'liverpool': 'uk.linkedin.com',
            'sheffield': 'uk.linkedin.com',
            'cardiff': 'uk.linkedin.com',
            'belfast': 'uk.linkedin.com',
            
            # Canada
            'canada': 'ca.linkedin.com',
            'toronto': 'ca.linkedin.com',
            'vancouver': 'ca.linkedin.com',
            'montreal': 'ca.linkedin.com',
            'calgary': 'ca.linkedin.com',
            'ottawa': 'ca.linkedin.com',
            'quebec': 'ca.linkedin.com',
            
            # Australia
            'australia': 'au.linkedin.com',
            'sydney': 'au.linkedin.com',
            'melbourne': 'au.linkedin.com',
            'brisbane': 'au.linkedin.com',
            'perth': 'au.linkedin.com',
            'adelaide': 'au.linkedin.com',
            'canberra': 'au.linkedin.com',
            
            # Major European markets
            'france': 'fr.linkedin.com',
            'paris': 'fr.linkedin.com',
            'lyon': 'fr.linkedin.com',
            'marseille': 'fr.linkedin.com',
            
            'germany': 'de.linkedin.com',
            'berlin': 'de.linkedin.com',
            'munich': 'de.linkedin.com',
            'hamburg': 'de.linkedin.com',
            'cologne': 'de.linkedin.com',
            'frankfurt': 'de.linkedin.com',
            
            'spain': 'es.linkedin.com',
            'madrid': 'es.linkedin.com',
            'barcelona': 'es.linkedin.com',
            
            'italy': 'it.linkedin.com',
            'rome': 'it.linkedin.com',
            'milan': 'it.linkedin.com',
            
            'netherlands': 'nl.linkedin.com',
            'amsterdam': 'nl.linkedin.com',
            'rotterdam': 'nl.linkedin.com',
            
            # Other markets
            'india': 'in.linkedin.com',
            'mumbai': 'in.linkedin.com',
            'delhi': 'in.linkedin.com',
            'bangalore': 'in.linkedin.com',
            'chennai': 'in.linkedin.com',
            'hyderabad': 'in.linkedin.com',
            'pune': 'in.linkedin.com',
            
            'brazil': 'br.linkedin.com',
            'sao paulo': 'br.linkedin.com',
            'rio de janeiro': 'br.linkedin.com',
            
            'japan': 'jp.linkedin.com',
            'tokyo': 'jp.linkedin.com',
            'osaka': 'jp.linkedin.com',
            
            'singapore': 'sg.linkedin.com',
            'hong kong': 'sg.linkedin.com',
        }
    
    def _determine_target_domain_and_location(self) -> tuple:
        """Determine the target LinkedIn domain and location search terms."""
        try:
            # Get location from config
            location = self.config.get('location', '').lower().strip()
            
            # Check if we have a location config with more details
            location_config = self.config.get('location_config', {})
            
            # Start with the main location
            location_terms = [self.config.get('location', '')]
            
            # Check for country-specific domain
            target_domain = 'linkedin.com'  # Default
            
            if location_config and not location_config.get('manual_input', True):
                # Location was selected from database
                country = location_config.get('country', '').lower()
                city = location_config.get('city', '').lower()
                
                # Check for domain mapping
                if country in self.linkedin_domains:
                    target_domain = self.linkedin_domains[country]
                elif city in self.linkedin_domains:
                    target_domain = self.linkedin_domains[city]
                
                # Add location variations from config
                if location_config.get('location_variations'):
                    location_terms.extend(location_config['location_variations'])
                
                if location_config.get('country_terms'):
                    location_terms.extend(location_config['country_terms'])
                    
            else:
                # Manual location entry - try to detect domain
                location_words = location.split()
                for word in location_words:
                    if word in self.linkedin_domains:
                        target_domain = self.linkedin_domains[word]
                        break
                
                # For manual entry, also check the full location string
                if location in self.linkedin_domains:
                    target_domain = self.linkedin_domains[location]
            
            # Clean up location terms and remove duplicates
            clean_location_terms = []
            for term in location_terms:
                if term and term.strip():
                    clean_term = term.strip()
                    if clean_term not in clean_location_terms:
                        clean_location_terms.append(clean_term)
            
            # If no location terms, add the main location
            if not clean_location_terms:
                clean_location_terms = [self.config.get('location', '')]
            
            return target_domain, clean_location_terms
            
        except Exception as e:
            logger.error(f"Error determining target domain: {e}")
            return 'linkedin.com', [self.config.get('location', '')]
    
    def _get_job_titles_to_search(self) -> List[str]:
        """Get job titles from config."""
        if 'job_titles' in self.config and self.config['job_titles']:
            return self.config['job_titles']
        return []
    
    def _setup_browser(self) -> Optional[webdriver.Chrome]:
        """Setup browser for Enhanced Recruitment Geek tool."""
        try:
            print("Setting up browser for Enhanced Recruitment Geek tool...")
            
            chrome_options = Options()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-extensions')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument(f'user-agent={random.choice(self.user_agents)}')
            chrome_options.add_argument('--window-size=1400,1000')
            
            # Add preferences for downloads if needed
            chrome_options.add_experimental_option('prefs', {
                'profile.default_content_setting_values.notifications': 2,
                'profile.default_content_settings.popups': 0,
                'profile.managed_default_content_settings.images': 2
            })
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(30)
            driver.implicitly_wait(10)
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            print("✅ Browser setup successful for Enhanced Recruitment Geek")
            return driver
            
        except Exception as e:
            print(f"❌ Enhanced browser setup error: {e}")
            return None
    
    def _navigate_to_recruitment_geek(self) -> bool:
        """Navigate to the Recruitment Geek LinkedIn tool."""
        try:
            print(f"Navigating to Enhanced Recruitment Geek tool: {self.recruitment_geek_url}")
            self.driver.get(self.recruitment_geek_url)
            
            # Wait for the page to load
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # Look for the LinkedIn search form elements
            # The actual form elements will depend on the current website structure
            time.sleep(3)  # Allow page to fully render
            
            print("✅ Successfully navigated to Enhanced Recruitment Geek tool")
            return True
            
        except Exception as e:
            print(f"❌ Error navigating to Enhanced Recruitment Geek: {e}")
            return False
    
    def _paste_enhanced_boolean_search(self, enhanced_query: str) -> bool:
        """Paste the enhanced Boolean search string with site: domains into Recruitment Geek."""
        try:
            # Wait a bit for any dynamic content to load
            time.sleep(2)
            
            # Look for the main search input box with enhanced selectors
            search_selectors = [
                'input[type="text"]',
                'input[type="search"]',
                'textarea',
                'input.form-control',
                'input[placeholder*="search"]',
                'input[placeholder*="skills"]',
                'input[placeholder*="location"]',
                'input[placeholder*="keyword"]',
                'input[placeholder*="query"]',
                '#search',
                '.search-input',
                'input[name="q"]',
                'input[name="search"]',
                'input[name="keywords"]'
            ]
            
            search_field = None
            for selector in search_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    # Get the largest visible input field (likely the main search box)
                    for element in elements:
                        if element.is_displayed() and element.is_enabled():
                            # Check if it's likely the main search field
                            field_size = element.size
                            if field_size['width'] > 100 and field_size['height'] > 20:
                                search_field = element
                                break
                    if search_field:
                        break
                except NoSuchElementException:
                    continue
            
            if search_field:
                # Clear the field and paste the enhanced Boolean query
                search_field.clear()
                time.sleep(0.5)
                search_field.send_keys(enhanced_query)
                print(f"✅ Pasted Enhanced Boolean query: {enhanced_query[:120]}...")
                time.sleep(1)
                return True
            else:
                print("⚠️ Could not find search input field for enhanced query")
                return False
            
        except Exception as e:
            print(f"❌ Error pasting enhanced Boolean search: {e}")
            return False
    
    def _submit_enhanced_search(self) -> bool:
        """Submit the enhanced search by pressing Enter or clicking search button."""
        try:
            # First try to press Enter in the search field
            try:
                search_field = self.driver.find_element(By.CSS_SELECTOR, 'input[type="text"], input[type="search"], textarea')
                search_field.send_keys(Keys.RETURN)
                print("✅ Enhanced search submitted via Enter key")
                time.sleep(3)
                return True
            except:
                pass
            
            # If Enter doesn't work, look for search/submit buttons with enhanced selectors
            submit_selectors = [
                'button[type="submit"]',
                'input[type="submit"]',
                'button[class*="search"]',
                'button[id*="search"]',
                'button[class*="btn"]',
                'button.btn',
                'button[class*="submit"]',
                '.search-button',
                'button:contains("Search")',
                'button:contains("Go")',
                'input[value*="Search"]',
                'input[value*="Go"]'
            ]
            
            submit_button = None
            for selector in submit_selectors:
                try:
                    submit_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    # Check if button is visible and clickable
                    if submit_button.is_displayed() and submit_button.is_enabled():
                        break
                except NoSuchElementException:
                    continue
            
            if submit_button:
                # Try clicking the button
                try:
                    submit_button.click()
                except:
                    # If regular click fails, try JavaScript click
                    self.driver.execute_script("arguments[0].click();", submit_button)
                
                print("✅ Enhanced search submitted via button click")
                time.sleep(5)  # Wait longer for results to load
                return True
            else:
                print("⚠️ Could not find search button, enhanced search may have been submitted via Enter")
                return True
            
        except Exception as e:
            print(f"❌ Error submitting enhanced search: {e}")
            return False
    
    def _extract_enhanced_results_from_page(self, search_query: str) -> List[LinkedInCandidate]:
        """Extract LinkedIn profile results with enhanced validation and domain tracking."""
        candidates = []
        
        try:
            # Wait for results to appear
            time.sleep(3)
            
            # Look for LinkedIn profile links in results with enhanced selectors
            link_selectors = [
                'a[href*="linkedin.com/in/"]',
                'a[href*="www.linkedin.com/in/"]',
                f'a[href*="{self.target_domain}/in/"]',
                f'a[href*="www.{self.target_domain}/in/"]',
                'a[href*=".linkedin.com/in/"]'  # Catch any LinkedIn domain
            ]
            
            profile_links = []
            for selector in link_selectors:
                try:
                    links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    profile_links.extend(links)
                except:
                    continue
            
            # Remove duplicates by href
            unique_links = {}
            for link in profile_links:
                href = link.get_attribute('href')
                if href and href not in unique_links:
                    unique_links[href] = link
            
            profile_links = list(unique_links.values())
            print(f"Found {len(profile_links)} potential LinkedIn profile links")
            
            for link in profile_links:
                try:
                    href = link.get_attribute('href')
                    if not href or href in self.processed_urls:
                        continue
                    
                    # Enhanced validation - check if URL matches our search criteria
                    if not self._validate_linkedin_url(href):
                        continue
                    
                    # Extract text that might contain the name
                    link_text = link.text.strip()
                    
                    # Try to get name from parent elements or nearby text with enhanced extraction
                    name_text = self._extract_name_context(link, link_text)
                    
                    # Try to extract name from the link text or nearby elements
                    name_data = self._extract_enhanced_name_from_text(name_text, href)
                    
                    if name_data:
                        # Extract LinkedIn domain from URL
                        extracted_domain = self._extract_linkedin_domain_from_url(href)
                        
                        candidate = LinkedInCandidate(
                            first_name=name_data['first_name'],
                            last_name=name_data['last_name'],
                            linkedin_url=href,
                            company_name=self.config.get('company_name', ''),
                            location=self.config.get('location', ''),
                            confidence=self._determine_enhanced_confidence(name_text, href, search_query),
                            source=f'Enhanced Recruitment Geek ({extracted_domain})',
                            extraction_date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            linkedin_domain=extracted_domain,
                            search_query=search_query
                        )
                        
                        candidates.append(candidate)
                        self.processed_urls.add(href)
                        print(f"✅ Enhanced extraction: {candidate.first_name} {candidate.last_name} ({extracted_domain})")
                
                except Exception as e:
                    logger.debug(f"Error processing enhanced link: {e}")
                    continue
            
        except Exception as e:
            print(f"❌ Error extracting enhanced results: {e}")
        
        return candidates
    
    def _validate_linkedin_url(self, url: str) -> bool:
        """Enhanced LinkedIn URL validation."""
        try:
            url_lower = url.lower()
            
            # Must contain linkedin.com/in/
            if '/in/' not in url_lower or 'linkedin.com' not in url_lower:
                return False
            
            # Should not be a company page or other LinkedIn page type
            invalid_patterns = [
                '/company/', '/school/', '/groups/', '/events/',
                '/jobs/', '/feed/', '/notifications/', '/messaging/'
            ]
            
            for pattern in invalid_patterns:
                if pattern in url_lower:
                    return False
            
            return True
            
        except Exception:
            return False
    
    def _extract_name_context(self, link_element, link_text: str) -> str:
        """Extract enhanced name context from link and surrounding elements."""
        try:
            name_text = link_text
            
            # If link text is empty or too short, try parent elements
            if not name_text or len(name_text) < 5:
                try:
                    # Try immediate parent
                    parent = link_element.find_element(By.XPATH, './..')
                    parent_text = parent.text.strip()
                    if len(parent_text) > len(name_text):
                        name_text = parent_text
                except:
                    pass
                
                # Try grandparent if still not good
                if not name_text or len(name_text) < 5:
                    try:
                        grandparent = link_element.find_element(By.XPATH, './../..')
                        grandparent_text = grandparent.text.strip()
                        if len(grandparent_text) > len(name_text) and len(grandparent_text) < 200:
                            name_text = grandparent_text
                    except:
                        pass
            
            # Try to get text from adjacent elements
            if not name_text or len(name_text) < 5:
                try:
                    # Try previous sibling
                    prev_sibling = link_element.find_element(By.XPATH, './preceding-sibling::*[1]')
                    prev_text = prev_sibling.text.strip()
                    if prev_text and len(prev_text) < 100:
                        name_text = prev_text + ' ' + name_text
                except:
                    pass
                
                try:
                    # Try next sibling
                    next_sibling = link_element.find_element(By.XPATH, './following-sibling::*[1]')
                    next_text = next_sibling.text.strip()
                    if next_text and len(next_text) < 100:
                        name_text = name_text + ' ' + next_text
                except:
                    pass
            
            return name_text.strip()
            
        except Exception as e:
            logger.debug(f"Error extracting name context: {e}")
            return link_text
    
    def _extract_enhanced_name_from_text(self, text: str, url: str) -> Optional[Dict]:
        """Enhanced name extraction with better validation."""
        try:
            if not text:
                return None
            
            # Clean up the text
            text = re.sub(r'\s+', ' ', text).strip()
            
            # Enhanced name patterns
            name_patterns = [
                # Standard "FirstName LastName" pattern
                r'^([A-Z][a-z]+)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',
                # "FirstName LastName" anywhere in text
                r'\b([A-Z][a-z]+)\s+([A-Z][a-z]+)\b',
                # Handle names with middle initials
                r'\b([A-Z][a-z]+)\s+[A-Z]\.\s+([A-Z][a-z]+)\b',
                # Handle hyphenated names
                r'\b([A-Z][a-z]+-[A-Z][a-z]+)\s+([A-Z][a-z]+)\b',
                r'\b([A-Z][a-z]+)\s+([A-Z][a-z]+-[A-Z][a-z]+)\b',
            ]
            
            for pattern in name_patterns:
                matches = re.finditer(pattern, text)
                for match in matches:
                    first_name = match.group(1).strip()
                    last_name = match.group(2).strip()
                    
                    # Enhanced name validation
                    if (self._is_enhanced_valid_name_part(first_name) and 
                        self._is_enhanced_valid_name_part(last_name)):
                        return {
                            'first_name': first_name,
                            'last_name': last_name
                        }
            
            return None
            
        except Exception as e:
            logger.debug(f"Error extracting enhanced name from text: {e}")
            return None
    
    def _is_enhanced_valid_name_part(self, name_part: str) -> bool:
        """Enhanced name part validation."""
        if not name_part or len(name_part) < 2 or len(name_part) > 30:
            return False
        
        # Must contain only letters, hyphens, and apostrophes
        if not re.match(r"^[a-zA-Z\-']+$", name_part):
            return False
        
        # Enhanced false positives list
        false_positives = {
            'linkedin', 'profile', 'company', 'limited', 'group', 'ltd',
            'corporation', 'corp', 'inc', 'llc', 'business', 'enterprise',
            'consulting', 'services', 'solutions', 'management', 'director',
            'manager', 'executive', 'president', 'officer', 'employee',
            'recruitment', 'geek', 'search', 'results', 'view', 'more',
            'connect', 'follow', 'message', 'contact', 'about', 'experience',
            'education', 'skills', 'recommendations', 'activity', 'interests'
        }
        
        if name_part.lower() in false_positives:
            return False
        
        return True
    
    def _extract_linkedin_domain_from_url(self, url: str) -> str:
        """Extract LinkedIn domain from URL."""
        try:
            domain_match = re.search(r'((?:[a-z]{2}\.)?linkedin\.com)', url.lower())
            if domain_match:
                return domain_match.group(1)
            return 'linkedin.com'
        except:
            return 'linkedin.com'
    
    def _determine_enhanced_confidence(self, name_text: str, url: str, search_query: str) -> str:
        """Enhanced confidence determination based on multiple factors."""
        score = 0
        
        # Base score for enhanced search with site: domains
        if f'site:{self.target_domain}' in search_query:
            score += 3  # Bonus for targeted domain search
        
        # URL domain matching
        url_domain = self._extract_linkedin_domain_from_url(url)
        if url_domain == self.target_domain:
            score += 2  # Matches target domain
        elif url_domain != 'linkedin.com':
            score += 1  # Country-specific domain but not target
        
        # Company name presence in context
        company = self.config['company_name']
        if company.lower() in name_text.lower():
            score += 3
        
        # Location term presence
        for location_term in self.location_terms:
            if location_term.lower() in name_text.lower():
                score += 1
                break
        
        # Job title presence (if we have job titles to search)
        if self.job_titles:
            for job_title in self.job_titles:
                if job_title.lower() in name_text.lower():
                    score += 2
                    break
        
        # URL quality indicators
        if len(url.split('/')) >= 5:  # Proper LinkedIn profile URL structure
            score += 1
        
        # Enhanced confidence thresholds
        if score >= 6:
            return 'high'
        elif score >= 3:
            return 'medium'
        else:
            return 'low'
    
    def search_with_enhanced_recruitment_geek(self) -> List[Dict]:
        """Main enhanced search method using site-specific domains + location targeting."""
        company_name = self.config['company_name']
        all_candidates = []
        
        print(f"Starting Enhanced Recruitment Geek search for {company_name}")
        print(f"Using site:{self.target_domain} + location targeting")
        
        try:
            self.driver = self._setup_browser()
            if not self.driver:
                return []
            
            # Navigate to Recruitment Geek
            if not self._navigate_to_recruitment_geek():
                return []
            
            # Create enhanced Boolean search queries with site: domains
            enhanced_queries = self._create_enhanced_boolean_queries()
            
            for i, query_info in enumerate(enhanced_queries, 1):
                print(f"\n--- Enhanced Boolean Search {i}/{len(enhanced_queries)} ---")
                print(f"Query: {query_info['query']}")
                print(f"Type: {query_info['description']}")
                
                # Paste enhanced Boolean search string and submit
                if self._paste_enhanced_boolean_search(query_info['query']):
                    if self._submit_enhanced_search():
                        # Extract results with enhanced validation
                        page_candidates = self._extract_enhanced_results_from_page(query_info['query'])
                        all_candidates.extend(page_candidates)
                        print(f"Found {len(page_candidates)} candidates from enhanced search")
                        
                        # Navigate back for next search if needed
                        if i < len(enhanced_queries):
                            # Go back to the main page for next search
                            self.driver.get(self.recruitment_geek_url)
                            time.sleep(3)
                    else:
                        print("Failed to submit enhanced search")
                else:
                    print("Failed to paste enhanced Boolean search")
                
                # Add delay between searches
                time.sleep(random.uniform(3, 5))
                
                # Break if we have enough results
                if len(all_candidates) >= 50:  # Reasonable limit
                    print(f"Reached candidate limit, stopping enhanced search")
                    break
        
        except Exception as e:
            print(f"❌ Enhanced search process error: {e}")
        
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                    print("✅ Browser closed")
                except:
                    pass
        
        print(f"Enhanced Recruitment Geek search completed. Found {len(all_candidates)} total candidates")
        return [candidate.__dict__ for candidate in all_candidates]
    
    def _create_enhanced_boolean_queries(self) -> List[Dict]:
        """Create enhanced Boolean queries with site-specific domains + location terms."""
        company = self.config['company_name']
        queries = []
        
        if self.job_titles:
            print(f"Creating enhanced targeted searches for {len(self.job_titles)} job titles...")
            print(f"Using site:{self.target_domain} with location terms: {self.location_terms}")
            
            # Job title specific searches with site: domain + location
            for title in self.job_titles:
                for location_term in self.location_terms[:2]:  # Use top 2 location variations
                    enhanced_queries = [
                        {
                            'query': f'site:{self.target_domain} "{company}" "{title}" "{location_term}"',
                            'description': f'Enhanced Domain+Location: {title} at {company} in {location_term}',
                            'type': 'enhanced_title_location'
                        },
                        {
                            'query': f'site:{self.target_domain} "{title}" at "{company}" "{location_term}"',
                            'description': f'Enhanced At-Company: {title} at {company} ({location_term})',
                            'type': 'enhanced_title_at_company'
                        }
                    ]
                    queries.extend(enhanced_queries)
                
                # Domain-only queries for broader reach
                domain_queries = [
                    {
                        'query': f'site:{self.target_domain} "{company}" "{title}"',
                        'description': f'Enhanced Domain-Only: {title} at {company}',
                        'type': 'enhanced_title_domain'
                    }
                ]
                queries.extend(domain_queries)
            
            # Add enhanced general executive queries
            for location_term in self.location_terms[:2]:
                exec_queries = [
                    {
                        'query': f'site:{self.target_domain} "{company}" (Director OR Manager) "{location_term}"',
                        'description': f'Enhanced Leadership: Directors/Managers at {company} in {location_term}',
                        'type': 'enhanced_executive'
                    },
                    {
                        'query': f'site:{self.target_domain} "{company}" (CEO OR President OR "Vice President") "{location_term}"',
                        'description': f'Enhanced Senior Exec: C-Suite at {company} in {location_term}',
                        'type': 'enhanced_senior_exec'
                    }
                ]
                queries.extend(exec_queries)
            
        else:
            print(f"Creating enhanced general company searches...")
            print(f"Using site:{self.target_domain} with location terms: {self.location_terms}")
            
            # General company searches with enhanced targeting
            for location_term in self.location_terms:
                general_queries = [
                    {
                        'query': f'site:{self.target_domain} "{company}" "{location_term}"',
                        'description': f'Enhanced General: {company} employees in {location_term}',
                        'type': 'enhanced_company_location'
                    },
                    {
                        'query': f'site:{self.target_domain} "{company}" employee "{location_term}"',
                        'description': f'Enhanced Employee Search: {company} employees ({location_term})',
                        'type': 'enhanced_employee_search'
                    },
                    {
                        'query': f'site:{self.target_domain} works at "{company}" "{location_term}"',
                        'description': f'Enhanced Works-At: Works at {company} ({location_term})',
                        'type': 'enhanced_works_at'
                    }
                ]
                queries.extend(general_queries)
            
            # Add domain-wide queries for maximum coverage
            domain_queries = [
                {
                    'query': f'site:{self.target_domain} "{company}"',
                    'description': f'Enhanced Domain-Wide: {company} (all {self.target_domain})',
                    'type': 'enhanced_domain_wide'
                },
                {
                    'query': f'site:{self.target_domain} "{company}" (Director OR Manager)',
                    'description': f'Enhanced Management: {company} Leadership ({self.target_domain})',
                    'type': 'enhanced_management'
                }
            ]
            queries.extend(domain_queries)
        
        print(f"Created {len(queries)} enhanced Boolean queries for Recruitment Geek")
        return queries
    
    def save_enhanced_results(self, candidates: List[Dict]) -> bool:
        """Save enhanced results to files."""
        try:
            # Convert to standard employee format for compatibility
            employees = []
            for candidate in candidates:
                employee = {
                    'first_name': candidate.get('first_name', ''),
                    'last_name': candidate.get('last_name', ''),
                    'title': 'LinkedIn Profile',  # We don't extract titles from Recruitment Geek
                    'link': candidate.get('linkedin_url', ''),
                    'company_name': candidate.get('company_name', ''),
                    'location': candidate.get('location', ''),
                    'source': candidate.get('source', ''),
                    'confidence': candidate.get('confidence', 'medium'),
                    'linkedin_domain': candidate.get('linkedin_domain', ''),
                    'search_query': candidate.get('search_query', ''),
                    'extraction_date': candidate.get('extraction_date', ''),
                    'needs_verification': True
                }
                employees.append(employee)
            
            # Save to enhanced-specific file
            enhanced_file = self.script_dir / "enhanced_recruitment_geek_results.json"
            with open(enhanced_file, 'w', encoding='utf-8') as f:
                json.dump(employees, f, indent=4, ensure_ascii=False)
            
            # Also save to standard merged file for compatibility
            merged_file = self.script_dir / "merged_employees.json"
            if merged_file.exists():
                with open(merged_file, 'r', encoding='utf-8') as f:
                    existing = json.load(f)
                
                existing_urls = {emp.get('link', '') for emp in existing}
                new_count = 0
                for emp in employees:
                    if emp.get('link') and emp['link'] not in existing_urls:
                        existing.append(emp)
                        new_count += 1
                
                with open(merged_file, 'w', encoding='utf-8') as f:
                    json.dump(existing, f, indent=4, ensure_ascii=False)
                
                print(f"Added {new_count} new enhanced candidates to merged file")
            else:
                with open(merged_file, 'w', encoding='utf-8') as f:
                    json.dump(employees, f, indent=4, ensure_ascii=False)
            
            print(f"Enhanced results saved: {len(employees)} candidates")
            return True
            
        except Exception as e:
            print(f"Error saving enhanced results: {e}")
            return False
    
    def create_enhanced_excel_report(self, candidates: List[Dict]) -> bool:
        """Create enhanced Excel report with domain and query tracking."""
        try:
            if not candidates:
                print("⚠️ No candidates for enhanced Excel report")
                return False
            
            print("📊 Creating Enhanced Recruitment Geek Excel report...")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Enhanced Recruitment Geek"
            
            # Headers
            headers = [
                "First Name", "Last Name", "LinkedIn URL", "LinkedIn Domain", 
                "Company", "Location", "Confidence", "Source", "Search Query", "Extraction Date"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row, candidate in enumerate(candidates, 2):
                ws.cell(row=row, column=1, value=candidate.get('first_name', ''))
                ws.cell(row=row, column=2, value=candidate.get('last_name', ''))
                
                # LinkedIn URL with hyperlink
                url = candidate.get('linkedin_url', '')
                url_cell = ws.cell(row=row, column=3, value=url)
                if url:
                    url_cell.hyperlink = url
                    url_cell.font = Font(color="0000FF", underline="single")
                
                # LinkedIn domain with highlighting for country-specific domains
                domain_cell = ws.cell(row=row, column=4, value=candidate.get('linkedin_domain', ''))
                if candidate.get('linkedin_domain', '') != 'linkedin.com':
                    domain_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                ws.cell(row=row, column=5, value=candidate.get('company_name', ''))
                ws.cell(row=row, column=6, value=candidate.get('location', ''))
                
                # Confidence with colors
                conf_cell = ws.cell(row=row, column=7, value=candidate.get('confidence', '').upper())
                if conf_cell.value == 'HIGH':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif conf_cell.value == 'MEDIUM':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row, column=8, value=candidate.get('source', ''))
                ws.cell(row=row, column=9, value=candidate.get('search_query', ''))
                ws.cell(row=row, column=10, value=candidate.get('extraction_date', ''))
            
            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 60)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet(title="Enhancement Summary")
            
            # Calculate enhanced statistics
            domain_counts = {}
            confidence_counts = {}
            query_type_counts = {}
            
            for candidate in candidates:
                domain = candidate.get('linkedin_domain', 'unknown')
                domain_counts[domain] = domain_counts.get(domain, 0) + 1
                
                conf = candidate.get('confidence', 'unknown')
                confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
                
                # Extract query type from search query
                query = candidate.get('search_query', '')
                if f'site:{self.target_domain}' in query:
                    query_type_counts['enhanced_domain'] = query_type_counts.get('enhanced_domain', 0) + 1
                else:
                    query_type_counts['standard'] = query_type_counts.get('standard', 0) + 1
            
            summary_data = [
                ["Enhanced Recruitment Geek Results", ""],
                ["", ""],
                ["Company", self.config.get('company_name', '')],
                ["Location", self.config.get('location', '')],
                ["Target LinkedIn Domain", self.target_domain],
                ["Location Terms Used", ', '.join(self.location_terms)],
                ["Total Candidates Found", len(candidates)],
                ["Enhancement Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["", ""],
                ["LinkedIn Domain Breakdown", ""],
            ]
            
            for domain, count in sorted(domain_counts.items()):
                summary_data.append([f"  {domain}", count])
            
            summary_data.extend([
                ["", ""],
                ["Confidence Breakdown", ""],
            ])
            
            for conf, count in sorted(confidence_counts.items()):
                summary_data.append([f"  {conf.title()}", count])
            
            summary_data.extend([
                ["", ""],
                ["Enhancement Benefits", ""],
                ["✓ Site-specific domain targeting", f"Used site:{self.target_domain}"],
                ["✓ Location term integration", f"{len(self.location_terms)} location variations"],
                ["✓ Enhanced name validation", "Improved false positive filtering"],
                ["✓ Domain-aware confidence scoring", "Better result quality assessment"],
                ["", ""],
                ["Ready for LinkedIn Verification", "Yes - All profiles need verification"],
            ])
            
            for row, (label, value) in enumerate(summary_data, 1):
                cell1 = summary_sheet.cell(row=row, column=1, value=label)
                cell2 = summary_sheet.cell(row=row, column=2, value=value)
                
                if label and not label.startswith(('  ', '✓')):
                    cell1.font = Font(bold=True)
                    if label == "Enhanced Recruitment Geek Results":
                        cell1.font = Font(bold=True, size=14)
            
            summary_sheet.column_dimensions['A'].width = 35
            summary_sheet.column_dimensions['B'].width = 30
            
            # Save file
            company_clean = self.config.get('company_name', 'Company').replace(' ', '_')
            domain_suffix = self.target_domain.replace('.linkedin.com', '').replace('linkedin.com', 'global')
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            filename = f"{company_clean}_Enhanced_RecruitmentGeek_{domain_suffix}_{timestamp}.xlsx"
            filepath = self.script_dir / filename
            
            wb.save(filepath)
            
            if filepath.exists():
                print(f"✅ Enhanced Excel report created: {filename}")
                
                try:
                    if sys.platform == 'win32':
                        os.startfile(str(filepath))
                    elif sys.platform == 'darwin':
                        subprocess.call(['open', str(filepath)])
                    else:
                        subprocess.call(['xdg-open', str(filepath)])
                    print("📂 File opened automatically")
                except:
                    print(f"📂 Please open manually: {filepath}")
                
                return True
            else:
                print("❌ Enhanced Excel file creation failed")
                return False
                
        except Exception as e:
            print(f"❌ Enhanced Excel creation error: {e}")
            return False

def main():
    """Main execution for Enhanced Recruitment Geek tool."""
    try:
        script_dir = Path(__file__).parent.absolute()
        config_path = script_dir / "company_config.json"
        
        if not config_path.exists():
            print("❌ Configuration file not found.")
            print("Please run script1_input_collection.py first.")
            sys.exit(1)
        
        print("=" * 70)
        print("ENHANCED RECRUITMENT GEEK - SITE DOMAINS + LOCATION TARGETING")
        print("=" * 70)
        print("This enhanced version uses site:uk.linkedin.com style queries")
        print("combined with location terms to work around Recruitment Geek's")
        print("poor location filtering and provide much better targeted results.")
        
        scraper = EnhancedRecruitmentGeekScraper(str(config_path))
        
        print(f"\n📊 Enhanced Configuration:")
        print(f"Company: {scraper.config['company_name']}")
        print(f"Location: {scraper.config['location']}")
        print(f"Target LinkedIn Domain: {scraper.target_domain}")
        print(f"Location Search Terms: {', '.join(scraper.location_terms)}")
        print(f"Job titles: {len(scraper.job_titles)}")
        print(f"Tool URL: {scraper.recruitment_geek_url}")
        
        if scraper.job_titles:
            print(f"Search strategy: Enhanced company + job title targeting")
            if len(scraper.job_titles) <= 5:
                print(f"Titles: {', '.join(scraper.job_titles)}")
            else:
                print(f"Sample titles: {', '.join(scraper.job_titles[:3])}... (+{len(scraper.job_titles)-3} more)")
        else:
            print("Search strategy: Enhanced general company search")
        
        print(f"\n💡 Enhanced Recruitment Geek Benefits:")
        print(f"   ✅ Uses actual Recruitment Geek website interface")
        if scraper.target_domain != 'linkedin.com':
            print(f"   ✅ Site-specific domain targeting (site:{scraper.target_domain})")
        else:
            print(f"   ✅ Global LinkedIn targeting (site:linkedin.com)")
        print(f"   ✅ Location terms combined with domain filtering")
        print(f"   ✅ Works around Recruitment Geek's poor location filter")
        print(f"   ✅ Enhanced name validation and confidence scoring")
        print(f"   ✅ Better result quality through combined targeting")
        
        print(f"\n🔍 Sample Enhanced Boolean Queries:")
        sample_company = scraper.config.get('company_name', 'YourCompany')
        sample_location = scraper.location_terms[0] if scraper.location_terms else 'YourLocation'
        print(f'   site:{scraper.target_domain} "{sample_company}" "{sample_location}"')
        if scraper.job_titles:
            sample_title = scraper.job_titles[0]
            print(f'   site:{scraper.target_domain} "{sample_title}" at "{sample_company}" "{sample_location}"')
        print(f'   site:{scraper.target_domain} "{sample_company}" (Director OR Manager) "{sample_location}"')
        
        print(f"\n⚠️ Important Notes:")
        print(f"   • This script will open the Recruitment Geek website")
        print(f"   • Enhanced Boolean queries will be pasted automatically")
        print(f"   • Results will be extracted with improved validation")
        print(f"   • Domain-specific targeting provides better accuracy")
        
        proceed = input(f"\n🚀 Start Enhanced Recruitment Geek search? (y/n, default: y): ").strip().lower()
        if proceed == 'n':
            print("Cancelled.")
            return
        
        # Perform enhanced search using the tool
        candidates = scraper.search_with_enhanced_recruitment_geek()
        
        if candidates:
            print(f"\n✅ Enhanced Recruitment Geek search completed! Found {len(candidates)} profiles")
            
            # Save enhanced results
            if scraper.save_enhanced_results(candidates):
                # Show enhanced summary
                domain_counts = {}
                confidence_counts = {}
                for candidate in candidates:
                    domain = candidate.get('linkedin_domain', 'unknown')
                    domain_counts[domain] = domain_counts.get(domain, 0) + 1
                    
                    conf = candidate.get('confidence', 'unknown')
                    confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
                
                print(f"\n🔗 Enhanced LinkedIn profiles extracted:")
                print(f"   • Tool: {scraper.recruitment_geek_url}")
                print(f"   • Enhancement: site:{scraper.target_domain} + location targeting")
                print(f"   • Data quality: Improved validation and filtering")
                
                print(f"\nLinkedIn domain breakdown:")
                for domain, count in sorted(domain_counts.items()):
                    if domain == scraper.target_domain:
                        print(f"  - {domain}: {count} ✅ (Target domain)")
                    else:
                        print(f"  - {domain}: {count}")
                
                print(f"\nConfidence breakdown:")
                for conf, count in sorted(confidence_counts.items()):
                    print(f"  - {conf.title()}: {count}")
                
                # Show sample enhanced results
                print(f"\n👥 Sample enhanced profiles found:")
                for i, candidate in enumerate(candidates[:5], 1):
                    name = f"{candidate.get('first_name', '')} {candidate.get('last_name', '')}"
                    domain = candidate.get('linkedin_domain', 'unknown')
                    conf = candidate.get('confidence', 'unknown')
                    print(f"   {i}. {name} ({domain}, {conf} confidence)")
                
                if len(candidates) > 5:
                    print(f"   ... and {len(candidates) - 5} more")
                
                # Create enhanced Excel report
                print(f"\n📋 Generating enhanced Excel report...")
                if scraper.create_enhanced_excel_report(candidates):
                    print("✅ Enhanced Excel report created successfully")
                    print("   • Results from Enhanced Recruitment Geek tool")
                    print("   • Site-specific domain targeting applied")
                    print("   • Location filtering enhancement included")
                    print("   • Ready for LinkedIn verification")
                    
                    # Offer verification
                    print(f"\n🚀 Ready to launch LinkedIn verification...")
                    launch_verification = input("Launch LinkedIn verification now? (y/n, default: y): ").strip().lower()
                    
                    if launch_verification != 'n':
                        verification_script = script_dir / "script5_linkedin_verification.py"
                        if verification_script.exists():
                            try:
                                print("✅ LinkedIn verification script launched!")
                                if sys.platform == 'win32':
                                    subprocess.Popen([sys.executable, str(verification_script)], 
                                                   creationflags=subprocess.CREATE_NEW_CONSOLE)
                                else:
                                    subprocess.Popen([sys.executable, str(verification_script)])
                                print("📋 The verification process will:")
                                print("   • Visit each LinkedIn profile to extract current job info")
                                print("   • Identify who currently works at your target company") 
                                print("   • Create a final Excel report with verified employees")
                                print("   • Use the enhanced Recruitment Geek results as input")
                            except Exception as e:
                                print(f"⚠️  Could not launch verification script: {e}")
                                print("Please run script5_linkedin_verification.py manually.")
                        else:
                            print("⚠️  Verification script not found")
                            print("Please ensure script5_linkedin_verification.py exists")
                    else:
                        print("LinkedIn verification skipped.")
                        print("Run script5_linkedin_verification.py when ready to verify profiles.")
                else:
                    print("⚠️ Enhanced Excel report creation failed")
            else:
                print("Failed to save enhanced results")
                sys.exit(1)
        else:
            print("❌ No LinkedIn profiles found with enhanced search.")
            print("This could be due to:")
            print("- Recruitment Geek website structure changes")
            print("- Network connectivity issues") 
            print("- Enhanced query complexity blocking")
            print("- Limited search results for the company/location combination")
            print(f"- No profiles available on {scraper.target_domain}")
            if scraper.job_titles:
                print("- Job titles might be too specific - try broader titles")
            else:
                print("- Consider adding specific job titles for more targeted results")
            
            print(f"\n💡 Troubleshooting:")
            print(f"   • Check {scraper.recruitment_geek_url} manually")
            print(f"   • Verify site:{scraper.target_domain} queries work in the tool")
            print(f"   • Try the standard Recruitment Geek script (script2_geek.py)")
            print(f"   • Consider using the enhanced LinkedIn search (script2_web_scraping.py)")
    
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()