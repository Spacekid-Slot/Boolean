#!/usr/bin/env python3
"""
LinkedIn Search via Searx - COMPLETE VERSION

This script replaces the Bing-based LinkedIn search with Searx (priv.au)
using the syntax: "uk.linkedin.com/in/" london company jobtitle

COMPLETE FEATURES:
- Searx API + Browser fallback
- Full interactive workflow
- Excel report generation
- LinkedIn verification integration
- Proper error handling and cleanup
"""

import json
import logging
import os
import random
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import List, Dict, Optional, Set
from urllib.parse import quote_plus
from datetime import datetime
import requests

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def install_dependencies():
    """Install required packages."""
    required = ['selenium', 'webdriver-manager', 'openpyxl', 'requests']
    for package in required:
        try:
            if package == 'webdriver-manager':
                import webdriver_manager
            elif package == 'openpyxl':
                import openpyxl
            elif package == 'requests':
                import requests
            else:
                __import__(package)
        except ImportError:
            print(f"Installing {package}...")
            subprocess.call([sys.executable, "-m", "pip", "install", package])

install_dependencies()

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
    from webdriver_manager.chrome import ChromeDriverManager
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import requests
except ImportError as e:
    print(f"Failed to import required packages: {e}")
    sys.exit(1)

class LinkedInSearxSearcher:
    """LinkedIn searcher using Searx instead of Bing with full interactive workflow."""
    
    def __init__(self, config_path: str):
        self.config = self._load_config(config_path)
        self.script_dir = Path(__file__).parent
        self.driver = None
        self.found_candidates = []
        self.processed_urls = set()
        
        # Searx configuration
        self.primary_searx = "https://priv.au"
        self.backup_searx = ["https://searx.be", "https://search.privacyguides.net"]
        self.working_searx = self._test_searx_instance()
        
        # LinkedIn domain mapping for different countries
        self.linkedin_domains = self._get_linkedin_domains()
        self.target_domain = self._determine_target_domain()
        
    def _load_config(self, config_path: str) -> dict:
        """Load configuration."""
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ Configuration error: {e}")
            sys.exit(1)
    
    def _get_linkedin_domains(self) -> dict:
        """Get LinkedIn domain mappings."""
        return {
            'uk': 'uk.linkedin.com',
            'united kingdom': 'uk.linkedin.com',
            'england': 'uk.linkedin.com',
            'scotland': 'uk.linkedin.com',
            'wales': 'uk.linkedin.com',
            'london': 'uk.linkedin.com',
            'edinburgh': 'uk.linkedin.com',
            'manchester': 'uk.linkedin.com',
            'canada': 'ca.linkedin.com',
            'australia': 'au.linkedin.com',
            'france': 'fr.linkedin.com',
            'germany': 'de.linkedin.com',
            'spain': 'es.linkedin.com',
            'italy': 'it.linkedin.com',
            'netherlands': 'nl.linkedin.com',
        }
    
    def _determine_target_domain(self) -> str:
        """Determine target LinkedIn domain from location."""
        location = self.config.get('location', '').lower()
        
        for location_key, domain in self.linkedin_domains.items():
            if location_key in location:
                logger.info(f"Using LinkedIn domain: {domain}")
                return domain
        
        logger.info("Using default LinkedIn domain: linkedin.com")
        return 'linkedin.com'
    
    def _test_searx_instance(self) -> str:
        """Test Searx instances and return working one."""
        print("🔍 Testing Searx instances...")
        
        # Test primary first
        try:
            response = requests.get(f"{self.primary_searx}/search", 
                                  params={"q": "test", "format": "json"}, 
                                  timeout=5)
            if response.status_code == 200:
                print(f"✅ {self.primary_searx} - Working (PRIMARY)")
                return self.primary_searx
        except Exception as e:
            print(f"❌ {self.primary_searx} - Error: {e}")
        
        # Test backups
        for backup in self.backup_searx:
            try:
                response = requests.get(f"{backup}/search", 
                                      params={"q": "test", "format": "json"}, 
                                      timeout=5)
                if response.status_code == 200:
                    print(f"✅ {backup} - Working (backup)")
                    return backup
            except Exception as e:
                print(f"❌ {backup} - Error: {e}")
        
        print(f"⚠️ No instances responding, using {self.primary_searx} anyway")
        return self.primary_searx
    
    def display_configuration(self):
        """Display current configuration and get user confirmation."""
        print("\n" + "=" * 60)
        print("📋 SEARX LINKEDIN SEARCH CONFIGURATION")
        print("=" * 60)
        print(f"🏢 Company: {self.config.get('company_name', 'Not set')}")
        print(f"📍 Location: {self.config.get('location', 'Not set')}")
        print(f"🔍 Search Engine: {self.working_searx}")
        print(f"🌐 Target LinkedIn Domain: {self.target_domain}")
        
        job_titles = self.config.get('job_titles', [])
        if job_titles:
            print(f"🎯 JOB TITLES TO SEARCH FOR ({len(job_titles)}):")
            for i, title in enumerate(job_titles[:10], 1):
                print(f"   {i}. {title}")
            if len(job_titles) > 10:
                print(f"   ... and {len(job_titles) - 10} more")
        else:
            print("🎯 JOB TITLES: Using general search (no specific titles)")
        
        print(f"📄 Search method: Searx API + Browser fallback")
        print("=" * 60)
        
        while True:
            choice = input("\nProceed with Searx LinkedIn search? (y/n): ").strip().lower()
            if choice in ['y', 'yes', '']:
                return True
            elif choice in ['n', 'no']:
                print("❌ Search cancelled.")
                return False
            else:
                print("Please enter 'y' for yes or 'n' for no.")
    
    def display_search_strategy(self):
        """Display search strategy."""
        company = self.config.get('company_name', '')
        location = self.config.get('location', '')
        job_titles = self.config.get('job_titles', [])
        
        print("\n" + "=" * 60)
        print("🎯 SEARX SEARCH STRATEGY")
        print("=" * 60)
        
        print(f"📝 Search Syntax Examples:")
        print(f"   \"{self.target_domain}/in/\" {location.lower()} {company.lower()} sales")
        print(f"   \"{self.target_domain}/in/\" {company.lower()} director {location.lower()}")
        
        if job_titles:
            print(f"🎯 Job Title Targeting:")
            sample_titles = job_titles[:3]
            for title in sample_titles:
                print(f"   \"{self.target_domain}/in/\" {company.lower()} {title.lower()} {location.lower()}")
            if len(job_titles) > 3:
                print(f"   ... and {len(job_titles) - 3} more variations")
            
            total_queries = len(job_titles) * 3 + 6  # 3 per title + 6 general
            print(f"📊 TOTAL QUERIES: {total_queries}")
            print(f"⏱️ ESTIMATED TIME: {(total_queries * 2) // 60} minutes")
        else:
            print(f"🔍 General company search")
            print(f"📊 TOTAL QUERIES: 6")
            print(f"⏱️ ESTIMATED TIME: 1-2 minutes")
        
        print(f"🚀 Advantage: No blocking, better coverage than Bing")
        print("=" * 60)
        
        while True:
            choice = input("\nStart Searx LinkedIn search? (y/n): ").strip().lower()
            if choice in ['y', 'yes', '']:
                return True
            elif choice in ['n', 'no']:
                print("❌ Search cancelled.")
                return False
            else:
                print("Please enter 'y' for yes or 'n' for no.")
    
    def create_searx_queries(self):
        """Create search queries using exact Searx X-ray syntax."""
        company = self.config.get('company_name', '').lower()
        location = self.config.get('location', '').lower()
        job_titles = self.config.get('job_titles', [])
        
        queries = []
        
        # Domain path for search
        domain_path = f"{self.target_domain}/in/"
        
        if job_titles:
            logger.info(f"Creating targeted searches for {len(job_titles)} job titles...")
            
            # Job title specific searches using your exact syntax
            for title in job_titles:
                title_lower = title.lower()
                title_queries = [
                    f'"{domain_path}" {location} {company} {title_lower}',
                    f'"{domain_path}" {company} {title_lower} {location}',
                    f'"{domain_path}" {company} {title_lower}'  # Without location as backup
                ]
                queries.extend(title_queries)
            
            # Add general executive queries
            exec_queries = [
                f'"{domain_path}" {location} {company} director',
                f'"{domain_path}" {location} {company} manager',
                f'"{domain_path}" {location} {company} sales',
                f'"{domain_path}" {company} director {location}',
                f'"{domain_path}" {company} manager {location}',
                f'"{domain_path}" {location} {company}'  # General fallback
            ]
            queries.extend(exec_queries)
            
            # Add global backup searches (linkedin.com instead of country-specific)
            if self.target_domain != 'linkedin.com':
                global_queries = [
                    f'"linkedin.com/in/" {location} {company} sales',
                    f'"linkedin.com/in/" {company} director {location}',
                    f'"linkedin.com/in/" {location} {company}'
                ]
                queries.extend(global_queries)
            
        else:
            logger.info(f"Creating general company searches...")
            
            # General company searches using your syntax
            queries = [
                f'"{domain_path}" {location} {company}',
                f'"{domain_path}" {company} {location}',
                f'"{domain_path}" {company} sales {location}',
                f'"{domain_path}" {company} director {location}',
                f'"{domain_path}" {company} manager {location}',
                f'"{domain_path}" {company}'  # Broadest search
            ]
            
            # Add global backup
            if self.target_domain != 'linkedin.com':
                queries.extend([
                    f'"linkedin.com/in/" {location} {company}',
                    f'"linkedin.com/in/" {company} {location}'
                ])
        
        logger.info(f"Created {len(queries)} Searx search queries")
        return queries
    
    def search_profiles_via_searx(self):
        """Main search function using Searx API and browser fallback."""
        queries = self.create_searx_queries()
        
        print(f"\n🔍 Starting Searx LinkedIn search...")
        print(f"Using: {self.working_searx}")
        print("=" * 60)
        
        # Try API method first (faster)
        api_profiles = self._search_via_api(queries)
        
        if api_profiles:
            self.found_candidates.extend(api_profiles)
            print(f"✅ API search completed: {len(api_profiles)} profiles")
        else:
            print("⚠️ API search failed, trying browser method...")
            browser_profiles = self._search_via_browser(queries)
            self.found_candidates.extend(browser_profiles)
        
        print(f"\n🎉 Searx search completed! Found {len(self.found_candidates)} total profiles")
        return self.found_candidates
    
    def _search_via_api(self, queries: List[str]) -> List[Dict]:
        """Search using Searx JSON API."""
        profiles = []
        
        for i, query in enumerate(queries, 1):
            print(f"[{i}/{len(queries)}] API Search: {query[:60]}...")
            
            try:
                params = {
                    "q": query,
                    "format": "json",
                    "engines": "bing,google,duckduckgo",
                    "safesearch": "0"
                }
                
                response = requests.get(
                    f"{self.working_searx}/search",
                    params=params,
                    timeout=10,
                    headers={"User-Agent": "Mozilla/5.0 (compatible; LinkedIn-Research)"}
                )
                
                if response.status_code == 200:
                    data = response.json()
                    query_profiles = self._extract_profiles_from_json(data, query)
                    
                    new_profiles = [p for p in query_profiles if p['linkedin_url'] not in self.processed_urls]
                    profiles.extend(new_profiles)
                    
                    for profile in new_profiles:
                        self.processed_urls.add(profile['linkedin_url'])
                    
                    print(f"  ✅ Found {len(query_profiles)} total, {len(new_profiles)} new")
                else:
                    print(f"  ❌ API error: {response.status_code}")
                
                time.sleep(1)  # Small delay between queries
                
            except Exception as e:
                print(f"  ❌ API error: {e}")
                continue
        
        return profiles
    
    def _extract_profiles_from_json(self, data: dict, query: str) -> List[Dict]:
        """Extract LinkedIn profiles from Searx JSON response."""
        profiles = []
        
        try:
            results = data.get("results", [])
            
            for result in results:
                url = result.get("url", "")
                title = result.get("title", "")
                content = result.get("content", "")
                
                # Check if it's a LinkedIn profile
                if re.search(r'linkedin\.com/in/', url, re.IGNORECASE):
                    # Parse name from title
                    name_data = self._parse_name_from_title(title)
                    
                    if name_data:
                        profile = {
                            'first_name': name_data['first_name'],
                            'last_name': name_data['last_name'],
                            'title': self._extract_job_title(title, content),
                            'link': url,  # Use 'link' for compatibility with other scripts
                            'linkedin_url': url,
                            'company_name': self.config.get('company_name', ''),
                            'location': self.config.get('location', ''),
                            'confidence': self._determine_confidence(title, content, query),
                            'source': f'Searx Search ({self.working_searx})',
                            'linkedin_domain': self._extract_domain_from_url(url),
                            'search_query': query,
                            'extraction_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        profiles.append(profile)
        
        except Exception as e:
            logger.error(f"Error parsing Searx JSON results: {e}")
        
        return profiles
    
    def _search_via_browser(self, queries: List[str]) -> List[Dict]:
        """Fallback browser search if API fails."""
        profiles = []
        
        try:
            # Setup browser
            options = Options()
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-images')
            options.add_argument('--disable-extensions')
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
            
            for i, query in enumerate(queries[:8], 1):  # Limit browser searches
                print(f"[{i}/8] Browser Search: {query[:60]}...")
                
                try:
                    search_url = f"{self.working_searx}/search?q={quote_plus(query)}"
                    self.driver.get(search_url)
                    time.sleep(3)
                    
                    # Extract LinkedIn URLs from page
                    page_source = self.driver.page_source
                    linkedin_urls = self._extract_linkedin_urls_from_page(page_source)
                    
                    for url in linkedin_urls:
                        if url not in self.processed_urls:
                            # Try to extract name from URL
                            name_from_url = self._extract_name_from_url(url)
                            
                            profile = {
                                'first_name': name_from_url.get('first_name', 'Unknown'),
                                'last_name': name_from_url.get('last_name', 'Unknown'),
                                'title': 'LinkedIn Profile',
                                'link': url,
                                'linkedin_url': url,
                                'company_name': self.config.get('company_name', ''),
                                'location': self.config.get('location', ''),
                                'confidence': 'medium',
                                'source': f'Searx Browser ({self.working_searx})',
                                'linkedin_domain': self._extract_domain_from_url(url),
                                'search_query': query,
                                'extraction_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                            profiles.append(profile)
                            self.processed_urls.add(url)
                    
                    print(f"  ✅ Found {len(linkedin_urls)} LinkedIn URLs")
                    time.sleep(3)
                    
                except Exception as e:
                    print(f"  ❌ Browser error: {e}")
                    continue
        
        except Exception as e:
            print(f"❌ Browser setup error: {e}")
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
        
        return profiles
    
    def _extract_linkedin_urls_from_page(self, page_source: str) -> List[str]:
        """Extract LinkedIn URLs from page source."""
        pattern = r'https?://(?:www\.)?(?:[a-z]{2}\.)?linkedin\.com/in/[\w\-]+'
        urls = re.findall(pattern, page_source, re.IGNORECASE)
        
        # Remove duplicates and clean URLs
        unique_urls = []
        seen = set()
        for url in urls:
            clean_url = url.split('?')[0].rstrip('/')
            if clean_url not in seen and len(clean_url) > 20:  # Basic validation
                unique_urls.append(clean_url)
                seen.add(clean_url)
        
        return unique_urls
    
    def _extract_name_from_url(self, url: str) -> Dict[str, str]:
        """Extract potential name from LinkedIn URL."""
        try:
            # Extract the profile identifier from URL
            match = re.search(r'/in/([^/?]+)', url)
            if match:
                profile_id = match.group(1)
                
                # Try to parse as name (many LinkedIn URLs use firstname-lastname)
                if '-' in profile_id:
                    parts = profile_id.split('-')
                    if len(parts) >= 2:
                        first_name = parts[0].replace('-', '').title()
                        last_name = parts[1].replace('-', '').title()
                        
                        if self._is_valid_name_part(first_name) and self._is_valid_name_part(last_name):
                            return {'first_name': first_name, 'last_name': last_name}
            
        except Exception:
            pass
        
        return {'first_name': 'Unknown', 'last_name': 'Unknown'}
    
    def _is_valid_name_part(self, name: str) -> bool:
        """Validate if a string could be a name part."""
        if not name or len(name) < 2 or len(name) > 20:
            return False
        if not re.match(r'^[a-zA-Z]+$', name):
            return False
        # Exclude obvious non-names
        excluded = {'linkedin', 'profile', 'www', 'http', 'https', 'com', 'org'}
        return name.lower() not in excluded
    
    def _parse_name_from_title(self, title: str) -> Optional[Dict]:
        """Parse name from LinkedIn title."""
        try:
            # Clean up title
            title = re.sub(r'\s*-\s*LinkedIn.*', '', title, flags=re.IGNORECASE)
            title = title.strip()
            
            # Extract name pattern
            name_match = re.search(r'^([A-Z][a-z]+)\s+([A-Z][a-zA-Z\s]+?)(?:\s*[-|–]|$)', title)
            if name_match:
                first_name = name_match.group(1).strip()
                last_name = name_match.group(2).strip()
                
                if self._is_valid_name(first_name) and self._is_valid_name(last_name):
                    return {'first_name': first_name, 'last_name': last_name}
            
            return None
            
        except Exception:
            return None
    
    def _is_valid_name(self, name: str) -> bool:
        """Validate name part."""
        if not name or len(name) < 2 or len(name) > 25:
            return False
        if not re.match(r"^[a-zA-Z\-']+$", name):
            return False
        
        false_positives = {
            'linkedin', 'profile', 'company', 'limited', 'group',
            'director', 'manager', 'executive', 'president'
        }
        return name.lower() not in false_positives
    
    def _extract_job_title(self, title: str, content: str) -> str:
        """Extract job title from title/content."""
        # Look for title patterns
        patterns = [
            r'[-–]\s*([^|]+?)(?:\s+at\s+|\s+@\s+|\s+\|)',
            r'\|\s*([^|]+?)(?:\s+at\s+|\s+@\s+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, title, re.IGNORECASE)
            if match:
                job_title = match.group(1).strip()
                if 3 < len(job_title) < 100:
                    return job_title
        
        return "LinkedIn Profile"
    
    def _determine_confidence(self, title: str, content: str, query: str) -> str:
        """Determine confidence level."""
        score = 0
        company = self.config.get('company_name', '').lower()
        
        # Company name presence
        if company in title.lower():
            score += 3
        if company in content.lower():
            score += 2
        
        # Job title matching
        job_titles = self.config.get('job_titles', [])
        for job_title in job_titles:
            if job_title.lower() in title.lower() or job_title.lower() in content.lower():
                score += 3
                break
        
        # Domain matching
        if self.target_domain in query:
            score += 2
        
        # Location matching
        location = self.config.get('location', '').lower()
        if location in title.lower() or location in content.lower():
            score += 1
        
        if score >= 6:
            return 'high'
        elif score >= 3:
            return 'medium'
        else:
            return 'low'
    
    def _extract_domain_from_url(self, url: str) -> str:
        """Extract LinkedIn domain from URL."""
        domain_match = re.search(r'((?:[a-z]{2}\.)?linkedin\.com)', url, re.IGNORECASE)
        return domain_match.group(1) if domain_match else 'linkedin.com'
    
    def display_results_summary(self):
        """Display search results summary."""
        if not self.found_candidates:
            print("\n❌ No LinkedIn profiles found!")
            print("\nThis could be due to:")
            print("• Company name spelling or location issues") 
            print("• Limited public LinkedIn profiles")
            print("• Search engine restrictions")
            print("• Job titles too specific")
            return False
        
        print("\n" + "=" * 60)
        print("📊 SEARX SEARCH RESULTS SUMMARY")
        print("=" * 60)
        print(f"✅ Found {len(self.found_candidates)} LinkedIn profiles via Searx")
        
        # Show confidence breakdown
        confidence_counts = {}
        domain_counts = {}
        
        for candidate in self.found_candidates:
            conf = candidate.get('confidence', 'unknown')
            confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
            
            domain = candidate.get('linkedin_domain', 'unknown')
            domain_counts[domain] = domain_counts.get(domain, 0) + 1
        
        print("\n📈 Confidence Levels:")
        for conf, count in confidence_counts.items():
            print(f"   {conf.title()}: {count}")
        
        print("\n🌐 LinkedIn Domains:")
        for domain, count in domain_counts.items():
            print(f"   {domain}: {count}")
        
        # Show sample results
        print(f"\n👥 Sample Profiles Found:")
        for i, candidate in enumerate(self.found_candidates[:5], 1):
            name = f"{candidate.get('first_name', '')} {candidate.get('last_name', '')}"
            title = candidate.get('title', 'Unknown')
            domain = candidate.get('linkedin_domain', 'unknown')
            print(f"   {i}. {name} - {title} ({domain})")
        
        if len(self.found_candidates) > 5:
            print(f"   ... and {len(self.found_candidates) - 5} more")
        
        print("=" * 60)
        
        while True:
            print("\nWhat would you like to do?")
            print("1. Save results and create Excel report")
            print("2. Try different search terms")
            print("3. Cancel and exit")
            
            choice = input("\nYour choice (1/2/3): ").strip()
            
            if choice == '1':
                return True
            elif choice == '2':
                print("💡 To modify search terms, please run script1_input_collection.py")
                return False
            elif choice == '3':
                print("❌ Search cancelled.")
                return False
            else:
                print("Please enter 1, 2, or 3.")
    
    def save_results(self):
        """Save search results to JSON file."""
        try:
            # Save in the same format as other scripts for compatibility
            output_file = 'linkedin_candidates.json'
            with open(output_file, 'w') as f:
                json.dump(self.found_candidates, f, indent=4)
            
            print(f"💾 Results saved to {output_file}")
            return True
        except Exception as e:
            print(f"❌ Error saving results: {e}")
            return False
    
    def create_excel_report(self):
        """Create Excel report with Searx results."""
        try:
            print("📊 Creating Searx Excel report...")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Searx LinkedIn Profiles"
            
            # Headers
            headers = ["First Name", "Last Name", "Job Title", "LinkedIn URL", "LinkedIn Domain", 
                      "Company", "Location", "Confidence", "Source"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, candidate in enumerate(self.found_candidates, 2):
                ws.cell(row=row, column=1, value=candidate.get('first_name', ''))
                ws.cell(row=row, column=2, value=candidate.get('last_name', ''))
                ws.cell(row=row, column=3, value=candidate.get('title', ''))
                
                # LinkedIn URL with hyperlink
                url_cell = ws.cell(row=row, column=4, value=candidate.get('linkedin_url', ''))
                if candidate.get('linkedin_url'):
                    url_cell.hyperlink = candidate['linkedin_url']
                    url_cell.font = Font(color="0000FF", underline="single")
                
                # Domain highlighting
                domain_cell = ws.cell(row=row, column=5, value=candidate.get('linkedin_domain', ''))
                if candidate.get('linkedin_domain') != 'linkedin.com':
                    domain_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                ws.cell(row=row, column=6, value=candidate.get('company_name', ''))
                ws.cell(row=row, column=7, value=candidate.get('location', ''))
                
                # Confidence with color
                conf_cell = ws.cell(row=row, column=8, value=candidate.get('confidence', '').title())
                if conf_cell.value == 'High':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif conf_cell.value == 'Medium':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                ws.cell(row=row, column=9, value=candidate.get('source', ''))
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 60)
            
            # Save file
            company = self.config.get('company_name', 'Company').replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{company}_Searx_LinkedIn_{timestamp}.xlsx"
            wb.save(filename)
            
            print(f"✅ Searx Excel report created: {filename}")
            
            # Try to open
            try:
                if sys.platform == 'win32':
                    os.startfile(filename)
                elif sys.platform == 'darwin':
                    subprocess.call(['open', filename])
                else:
                    subprocess.call(['xdg-open', filename])
                print("📂 File opened automatically")
            except:
                print(f"📂 Please open manually: {filename}")
            
            return True
            
        except Exception as e:
            print(f"❌ Searx Excel creation error: {e}")
            return False
    
    def offer_verification(self):
        """Offer to run LinkedIn verification."""
        if not self.found_candidates:
            return
        
        print("\n" + "=" * 60)
        print("🔗 LINKEDIN VERIFICATION AVAILABLE")
        print("=" * 60)
        print(f"Found {len(self.found_candidates)} LinkedIn profiles via Searx")
        print("\nVerification can:")
        print("✅ Extract current job titles")
        print("✅ Identify current company employees") 
        print("✅ Get accurate employment data")
        print("✅ Create verified employee report")
        print("=" * 60)
        
        while True:
            choice = input("\nLaunch LinkedIn verification now? (y/n): ").strip().lower()
            if choice in ['y', 'yes']:
                verification_script = self.script_dir / "script5_linkedin_verification.py"
                if verification_script.exists():
                    try:
                        print("🚀 Launching LinkedIn verification...")
                        if sys.platform == 'win32':
                            subprocess.Popen([sys.executable, str(verification_script)], 
                                           creationflags=subprocess.CREATE_NEW_CONSOLE)
                        else:
                            subprocess.Popen([sys.executable, str(verification_script)])
                        print("✅ Verification script launched!")
                        break
                    except Exception as e:
                        print(f"❌ Error launching verification: {e}")
                        break
                else:
                    print("❌ Verification script not found")
                    # Try alternative script names
                    alt_scripts = ["paste.py", "linkedin_verification.py"]
                    for alt_script in alt_scripts:
                        alt_path = self.script_dir / alt_script
                        if alt_path.exists():
                            try:
                                print(f"🚀 Found alternative verification script: {alt_script}")
                                if sys.platform == 'win32':
                                    subprocess.Popen([sys.executable, str(alt_path)], 
                                                   creationflags=subprocess.CREATE_NEW_CONSOLE)
                                else:
                                    subprocess.Popen([sys.executable, str(alt_path)])
                                print("✅ Alternative verification script launched!")
                                return
                            except Exception as e:
                                print(f"❌ Error launching {alt_script}: {e}")
                    break
            elif choice in ['n', 'no']:
                print("ℹ️ Verification skipped. You can run it later.")
                print("💡 Available verification scripts:")
                print("   - script5_linkedin_verification.py")
                print("   - paste.py (manual login version)")
                break
            else:
                print("Please enter 'y' for yes or 'n' for no.")
    
    def cleanup(self):
        """Close browser and cleanup."""
        if self.driver:
            try:
                self.driver.quit()
                print("🌐 Browser closed")
            except:
                pass

def main():
    """Main execution with full interactive workflow."""
    try:
        print("=" * 60)
        print("🔍 SEARX LINKEDIN SEARCH - COMPLETE VERSION")
        print("=" * 60)
        print("This script uses Searx (priv.au) instead of Bing for better")
        print("LinkedIn profile discovery without blocking issues.")
        
        # Install dependencies first
        print("\n🔧 Checking dependencies...")
        install_dependencies()
        
        try:
            print("✅ All dependencies loaded successfully")
        except ImportError as e:
            print(f"❌ Import error: {e}")
            input("Press Enter to exit...")
            return
        
        # Load configuration
        config_path = Path(__file__).parent / "company_config.json"
        if not config_path.exists():
            print("❌ Configuration file not found!")
            print("Please run script1_input_collection.py first")
            input("Press Enter to exit...")
            return
        
        # Initialize searcher
        searcher = LinkedInSearxSearcher(str(config_path))
        
        # Step 1: Display and confirm configuration
        if not searcher.display_configuration():
            return
        
        # Step 2: Display and confirm search strategy
        if not searcher.display_search_strategy():
            return
        
        try:
            # Step 3: Search for profiles with progress updates
            searcher.search_profiles_via_searx()
            
            # Step 4: Display results and get confirmation
            if not searcher.display_results_summary():
                return
            
            # Step 5: Save results
            if searcher.save_results():
                print("✅ Results saved successfully!")
                
                # Step 6: Create Excel report
                if searcher.create_excel_report():
                    print("✅ Excel report created successfully!")
                    
                    # Step 7: Offer verification
                    searcher.offer_verification()
                    
                    print(f"\n🎉 SUCCESS! Found {len(searcher.found_candidates)} LinkedIn profiles")
                    print("📊 Excel report created with clickable LinkedIn URLs")
                    print("🔗 LinkedIn verification available")
                    print("\n💡 Next Steps:")
                    print("   1. Review the Excel file")
                    print("   2. Click LinkedIn URLs to verify profiles")
                    print("   3. Run LinkedIn verification for automated checking")
                else:
                    print("⚠️ Excel report creation failed")
            else:
                print("❌ Failed to save results")
                
        finally:
            searcher.cleanup()
    
    except KeyboardInterrupt:
        print("\n\n❌ Search cancelled by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()