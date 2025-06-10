#!/usr/bin/env python3
"""
Enhanced Data Review and Validation Script

This script provides an interactive review process for employee data
with improved filtering, validation, and Excel generation capabilities.
"""

import json
import logging
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Set, Tuple
from dataclasses import dataclass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class EmployeeStats:
    """Statistics about employee data."""
    total: int
    high_confidence: int
    medium_confidence: int
    low_confidence: int
    sources: Dict[str, int]

class EmployeeDataProcessor:
    """Enhanced employee data processor with validation and review capabilities."""
    
    def __init__(self, config_path: str):
        """Initialize the data processor."""
        self.config = self._load_config(config_path)
        self.script_dir = Path(__file__).parent.absolute()
        
        # File paths
        self.verified_data_file = "verified_employee_data.json"
        self.processed_data_file = self.config.get('processed_data_file', 'processed_employee_data.json')
        
        # Load employee data
        self.employees = self._load_all_employee_data()
        
    def _load_config(self, config_path: str) -> dict:
        """Load configuration from JSON file."""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.error(f"Configuration error: {e}")
            sys.exit(1)
    
    def _get_employee_id(self, employee: Dict) -> str:
        """Generate a unique ID for an employee to detect duplicates."""
        first_name = employee.get('first_name', '').lower().strip()
        last_name = employee.get('last_name', '').lower().strip()
        return f"{first_name}_{last_name}"
    
    def _load_all_employee_data(self) -> List[Dict]:
        """Load employee data from all available sources with deduplication."""
        logger.info("Loading employee data from all sources...")
        
        all_employees = []
        existing_ids = set()
        
        # Check all potential data files in priority order
        potential_files = [
            "merged_employees.json",
            "linkedin_employees.json",
            "website_employees.json",
            "script2a_employees.json",
            "temp_employee_data.json",
            self.processed_data_file,
            "url_processed_employees.json"
        ]
        
        for filename in potential_files:
            file_path = self.script_dir / filename
            if file_path.exists():
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    if isinstance(data, list) and data:
                        logger.info(f"Loaded {len(data)} employees from {filename}")
                        
                        # Add unique employees
                        added_count = 0
                        for emp in data:
                            # Validate basic structure
                            if not self._is_valid_employee_record(emp):
                                continue
                                
                            emp_id = self._get_employee_id(emp)
                            if emp_id and emp_id not in existing_ids:
                                # Standardize and clean the record
                                cleaned_emp = self._clean_employee_record(emp)
                                all_employees.append(cleaned_emp)
                                existing_ids.add(emp_id)
                                added_count += 1
                        
                        if added_count > 0:
                            logger.info(f"  - Added {added_count} unique employees from {filename}")
                            
                except Exception as e:
                    logger.warning(f"Error loading {filename}: {e}")
        
        if not all_employees:
            logger.error("No employee data found. Please run data collection scripts first.")
            sys.exit(1)
        
        logger.info(f"Combined data from all sources: {len(all_employees)} unique employees")
        
        # Save processed data for consistency
        self._save_processed_data(all_employees)
        
        # Sort by confidence and name
        return self._sort_employees(all_employees)
    
    def _is_valid_employee_record(self, employee: Dict) -> bool:
        """Validate that an employee record has required fields."""
        required_fields = ['first_name', 'last_name']
        
        for field in required_fields:
            value = employee.get(field, '').strip()
            if not value or len(value) < 2:
                return False
        
        return True
    
    def _clean_employee_record(self, employee: Dict) -> Dict:
        """Clean and standardize an employee record."""
        cleaned = {
            'first_name': employee.get('first_name', '').strip().title(),
            'last_name': employee.get('last_name', '').strip().title(),
            'title': employee.get('title', 'Unknown').strip(),
            'source': employee.get('source', 'Unknown'),
            'confidence': employee.get('confidence', 'low').lower(),
            'location': employee.get('location', self.config.get('location', '')),
            'company_name': employee.get('company_name', self.config.get('company_name', '')),
        }
        
        # Handle different link field names
        link = (employee.get('link') or 
                employee.get('source_link') or 
                employee.get('url', ''))
        cleaned['link'] = link
        
        # Ensure confidence is valid
        if cleaned['confidence'] not in ['high', 'medium', 'low']:
            cleaned['confidence'] = 'low'
        
        return cleaned
    
    def _sort_employees(self, employees: List[Dict]) -> List[Dict]:
        """Sort employees by confidence level and name."""
        confidence_order = {'high': 0, 'medium': 1, 'low': 2}
        
        return sorted(
            employees,
            key=lambda x: (
                confidence_order.get(x.get('confidence', 'low'), 2),
                x.get('last_name', ''),
                x.get('first_name', '')
            )
        )
    
    def _save_processed_data(self, employees: List[Dict]) -> bool:
        """Save processed employee data."""
        try:
            processed_path = self.script_dir / self.processed_data_file
            with open(processed_path, 'w', encoding='utf-8') as f:
                json.dump(employees, f, indent=4, ensure_ascii=False)
            logger.info(f"Processed data saved to {self.processed_data_file}")
            return True
        except Exception as e:
            logger.error(f"Error saving processed data: {e}")
            return False
    
    def get_employee_stats(self) -> EmployeeStats:
        """Get statistics about the employee data."""
        confidence_counts = {'high': 0, 'medium': 0, 'low': 0}
        sources = {}
        
        for emp in self.employees:
            # Count by confidence
            conf = emp.get('confidence', 'low')
            confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
            
            # Count by source
            source = emp.get('source', 'Unknown')
            sources[source] = sources.get(source, 0) + 1
        
        return EmployeeStats(
            total=len(self.employees),
            high_confidence=confidence_counts['high'],
            medium_confidence=confidence_counts['medium'],
            low_confidence=confidence_counts['low'],
            sources=sources
        )
    
    def display_summary(self) -> None:
        """Display a summary of the employee data."""
        stats = self.get_employee_stats()
        
        print("\n" + "=" * 60)
        print(f"DATA REVIEW FOR {self.config['company_name'].upper()}")
        print("=" * 60)
        print(f"Total employees to review: {stats.total}")
        print(f"High confidence: {stats.high_confidence}")
        print(f"Medium confidence: {stats.medium_confidence}")
        print(f"Low confidence: {stats.low_confidence}")
        
        print("\nEmployee sources:")
        for source, count in sorted(stats.sources.items(), key=lambda x: x[1], reverse=True):
            print(f"  - {source}: {count} employees")
    
    def _clear_screen(self):
        """Clear the terminal screen."""
        os.system('cls' if os.name == 'nt' else 'clear')
    
    def _print_employee_details(self, employee: Dict, index: int, total: int) -> None:
        """Print detailed employee information for review."""
        self._clear_screen()
        print("\n" + "=" * 60)
        print(f"Employee {index}/{total} | Confidence: {employee.get('confidence', 'N/A').upper()}")
        print("=" * 60)
        print(f"Name:      {employee.get('first_name', '')} {employee.get('last_name', '')}")
        print(f"Title:     {employee.get('title', 'N/A')}")
        print(f"Source:    {employee.get('source', 'N/A')}")
        print(f"Location:  {employee.get('location', 'N/A')}")
        
        link = employee.get('link', 'N/A')
        print(f"Link:      {link}")
        print("=" * 60)
    
    def review_employees(self, auto_accept_high: bool = True, 
                        review_medium: bool = False) -> List[Dict]:
        """Interactive employee data review process."""
        logger.info("Starting employee review process...")
        
        verified_employees = []
        
        # Separate employees by confidence
        high_confidence = [emp for emp in self.employees if emp.get('confidence') == 'high']
        medium_confidence = [emp for emp in self.employees if emp.get('confidence') == 'medium']
        low_confidence = [emp for emp in self.employees if emp.get('confidence') == 'low']
        
        # Handle high confidence records
        if auto_accept_high and high_confidence:
            verified_employees.extend(high_confidence)
            logger.info(f"Auto-accepted {len(high_confidence)} high confidence records")
        elif high_confidence:
            logger.info("Reviewing high confidence records...")
            verified_employees.extend(
                self._review_employee_group(high_confidence, "high confidence")
            )
        
        # Handle medium confidence records
        if review_medium and medium_confidence:
            logger.info("Reviewing medium confidence records...")
            verified_employees.extend(
                self._review_employee_group(medium_confidence, "medium confidence")
            )
        elif medium_confidence:
            verified_employees.extend(medium_confidence)
            logger.info(f"Auto-accepted {len(medium_confidence)} medium confidence records")
        
        # Handle low confidence records
        if low_confidence:
            logger.info("Reviewing low confidence records...")
            verified_employees.extend(
                self._review_employee_group(low_confidence, "low confidence")
            )
        
        logger.info(f"Review complete. Verified {len(verified_employees)} of {len(self.employees)} employees")
        return verified_employees
    
    def _review_employee_group(self, employees: List[Dict], group_name: str) -> List[Dict]:
        """Review a group of employees interactively."""
        verified = []
        
        if not employees:
            return verified
        
        print(f"\nReviewing {len(employees)} {group_name} records.")
        print("Options for each record:")
        print("  'y' or Enter: Keep record")
        print("  'n': Skip record")
        print("  'q': Keep all remaining records")
        print("  's': Skip all remaining records")
        
        for i, employee in enumerate(employees, 1):
            self._print_employee_details(employee, i, len(employees))
            
            while True:
                choice = input("Keep this record? (y/n/q/s): ").strip().lower()
                
                if choice == 'q':
                    print("Keeping all remaining records.")
                    verified.extend(employees[i-1:])
                    return verified
                elif choice == 's':
                    print("Skipping all remaining records.")
                    return verified
                elif choice == 'n':
                    break  # Skip this record
                elif choice in ['y', '']:
                    verified.append(employee)
                    break
                else:
                    print("Please enter 'y', 'n', 'q', or 's'.")
        
        return verified
    
    def save_verified_data(self, verified_data: List[Dict]) -> bool:
        """Save verified employee data to JSON file."""
        try:
            # Check if other script's verified data exists
            verified_path = self.script_dir / self.verified_data_file
            
            if verified_path.exists():
                logger.info("Existing verified data found, merging...")
                with open(verified_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                
                # Merge with existing data
                existing_ids = {self._get_employee_id(emp) for emp in existing_data}
                combined_data = existing_data.copy()
                
                added_count = 0
                for emp in verified_data:
                    emp_id = self._get_employee_id(emp)
                    if emp_id not in existing_ids:
                        combined_data.append(emp)
                        existing_ids.add(emp_id)
                        added_count += 1
                
                final_data = combined_data
                logger.info(f"Merged {len(existing_data)} + {added_count} = {len(final_data)} total records")
            else:
                final_data = verified_data
                logger.info(f"Saving {len(final_data)} verified records")
            
            # Save the verified data
            with open(verified_path, 'w', encoding='utf-8') as f:
                json.dump(final_data, f, indent=4, ensure_ascii=False)
            
            logger.info(f"Verified data saved to {self.verified_data_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving verified data: {e}")
            return False
    
    def launch_excel_generator(self, employee_data: List[Dict] = None) -> bool:
        """Launch Excel generation with the verified data."""
        try:
            if employee_data:
                # Generate Excel directly
                return self._generate_excel_report(employee_data)
            else:
                # Launch external Excel script
                excel_scripts = [
                    "script4_excel_output.py",
                    "script3b_name_validator.py"
                ]
                
                for script_name in excel_scripts:
                    script_path = self.script_dir / script_name
                    if script_path.exists():
                        logger.info(f"Launching {script_name}")
                        if sys.platform == 'win32':
                            subprocess.Popen(
                                [sys.executable, str(script_path)],
                                creationflags=subprocess.CREATE_NEW_CONSOLE
                            )
                        else:
                            subprocess.Popen([sys.executable, str(script_path)])
                        return True
                
                # If no external script, generate directly
                return self._generate_excel_report(self.employees)
                
        except Exception as e:
            logger.error(f"Error launching Excel generator: {e}")
            return False
    
    def _generate_excel_report(self, employee_data: List[Dict]) -> bool:
        """Generate Excel report directly."""
        try:
            # Install openpyxl if needed
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                logger.info("Installing openpyxl...")
                subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            
            logger.info("Creating Excel report...")
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employee Data"
            
            # Headers
            headers = [
                "First Name", "Last Name", "Job Title", 
                "Source", "Confidence", "Link", "Location"
            ]
            
            # Write headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write employee data
            for row, emp in enumerate(employee_data, 2):
                ws.cell(row=row, column=1, value=emp.get('first_name', ''))
                ws.cell(row=row, column=2, value=emp.get('last_name', ''))
                ws.cell(row=row, column=3, value=emp.get('title', ''))
                ws.cell(row=row, column=4, value=emp.get('source', ''))
                
                # Confidence with color coding
                conf_cell = ws.cell(row=row, column=5, value=emp.get('confidence', '').upper())
                if conf_cell.value == 'HIGH':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif conf_cell.value == 'MEDIUM':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Link with hyperlink
                link = emp.get('link', '')
                link_cell = ws.cell(row=row, column=6, value=link)
                if link:
                    link_cell.hyperlink = link
                    link_cell.font = Font(color="0000FF", underline="single")
                
                ws.cell(row=row, column=7, value=emp.get('location', ''))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Add summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            summary_sheet["A1"] = "Company"
            summary_sheet["B1"] = self.config.get('company_name', '')
            summary_sheet["A2"] = "Location"
            summary_sheet["B2"] = self.config.get('location', '')
            summary_sheet["A3"] = "Total Employees"
            summary_sheet["B3"] = len(employee_data)
            summary_sheet["A4"] = "Generated On"
            summary_sheet["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Save Excel file
            company_name = self.config.get('company_name', 'Company').replace(' ', '_')
            location = self.config.get('location', 'Location').replace(' ', '_')
            output_file = f"{company_name}_{location}_Employees.xlsx"
            output_path = self.script_dir / output_file
            
            wb.save(output_path)
            logger.info(f"Excel report saved to {output_file}")
            
            # Try to open the file
            try:
                if sys.platform == 'win32':
                    os.startfile(str(output_path))
                elif sys.platform == 'darwin':
                    subprocess.call(['open', str(output_path)])
                else:
                    subprocess.call(['xdg-open', str(output_path)])
                logger.info("Excel file opened automatically")
            except Exception:
                logger.info("Please open the Excel file manually")
            
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return False

def main():
    """Main script execution function."""
    try:
        # Get script directory and config path
        script_dir = Path(__file__).parent.absolute()
        config_path = script_dir / "company_config.json"
        
        if not config_path.exists():
            logger.error("Configuration file not found. Please run script1_input_collection.py first.")
            sys.exit(1)
        
        # Display banner
        print("=" * 60)
        print("EMPLOYEE DATA REVIEW AND VALIDATION (ENHANCED)")
        print("=" * 60)
        
        # Initialize processor
        processor = EmployeeDataProcessor(str(config_path))
        
        # Display summary
        processor.display_summary()
        
        # Get user preferences
        print("\nReview Options:")
        skip_review = input("Skip manual review and keep all records? (y/n, default: n): ").strip().lower() == 'y'
        
        if not skip_review:
            review_medium = input("Review medium confidence records? (y/n, default: n): ").strip().lower() == 'y'
        else:
            review_medium = False
        
        # Perform review
        if skip_review:
            logger.info("Skipping review, keeping all records")
            verified_employees = processor.employees
        else:
            verified_employees = processor.review_employees(
                auto_accept_high=True,
                review_medium=review_medium
            )
        
        # Save verified data
        if processor.save_verified_data(verified_employees):
            print(f"\nVerification complete! Kept {len(verified_employees)} employees.")
            
            # Generate Excel report
            print("Generating Excel report...")
            if processor.launch_excel_generator(verified_employees):
                print("Excel report generated successfully!")
            else:
                print("Warning: Excel report generation failed")
        else:
            logger.error("Failed to save verified data")
            sys.exit(1)
        
        print("\nData processing complete.")
        
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(f"An error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()