#!/usr/bin/env python3
"""
Enhanced Name Validation Script

This script identifies and filters out non-names from employee data
using NLP techniques and comprehensive validation rules.
Now includes automatic Script 5 (LinkedIn verification) launching.
"""

import json
import logging
import os
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional
from dataclasses import dataclass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class ValidationResult:
    """Result of name validation."""
    is_valid: bool
    reason: str
    confidence: float

class NameValidator:
    """Enhanced name validator with NLP capabilities and learning."""
    
    def __init__(self, config_path: str):
        """Initialize the name validator."""
        self.config = self._load_config(config_path)
        self.script_dir = Path(__file__).parent.absolute()
        
        # File paths
        self.verified_data_file = "verified_employee_data.json"
        self.validated_output_file = "validated_employees.json"
        
        # Initialize validation resources
        self._ensure_dependencies()
        self.nlp_model = self._load_nlp_model()
        
        # Load name databases
        self.common_first_names = self._load_name_database('first_names.txt')
        self.common_last_names = self._load_name_database('last_names.txt')
        
        # Load custom exceptions
        self.custom_exceptions = self._load_custom_exceptions()
        
        # Known false positives
        self.false_positives = self._get_false_positives()
        
        # Load employee data
        self.employees = self._load_employee_data()
    
    def _load_config(self, config_path: str) -> dict:
        """Load configuration from JSON file."""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.error(f"Configuration error: {e}")
            sys.exit(1)
    
    def _ensure_dependencies(self):
        """Ensure required packages are installed."""
        try:
            import spacy
            # Try to load the model
            try:
                nlp = spacy.load("en_core_web_sm")
                logger.info("spaCy model loaded successfully")
            except OSError:
                logger.info("Installing spaCy English model...")
                subprocess.call([sys.executable, "-m", "spacy", "download", "en_core_web_sm"])
        except ImportError:
            logger.info("Installing spaCy...")
            subprocess.call([sys.executable, "-m", "pip", "install", "spacy"])
            subprocess.call([sys.executable, "-m", "spacy", "download", "en_core_web_sm"])
    
    def _load_nlp_model(self):
        """Load the spaCy NLP model for advanced name validation."""
        try:
            import spacy
            nlp = spacy.load("en_core_web_sm")
            logger.info("NLP model loaded for advanced validation")
            return nlp
        except Exception as e:
            logger.warning(f"Could not load NLP model: {e}")
            logger.info("Falling back to rule-based validation only")
            return None
    
    def _load_name_database(self, filename: str) -> Set[str]:
        """Load or create name database files."""
        file_path = self.script_dir / filename
        
        if not file_path.exists():
            # Create basic name databases
            if filename == 'first_names.txt':
                names = self._get_common_first_names()
            else:
                names = self._get_common_last_names()
            
            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                for name in sorted(names):
                    f.write(name + '\n')
            
            logger.info(f"Created {filename} with {len(names)} names")
            return names
        else:
            # Load existing file
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    names = {line.strip().lower() for line in f if line.strip()}
                logger.info(f"Loaded {len(names)} names from {filename}")
                return names
            except Exception as e:
                logger.error(f"Error loading {filename}: {e}")
                return set()
    
    def _get_common_first_names(self) -> Set[str]:
        """Get a comprehensive list of common first names."""
        return {
            # Male names
            "james", "john", "robert", "michael", "william", "david", "richard", "joseph", 
            "thomas", "christopher", "charles", "daniel", "matthew", "anthony", "mark", 
            "donald", "steven", "paul", "andrew", "joshua", "kenneth", "kevin", "brian",
            "george", "timothy", "ronald", "jason", "edward", "jeffrey", "ryan", "jacob",
            "gary", "nicholas", "eric", "jonathan", "stephen", "larry", "justin", "scott",
            "brandon", "benjamin", "samuel", "gregory", "alexander", "patrick", "frank",
            "raymond", "jack", "dennis", "jerry", "tyler", "aaron", "jose", "henry",
            "adam", "douglas", "nathan", "peter", "zachary", "kyle", "noah", "alan",
            "ethan", "jeremy", "lionel", "angel", "jordan", "wayne", "arthur", "mason",
            "roy", "ralph", "eugene", "louis", "philip", "bobby", "harold", "lawrence",
            
            # Female names
            "mary", "patricia", "jennifer", "linda", "elizabeth", "barbara", "susan",
            "jessica", "sarah", "karen", "nancy", "lisa", "betty", "helen", "sandra",
            "donna", "carol", "ruth", "sharon", "michelle", "laura", "sarah", "kimberly",
            "deborah", "dorothy", "lisa", "nancy", "karen", "betty", "helen", "sandra",
            "donna", "carol", "ruth", "sharon", "michelle", "laura", "emily", "kimberly",
            "deborah", "dorothy", "amy", "angela", "ashley", "brenda", "emma", "olivia",
            "cynthia", "marie", "janet", "catherine", "frances", "christine", "samantha",
            "debra", "rachel", "carolyn", "janet", "virginia", "maria", "heather", "diane",
            "julie", "joyce", "victoria", "kelly", "christina", "joan", "evelyn", "lauren",
            "judith", "megan", "cheryl", "andrea", "hannah", "jacqueline", "martha", "gloria",
            "teresa", "sara", "janice", "marie", "julia", "heather", "katherine", "alice",
            "madison", "julia", "katherine", "alice", "madison", "abigail", "stephanie",
            "carolyn", "janet", "virginia", "maria", "heather", "diane", "julie", "joyce",
            
            # Modern/International names
            "aiden", "liam", "noah", "ethan", "mason", "lucas", "oliver", "jacob", "emma",
            "sophia", "isabella", "ava", "mia", "emily", "abigail", "madison", "charlotte",
            "harper", "sofia", "avery", "ella", "scarlett", "grace", "chloe", "victoria",
            "riley", "aria", "lily", "aubrey", "zoey", "penelope", "lillian", "addison",
            "layla", "natalie", "camila", "hannah", "brooklyn", "zoe", "nora", "leah"
        }
    
    def _get_common_last_names(self) -> Set[str]:
        """Get a comprehensive list of common last names."""
        return {
            "smith", "johnson", "williams", "jones", "brown", "davis", "miller", "wilson",
            "moore", "taylor", "anderson", "thomas", "jackson", "white", "harris", "martin",
            "thompson", "garcia", "martinez", "robinson", "clark", "rodriguez", "lewis",
            "lee", "walker", "hall", "allen", "young", "hernandez", "king", "wright",
            "lopez", "hill", "scott", "green", "adams", "baker", "gonzalez", "nelson",
            "carter", "mitchell", "perez", "roberts", "turner", "phillips", "campbell",
            "parker", "evans", "edwards", "collins", "stewart", "sanchez", "morris",
            "rogers", "reed", "cook", "morgan", "bell", "murphy", "bailey", "rivera",
            "cooper", "richardson", "cox", "howard", "ward", "torres", "peterson", "gray",
            "ramirez", "james", "watson", "brooks", "kelly", "sanders", "price", "bennett",
            "wood", "barnes", "ross", "henderson", "coleman", "jenkins", "perry", "powell",
            "long", "patterson", "hughes", "flores", "washington", "butler", "simmons",
            "foster", "gonzales", "bryant", "alexander", "russell", "griffin", "diaz",
            "hayes", "myers", "ford", "hamilton", "graham", "sullivan", "wallace", "woods",
            "cole", "west", "jordan", "owens", "reynolds", "fisher", "ellis", "harrison",
            "gibson", "mcdonald", "cruz", "marshall", "ortiz", "gomez", "murray", "freeman",
            "wells", "webb", "simpson", "stevens", "tucker", "porter", "hunter", "hicks",
            "crawford", "henry", "boyd", "mason", "morales", "kennedy", "warren", "dixon",
            "ramos", "reyes", "burns", "gordon", "shaw", "holmes", "rice", "robertson",
            "hunt", "black", "daniels", "palmer", "mills", "nichols", "grant", "knight",
            
            # UK-specific surnames
            "clarke", "davies", "evans", "hughes", "roberts", "wright", "walker", "robinson",
            "thompson", "taylor", "white", "anderson", "harris", "clark", "lewis", "scott",
            "jones", "williams", "wilson", "brown", "johnson", "smith", "miller", "davis",
            "thomas", "jackson", "martin", "hall", "allen", "adams", "baker", "carter",
            "mitchell", "phillips", "campbell", "parker", "turner", "edwards", "collins",
            "stewart", "ward", "bell", "murphy", "cooper", "richardson", "bennett", "brooks",
            "kelly", "sanders", "wood", "barnes", "ross", "henderson", "jenkins", "perry",
            "powell", "long", "patterson", "hughes", "russell", "griffin", "hayes", "myers",
            "hamilton", "graham", "sullivan", "wallace", "cole", "west", "reynolds", "fisher",
            "ellis", "harrison", "gibson", "wells", "webb", "simpson", "stevens", "porter"
        }
    
    def _load_custom_exceptions(self) -> Dict[str, Set[str]]:
        """Load custom exceptions for names to always include or exclude."""
        exceptions_file = self.script_dir / "name_exceptions.json"
        
        if not exceptions_file.exists():
            # Create default exceptions file
            default_exceptions = {
                "always_include": [],
                "always_exclude": []
            }
            
            with open(exceptions_file, 'w', encoding='utf-8') as f:
                json.dump(default_exceptions, f, indent=4)
            
            return {"include": set(), "exclude": set()}
        else:
            try:
                with open(exceptions_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                return {
                    "include": set(name.lower().replace(" ", "_") for name in data.get("always_include", [])),
                    "exclude": set(name.lower().replace(" ", "_") for name in data.get("always_exclude", []))
                }
            except Exception as e:
                logger.error(f"Error loading custom exceptions: {e}")
                return {"include": set(), "exclude": set()}
    
    def _get_false_positives(self) -> Set[str]:
        """Get comprehensive list of false positive terms."""
        base_false_positives = {
            # Common business/property terms
            'property', 'building', 'office', 'street', 'avenue', 'road', 'lane',
            'terrace', 'close', 'court', 'house', 'apartment', 'suite', 'unit',
            'business', 'company', 'asset', 'service', 'management', 'development',
            'consulting', 'enterprise', 'investment', 'capital', 'finance', 'center',
            'centre', 'limited', 'incorporated', 'corporation', 'holdings', 'hotel',
            
            # Directional/descriptive terms
            'north', 'south', 'east', 'west', 'upper', 'lower', 'old', 'new',
            'great', 'little', 'saint', 'bridge', 'park', 'place', 'square',
            
            # Generic terms
            'property', 'building', 'office', 'room', 'floor', 'level', 'block',
            'development', 'estate', 'complex', 'apartments', 'flats', 'house',
            
            # Web/tech terms
            'copyright', 'privacy', 'policy', 'terms', 'conditions', 'cookies',
            'website', 'page', 'site', 'link', 'click', 'view', 'read', 'more'
        }
        
        # Add location-specific false positives
        location = self.config.get('location', '').lower()
        if 'edinburgh' in location:
            base_false_positives.update({
                'edinburgh', 'princes', 'royal', 'mile', 'meadows', 'holyrood',
                'grassmarket', 'morningside', 'stockbridge', 'newington',
                'pilrig', 'granton', 'newhaven', 'portobello', 'cramond',
                'balerno', 'corstorphine', 'duddingston', 'restalrig',
                'leith', 'haymarket', 'bruntsfield', 'murrayfield'
            })
        
        return base_false_positives
    
    def _load_employee_data(self) -> List[Dict]:
        """Load employee data for validation."""
        # Try to load from verified data first
        potential_files = [
            self.verified_data_file,
            "script3a_verified_employees.json",
            "merged_employees.json",
            "processed_employee_data.json",
            "linkedin_employees.json",
            "website_employees.json"
        ]
        
        for filename in potential_files:
            file_path = self.script_dir / filename
            if file_path.exists():
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    if isinstance(data, list) and data:
                        logger.info(f"Loaded {len(data)} employees from {filename}")
                        return data
                except Exception as e:
                    logger.warning(f"Error loading {filename}: {e}")
        
        logger.error("No employee data found for validation")
        return []
    
    def _validate_name_structure(self, first_name: str, last_name: str) -> ValidationResult:
        """Validate the basic structure of names."""
        # Length checks
        if len(first_name) < 2 or len(last_name) < 2:
            return ValidationResult(False, "too_short", 0.0)
        
        if len(first_name) > 25 or len(last_name) > 30:
            return ValidationResult(False, "too_long", 0.0)
        
        # Character validation
        if not re.match(r"^[a-zA-Z\-']+$", first_name) or not re.match(r"^[a-zA-Z\-']+$", last_name):
            return ValidationResult(False, "invalid_characters", 0.0)
        
        # Check for numbers
        if any(c.isdigit() for c in first_name + last_name):
            return ValidationResult(False, "contains_numbers", 0.0)
        
        return ValidationResult(True, "structure_valid", 0.7)
    
    def _validate_against_databases(self, first_name: str, last_name: str) -> ValidationResult:
        """Validate names against common name databases."""
        first_lower = first_name.lower()
        last_lower = last_name.lower()
        
        first_in_db = first_lower in self.common_first_names
        last_in_db = last_lower in self.common_last_names
        
        if first_in_db and last_in_db:
            return ValidationResult(True, "both_in_database", 0.9)
        elif first_in_db or last_in_db:
            return ValidationResult(True, "partial_database_match", 0.7)
        else:
            return ValidationResult(True, "not_in_database", 0.3)
    
    def _validate_with_nlp(self, first_name: str, last_name: str) -> ValidationResult:
        """Use NLP model to validate names."""
        if not self.nlp_model:
            return ValidationResult(True, "nlp_unavailable", 0.5)
        
        try:
            full_name = f"{first_name} {last_name}"
            doc = self.nlp_model(full_name)
            
            # Check if recognized as a person
            for ent in doc.ents:
                if ent.label_ == "PERSON":
                    # Check if the entity covers most of the name
                    if len(ent.text) >= len(full_name) * 0.7:
                        return ValidationResult(True, "nlp_person_entity", 0.8)
                elif ent.label_ in ["ORG", "LOC", "GPE"]:
                    # If recognized as organization or location, likely not a person
                    return ValidationResult(False, f"nlp_{ent.label_.lower()}_entity", 0.1)
            
            # If no entities found, use linguistic features
            # Check for proper noun pattern (capitalization)
            if first_name[0].isupper() and last_name[0].isupper():
                return ValidationResult(True, "proper_noun_pattern", 0.6)
            
            return ValidationResult(True, "nlp_neutral", 0.5)
            
        except Exception as e:
            logger.debug(f"NLP validation error: {e}")
            return ValidationResult(True, "nlp_error", 0.5)
    
    def _check_false_positives(self, first_name: str, last_name: str) -> ValidationResult:
        """Check against known false positives."""
        first_lower = first_name.lower()
        last_lower = last_name.lower()
        
        if first_lower in self.false_positives or last_lower in self.false_positives:
            return ValidationResult(False, "known_false_positive", 0.0)
        
        # Check for location patterns
        full_name = f"{first_name} {last_name}"
        location_patterns = [
            r'\b[A-Z][a-z]+ (?:Street|Road|Avenue|Lane|Drive|Place|Square|Terrace|Gardens|Park|Close|Way|Court|Crescent)\b',
            r'\b[A-Z][a-z]+ (?:Row|Mews|Circle|Heights|Hill|Estate|Acres|Bridge|Village|Gate|Cross|Plaza|Yard|Wharf|Quay)\b'
        ]
        
        for pattern in location_patterns:
            if re.search(pattern, full_name, re.IGNORECASE):
                return ValidationResult(False, "location_pattern", 0.0)
        
        return ValidationResult(True, "no_false_positive", 0.8)
    
    def _check_custom_exceptions(self, first_name: str, last_name: str) -> Optional[ValidationResult]:
        """Check custom exceptions for manual overrides."""
        name_key = f"{first_name.lower()}_{last_name.lower()}"
        
        if name_key in self.custom_exceptions["include"]:
            return ValidationResult(True, "custom_include", 1.0)
        
        if name_key in self.custom_exceptions["exclude"]:
            return ValidationResult(False, "custom_exclude", 0.0)
        
        return None
    
    def validate_name(self, first_name: str, last_name: str) -> ValidationResult:
        """Comprehensive name validation using multiple methods."""
        # Check custom exceptions first
        custom_result = self._check_custom_exceptions(first_name, last_name)
        if custom_result:
            return custom_result
        
        # Structure validation
        structure_result = self._validate_name_structure(first_name, last_name)
        if not structure_result.is_valid:
            return structure_result
        
        # False positive check
        fp_result = self._check_false_positives(first_name, last_name)
        if not fp_result.is_valid:
            return fp_result
        
        # Database validation
        db_result = self._validate_against_databases(first_name, last_name)
        
        # NLP validation
        nlp_result = self._validate_with_nlp(first_name, last_name)
        
        # Combine results with weighted scoring
        confidence_score = (
            structure_result.confidence * 0.2 +
            fp_result.confidence * 0.3 +
            db_result.confidence * 0.3 +
            nlp_result.confidence * 0.2
        )
        
        # Determine final validity
        is_valid = confidence_score >= 0.5
        
        # Determine primary reason
        if db_result.reason == "both_in_database":
            reason = "database_match"
        elif nlp_result.reason == "nlp_person_entity":
            reason = "nlp_confirmed"
        elif db_result.reason == "partial_database_match":
            reason = "partial_match"
        else:
            reason = "heuristic_analysis"
        
        return ValidationResult(is_valid, reason, confidence_score)
    
    def add_custom_exception(self, first_name: str, last_name: str, include: bool = True):
        """Add a name to the custom exceptions list."""
        name_key = f"{first_name.lower()}_{last_name.lower()}"
        exceptions_file = self.script_dir / "name_exceptions.json"
        
        try:
            # Load current exceptions
            with open(exceptions_file, 'r', encoding='utf-8') as f:
                exceptions = json.load(f)
            
            # Add to appropriate list
            full_name = f"{first_name} {last_name}"
            if include:
                if full_name not in exceptions["always_include"]:
                    exceptions["always_include"].append(full_name)
                # Remove from exclude list if present
                if full_name in exceptions["always_exclude"]:
                    exceptions["always_exclude"].remove(full_name)
                
                self.custom_exceptions["include"].add(name_key)
                self.custom_exceptions["exclude"].discard(name_key)
            else:
                if full_name not in exceptions["always_exclude"]:
                    exceptions["always_exclude"].append(full_name)
                # Remove from include list if present
                if full_name in exceptions["always_include"]:
                    exceptions["always_include"].remove(full_name)
                
                self.custom_exceptions["exclude"].add(name_key)
                self.custom_exceptions["include"].discard(name_key)
            
            # Save updated exceptions
            with open(exceptions_file, 'w', encoding='utf-8') as f:
                json.dump(exceptions, f, indent=4)
            
        except Exception as e:
            logger.error(f"Error updating custom exceptions: {e}")
    
    def validate_all_employees(self, interactive: bool = True) -> List[Dict]:
        """Validate all employee names with optional interactive review."""
        if not self.employees:
            logger.warning("No employees to validate")
            return []
        
        logger.info(f"Starting validation of {len(self.employees)} employees")
        
        valid_employees = []
        uncertain_employees = []
        invalid_employees = []
        
        # First pass - automatic validation
        for employee in self.employees:
            first_name = employee.get('first_name', '').strip()
            last_name = employee.get('last_name', '').strip()
            
            if not first_name or not last_name:
                continue
            
            result = self.validate_name(first_name, last_name)
            
            if result.is_valid and result.confidence >= 0.7:
                valid_employees.append((employee, result))
            elif result.is_valid and result.confidence >= 0.4:
                uncertain_employees.append((employee, result))
            else:
                invalid_employees.append((employee, result))
        
        logger.info(f"Validation results: {len(valid_employees)} valid, {len(uncertain_employees)} uncertain, {len(invalid_employees)} invalid")
        
        # Start with valid employees
        final_employees = [emp for emp, _ in valid_employees]
        
        # Handle uncertain employees interactively if requested
        if interactive and uncertain_employees:
            logger.info("Reviewing uncertain names...")
            final_employees.extend(
                self._review_uncertain_names(uncertain_employees)
            )
        elif uncertain_employees:
            # If not interactive, include uncertain employees
            final_employees.extend([emp for emp, _ in uncertain_employees])
            logger.info(f"Auto-included {len(uncertain_employees)} uncertain employees")
        
        logger.info(f"Final validation result: {len(final_employees)} employees validated")
        return final_employees
    
    def _review_uncertain_names(self, uncertain_employees: List[Tuple[Dict, ValidationResult]]) -> List[Dict]:
        """Review uncertain names interactively."""
        reviewed_employees = []
        
        print(f"\nReviewing {len(uncertain_employees)} uncertain names...")
        print("Options: 'y' (keep), 'n' (reject), 's' (skip remaining)")
        
        for i, (employee, result) in enumerate(uncertain_employees, 1):
            first_name = employee.get('first_name', '')
            last_name = employee.get('last_name', '')
            title = employee.get('title', 'Unknown')
            
            print(f"\nUncertain Name {i}/{len(uncertain_employees)}:")
            print(f"Name: {first_name} {last_name}")
            print(f"Title: {title}")
            print(f"Source: {employee.get('source', 'Unknown')}")
            print(f"Confidence: {result.confidence:.2f} ({result.reason})")
            
            while True:
                response = input("Keep this name? (y/n/s): ").strip().lower()
                if response == 's':
                    print("Skipping remaining uncertain names")
                    return reviewed_employees
                elif response == 'y':
                    reviewed_employees.append(employee)
                    self.add_custom_exception(first_name, last_name, include=True)
                    break
                elif response == 'n':
                    self.add_custom_exception(first_name, last_name, include=False)
                    break
                else:
                    print("Please enter 'y', 'n', or 's'")
        
        return reviewed_employees
    
    def save_validated_data(self, validated_employees: List[Dict]) -> bool:
        """Save validated employee data."""
        try:
            # Save to validated output file
            output_path = self.script_dir / self.validated_output_file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(validated_employees, f, indent=4, ensure_ascii=False)
            
            # Also update the main verified data file
            verified_path = self.script_dir / self.verified_data_file
            with open(verified_path, 'w', encoding='utf-8') as f:
                json.dump(validated_employees, f, indent=4, ensure_ascii=False)
            
            logger.info(f"Validated data saved to {self.validated_output_file} and {self.verified_data_file}")
            logger.info(f"Saved {len(validated_employees)} validated employees")
            return True
            
        except Exception as e:
            logger.error(f"Error saving validated data: {e}")
            return False
    
    def create_validation_report(self, validated_employees: List[Dict]) -> bool:
        """Create an Excel report with validated employee data."""
        try:
            # Install openpyxl if needed
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                logger.info("Installing openpyxl...")
                subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            
            logger.info("Creating validation report...")
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Validated Employees"
            
            # Headers
            headers = [
                "First Name", "Last Name", "Job Title", 
                "Source", "Confidence", "Link", "Location"
            ]
            
            # Write headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Write employee data
            for row, emp in enumerate(validated_employees, 2):
                ws.cell(row=row, column=1, value=emp.get('first_name', ''))
                ws.cell(row=row, column=2, value=emp.get('last_name', ''))
                ws.cell(row=row, column=3, value=emp.get('title', ''))
                ws.cell(row=row, column=4, value=emp.get('source', ''))
                
                # Confidence with color coding
                conf_cell = ws.cell(row=row, column=5, value=emp.get('confidence', '').upper())
                if conf_cell.value == 'HIGH':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif conf_cell.value == 'MEDIUM':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Link with hyperlink
                link = emp.get('link', '') or emp.get('source_link', '')
                link_cell = ws.cell(row=row, column=6, value=link)
                if link:
                    link_cell.hyperlink = link
                    link_cell.font = Font(color="0000FF", underline="single")
                
                ws.cell(row=row, column=7, value=emp.get('location', ''))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Add validation summary sheet
            summary_sheet = wb.create_sheet(title="Validation Summary")
            summary_sheet["A1"] = "Validation Report"
            summary_sheet["A1"].font = Font(bold=True, size=14)
            
            summary_sheet["A3"] = "Company"
            summary_sheet["B3"] = self.config.get('company_name', '')
            summary_sheet["A4"] = "Location"
            summary_sheet["B4"] = self.config.get('location', '')
            summary_sheet["A5"] = "Total Validated Employees"
            summary_sheet["B5"] = len(validated_employees)
            summary_sheet["A6"] = "Validation Date"
            summary_sheet["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            summary_sheet["A7"] = "Name Validation Applied"
            summary_sheet["B7"] = "Yes"
            
            # Statistics
            sources = {}
            confidence_levels = {}
            linkedin_count = 0
            
            for emp in validated_employees:
                source = emp.get('source', 'Unknown')
                sources[source] = sources.get(source, 0) + 1
                
                conf = emp.get('confidence', 'unknown')
                confidence_levels[conf] = confidence_levels.get(conf, 0) + 1
                
                # Count LinkedIn profiles
                link = emp.get('link', '') or emp.get('source_link', '')
                if link and 'linkedin.com/in/' in link.lower():
                    linkedin_count += 1
            
            # Add source breakdown
            row = 9
            summary_sheet[f"A{row}"] = "Sources:"
            summary_sheet[f"A{row}"].font = Font(bold=True)
            row += 1
            
            for source, count in sorted(sources.items(), key=lambda x: x[1], reverse=True):
                summary_sheet[f"A{row}"] = f"  {source}"
                summary_sheet[f"B{row}"] = count
                row += 1
            
            # Add confidence breakdown
            row += 1
            summary_sheet[f"A{row}"] = "Confidence Levels:"
            summary_sheet[f"A{row}"].font = Font(bold=True)
            row += 1
            
            for conf, count in sorted(confidence_levels.items()):
                summary_sheet[f"A{row}"] = f"  {conf.title()}"
                summary_sheet[f"B{row}"] = count
                row += 1
            
            # Add LinkedIn info
            if linkedin_count > 0:
                row += 1
                summary_sheet[f"A{row}"] = "LinkedIn Profiles Available:"
                summary_sheet[f"A{row}"].font = Font(bold=True)
                row += 1
                summary_sheet[f"A{row}"] = f"  Ready for verification"
                summary_sheet[f"B{row}"] = linkedin_count
            
            # Save Excel file
            company_name = self.config.get('company_name', 'Company').replace(' ', '_')
            location = self.config.get('location', 'Location').replace(' ', '_')
            output_file = f"{company_name}_{location}_Employees_Validated.xlsx"
            output_path = self.script_dir / output_file
            
            wb.save(output_path)
            logger.info(f"Validation report saved to {output_file}")
            
            # Try to open the file
            try:
                if sys.platform == 'win32':
                    os.startfile(str(output_path))
                elif sys.platform == 'darwin':
                    subprocess.call(['open', str(output_path)])
                else:
                    subprocess.call(['xdg-open', str(output_path)])
                logger.info("Excel file opened automatically")
            except Exception:
                logger.info("Please open the Excel file manually")
            
            return True
            
        except Exception as e:
            logger.error(f"Error creating validation report: {e}")
            return False
    
    def launch_next_script(self) -> bool:
        """Launch the next script in the pipeline."""
        next_scripts = [
            "script4_excel_output.py",
            "script5_linkedin_verification.py"  # Added Script 5
        ]
        
        for script_name in next_scripts:
            script_path = self.script_dir / script_name
            if script_path.exists():
                try:
                    logger.info(f"Launching {script_name}")
                    
                    # Special handling for LinkedIn verification
                    if script_name == "script5_linkedin_verification.py":
                        # Check if we have LinkedIn profiles first
                        linkedin_count = 0
                        if hasattr(self, 'employees') and self.employees:
                            for emp in self.employees:
                                link = emp.get('link', '') or emp.get('source_link', '')
                                if link and 'linkedin.com/in/' in link.lower():
                                    linkedin_count += 1
                        
                        if linkedin_count == 0:
                            logger.info("No LinkedIn profiles found, skipping Script 5")
                            continue
                        
                        print(f"\n🔗 LinkedIn profiles available for verification: {linkedin_count}")
                        print(f"Launching {script_name}...")
                    
                    if sys.platform == 'win32':
                        subprocess.Popen(
                            [sys.executable, str(script_path)],
                            creationflags=subprocess.CREATE_NEW_CONSOLE
                        )
                    else:
                        subprocess.Popen([sys.executable, str(script_path)])
                    return True
                except Exception as e:
                    logger.error(f"Error launching {script_name}: {e}")
        
        logger.info("No additional scripts to launch")
        return False

def main():
    """Main script execution function."""
    try:
        # Get script directory and config path
        script_dir = Path(__file__).parent.absolute()
        config_path = script_dir / "company_config.json"
        
        if not config_path.exists():
            logger.error("Configuration file not found. Please run script1_input_collection.py first.")
            sys.exit(1)
        
        # Display banner
        print("=" * 60)
        print("NAME VALIDATION SCRIPT (ENHANCED)")
        print("=" * 60)
        print("This script validates employee names to filter out false positives")
        
        # Initialize validator
        logger.info("Initializing name validator...")
        validator = NameValidator(str(config_path))
        
        if not validator.employees:
            print("No employee data found for validation.")
            print("Please run the data collection scripts first.")
            sys.exit(1)
        
        # Display summary
        print(f"\nCompany: {validator.config['company_name']}")
        print(f"Location: {validator.config['location']}")
        print(f"Employees to validate: {len(validator.employees)}")
        
        # Count LinkedIn profiles
        linkedin_count = 0
        for emp in validator.employees:
            link = emp.get('link', '') or emp.get('source_link', '')
            if link and 'linkedin.com/in/' in link.lower():
                linkedin_count += 1
        
        if linkedin_count > 0:
            print(f"LinkedIn profiles available: {linkedin_count}")
        
        # Determine interaction mode
        interactive = True
        auto_mode = False
        
        # Check command line arguments
        if len(sys.argv) > 1:
            if sys.argv[1].lower() in ['--auto', '-a', 'auto']:
                interactive = False
                auto_mode = True
                print("Running in automatic mode (no user interaction)")
        
        # Check environment variable (launched from other script)
        if os.getenv('LAUNCHED_FROM_SCRIPT') == 'true':
            interactive = False
            auto_mode = True
            print("Running in automatic mode (launched from another script)")
        
        if not auto_mode:
            # Ask user for interaction preference
            user_choice = input("\nRun in interactive mode to review uncertain names? (y/n, default: y): ").strip().lower()
            interactive = user_choice not in ['n', 'no']
        
        # Validate employee names
        print("\nStarting name validation...")
        validated_employees = validator.validate_all_employees(interactive=interactive)
        
        if validated_employees:
            # Save validated data
            if validator.save_validated_data(validated_employees):
                print(f"\nValidation completed successfully!")
                print(f"Validated {len(validated_employees)} of {len(validator.employees)} employees")
                
                # Show validation statistics
                sources = {}
                confidence_levels = {}
                final_linkedin_count = 0
                
                for emp in validated_employees:
                    source = emp.get('source', 'Unknown')
                    sources[source] = sources.get(source, 0) + 1
                    
                    conf = emp.get('confidence', 'unknown')
                    confidence_levels[conf] = confidence_levels.get(conf, 0) + 1
                    
                    link = emp.get('link', '') or emp.get('source_link', '')
                    if link and 'linkedin.com/in/' in link.lower():
                        final_linkedin_count += 1
                
                print("\nValidated employee sources:")
                for source, count in sorted(sources.items(), key=lambda x: x[1], reverse=True):
                    print(f"  - {source}: {count}")
                
                print("\nConfidence levels:")
                for conf, count in sorted(confidence_levels.items()):
                    print(f"  - {conf.title()}: {count}")
                
                if final_linkedin_count > 0:
                    print(f"\n🔗 {final_linkedin_count} LinkedIn profiles ready for verification")
                
                # Generate Excel report
                print("\nGenerating validation report...")
                if validator.create_validation_report(validated_employees):
                    print("Validation report generated successfully!")
                
                # Launch next script
                validator.launch_next_script()
            else:
                print("Failed to save validated data")
                sys.exit(1)
        else:
            print("No employees passed validation")
            sys.exit(1)
        
        print("\nName validation complete.")
        
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        print(f"An error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()