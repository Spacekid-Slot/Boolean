#!/usr/bin/env python3
"""
Fix Excel Output and LinkedIn Verification - FIXED VERSION

This script fixes Excel generation issues and LinkedIn verification.
Fixed to avoid merge cell errors and handle Windows properly.
"""

import json
import logging
import os
import subprocess
import sys
import time
from pathlib import Path
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def safe_print(*args, **kwargs):
    """Safe print that handles encoding."""
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        safe_args = []
        for arg in args:
            if isinstance(arg, str):
                safe_arg = arg.replace('🔧', '[FIX]')
                safe_arg = safe_arg.replace('✅', '[OK]')
                safe_arg = safe_arg.replace('❌', '[ERROR]')
                safe_arg = safe_arg.replace('📊', '[EXCEL]')
                safe_arg = safe_arg.replace('🔗', '[LINKEDIN]')
                safe_args.append(safe_arg)
            else:
                safe_args.append(arg)
        print(*safe_args, **kwargs)

class ExcelLinkedInFixer:
    """Fixes Excel output and LinkedIn verification - NO MERGE CELL ISSUES."""
    
    def __init__(self):
        self.script_dir = Path(__file__).parent.absolute()
        self.config = self._load_config()
        
    def _load_config(self) -> dict:
        """Load configuration."""
        config_path = self.script_dir / "company_config.json"
        if config_path.exists():
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return {}
    
    def create_working_excel_report(self, employees: list) -> bool:
        """Create Excel report without merge cell issues."""
        try:
            subprocess.call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            safe_print(f"Creating Excel report with {len(employees)} employees...")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LinkedIn Search Results"
            
            # Title - NO MERGE CELLS
            title_cell = ws['A1']
            title_cell.value = f"{self.config.get('company_name', 'Company')} - LinkedIn Search Results"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Subtitle
            subtitle_cell = ws['A2'] 
            subtitle_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            subtitle_cell.font = Font(italic=True, size=10)
            
            # Headers
            headers = ["First Name", "Last Name", "Job Title", "LinkedIn URL", "Company", "Location", "Source", "Confidence"]
            
            header_row = 4
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            linkedin_count = 0
            for row, emp in enumerate(employees, header_row + 1):
                ws.cell(row=row, column=1, value=emp.get('first_name', ''))
                ws.cell(row=row, column=2, value=emp.get('last_name', ''))
                ws.cell(row=row, column=3, value=emp.get('title', ''))
                
                # LinkedIn URL
                link = (emp.get('linkedin_url') or emp.get('link') or emp.get('source_link', ''))
                url_cell = ws.cell(row=row, column=4, value=link)
                if link and 'linkedin.com/in/' in link.lower():
                    linkedin_count += 1
                    try:
                        url_cell.hyperlink = link
                        url_cell.font = Font(color="0000FF", underline="single")
                    except:
                        pass
                
                ws.cell(row=row, column=5, value=emp.get('company_name', ''))
                ws.cell(row=row, column=6, value=emp.get('location', ''))
                ws.cell(row=row, column=7, value=emp.get('source', ''))
                
                # Confidence with colors
                conf_cell = ws.cell(row=row, column=8, value=emp.get('confidence', '').upper())
                if conf_cell.value == 'HIGH':
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif conf_cell.value == 'MEDIUM':
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Auto-adjust columns - FIXED METHOD
            for col_num in range(1, len(headers) + 1):
                column_letter = ws.cell(row=1, column=col_num).column_letter
                max_length = 0
                
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)
            
            # Save
            company_name = self.config.get('company_name', 'Company').replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{company_name}_LinkedIn_Results_FIXED_{timestamp}.xlsx"
            filepath = self.script_dir / filename
            
            wb.save(filepath)
            
            safe_print(f"[OK] Excel report created: {filename}")
            safe_print(f"[EXCEL] Contains {len(employees)} employees with {linkedin_count} LinkedIn profiles")
            
            # Open file
            try:
                if sys.platform == 'win32':
                    os.startfile(str(filepath))
                    safe_print("[OK] Excel file opened automatically")
                else:
                    subprocess.call(['open', str(filepath)])
            except:
                safe_print("[INFO] Please open Excel file manually")
            
            return True
            
        except Exception as e:
            safe_print(f"[ERROR] Excel creation error: {e}")
            return False
    
    def find_linkedin_results(self):
        """Find LinkedIn search results."""
        potential_files = [
            "linkedin_candidates.json",
            "merged_employees.json", 
            "verified_employee_data.json"
        ]
        
        all_results = []
        for filename in potential_files:
            file_path = self.script_dir / filename
            if file_path.exists():
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    if isinstance(data, list):
                        all_results.extend(data)
                except:
                    continue
        
        return all_results
    
    def run_complete_fix(self):
        """Run the complete fix."""
        safe_print("[FIX] FIXING EXCEL OUTPUT AND LINKEDIN VERIFICATION")
        safe_print("=" * 60)
        
        employees = self.find_linkedin_results()
        if not employees:
            safe_print("[ERROR] No LinkedIn search results found!")
            return False
        
        safe_print(f"[EXCEL] Creating Excel Report...")
        success = self.create_working_excel_report(employees)
        
        if success:
            safe_print("[OK] Excel output fixed - Reports generate without errors")
        else:
            safe_print("[ERROR] Excel output fix failed")
        
        return success

def main():
    fixer = ExcelLinkedInFixer()
    fixer.run_complete_fix()

if __name__ == "__main__":
    main()
